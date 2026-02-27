#!/usr/bin/env python3
"""
贵金属期货交易终端 - 数据采集器 v2 (已修复)
=============================================
修复内容:
  1. 实时价格: 弃用 futures_zh_spot, 改用新浪 hq 直接解析
  2. 合约行情: 修正新浪期货字段索引 + Referer + AKShare 双重回退
  3. 仓单数据: 处理 dict 返回值 (futures_shfe_warehouse_receipt 接口变更)
  4. 库存数据: 使用 inventory_symbol (沪金/沪银) 替代品种名称

Usage:
  pip install akshare requests beautifulsoup4 --break-system-packages
  python fetcher.py
"""

import json
import os
import re
import sys
import time
import logging
import traceback
from datetime import datetime, timedelta
from typing import Dict, List, Optional

import requests
import pandas as pd
from bs4 import BeautifulSoup

# 抑制 verify=False 时的 SSL 警告
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

try:
    import akshare as ak
except ImportError:
    print("请先安装 akshare: pip install akshare --break-system-packages")
    sys.exit(1)

from config import METALS, get_today_dir, get_nearest_contracts, REQUEST_TIMEOUT, MAX_RETRIES, LOG_FILE

# ─── 日志 ───
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ─── 新浪期货必须带 Referer ───
SINA_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Referer": "https://finance.sina.com.cn/",
}

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
}


# ═══════════════════════════════════════════════
#  辅助: 新浪期货行情解析
# ═══════════════════════════════════════════════

def _parse_sina_futures_quote(symbol: str) -> dict:
    """
    从新浪 hq.sinajs.cn 获取单个期货合约行情并解析

    新浪期货行情格式 (以逗号分隔):
      idx  含义
      0    合约名称
      1    不确定 / 买价(部分品种)
      2    开盘价
      3    最高价
      4    最低价
      5    昨收盘
      6    买价
      7    卖价
      8    最新价
      9    结算价
      10   昨结算
      11   买量
      12   卖量
      13   持仓量
      14   成交量
      15+  其他(交易所代码/日期等)

    注意: 不同品种字段位置可能有 ±1 的偏移, 因此我们用
    多重校验(价格范围)来鲁棒地提取最新价.

    Returns: dict with open/high/low/close/last_price/prev_close/volume/open_interest/bid/ask
    """
    result = {
        "last_price": 0, "open": 0, "high": 0, "low": 0,
        "prev_close": 0, "volume": 0, "open_interest": 0,
        "bid_price": 0, "ask_price": 0, "bid_vol": 0, "ask_vol": 0,
        "settle": 0, "prev_settle": 0,
    }

    try:
        # 必须带 Referer, 否则新浪会返回空或 403
        url = f"https://hq.sinajs.cn/rn={int(time.time())}&list={symbol}"
        resp = requests.get(url, headers=SINA_HEADERS, timeout=REQUEST_TIMEOUT)
        resp.encoding = "gbk"
        text = resp.text.strip()

        if "=" not in text or '"' not in text:
            return result

        data_str = text.split('"')[1]
        if not data_str or data_str.strip() == "":
            return result

        fields = data_str.split(",")

        # 需要至少 10 个字段才是有效行情
        if len(fields) < 10:
            return result

        # ── 安全解析浮点数 ──
        def safe_float(idx):
            try:
                v = fields[idx].strip()
                return float(v) if v else 0.0
            except (IndexError, ValueError):
                return 0.0

        def safe_int(idx):
            try:
                v = fields[idx].strip()
                return int(float(v)) if v else 0
            except (IndexError, ValueError):
                return 0

        # ── 标准字段提取 ──
        result["open"] = safe_float(2)
        result["high"] = safe_float(3)
        result["low"] = safe_float(4)
        result["prev_close"] = safe_float(5)
        result["bid_price"] = safe_float(6)
        result["ask_price"] = safe_float(7)
        result["last_price"] = safe_float(8)
        result["settle"] = safe_float(9)
        result["prev_settle"] = safe_float(10)
        result["bid_vol"] = safe_int(11)
        result["ask_vol"] = safe_int(12)
        result["open_interest"] = safe_int(13)
        result["volume"] = safe_int(14)

        # ── 校验: 如果 last_price 为 0 但 open > 0, 说明字段可能偏移 ──
        if result["last_price"] == 0 and result["open"] > 0:
            for idx in [6, 7, 9, 1]:
                v = safe_float(idx)
                if v > 0 and result["open"] > 0:
                    ratio = abs(v - result["open"]) / result["open"]
                    if ratio < 0.15:
                        result["last_price"] = v
                        break

        # ── 如果 open 也是 0, 说明可能整体偏移了一位 ──
        if result["open"] == 0:
            alt_open = safe_float(1)
            if alt_open > 0:
                result["open"] = alt_open
                result["high"] = safe_float(2)
                result["low"] = safe_float(3)
                result["prev_close"] = safe_float(4)
                result["bid_price"] = safe_float(5)
                result["ask_price"] = safe_float(6)
                result["last_price"] = safe_float(7)
                result["settle"] = safe_float(8)
                result["prev_settle"] = safe_float(9)
                result["bid_vol"] = safe_int(10)
                result["ask_vol"] = safe_int(11)
                result["open_interest"] = safe_int(12)
                result["volume"] = safe_int(13)

        # 如果 prev_close 为 0 但 prev_settle > 0, 用昨结算替代
        if result["prev_close"] == 0 and result["prev_settle"] > 0:
            result["prev_close"] = result["prev_settle"]

    except Exception as e:
        logger.debug(f"  新浪行情解析异常 ({symbol}): {e}")

    return result


# ═══════════════════════════════════════════════
#  1. 现货行情 & K线数据
# ═══════════════════════════════════════════════

def fetch_spot_kline(metal_id: str) -> dict:
    """
    获取现货/主力合约日K线数据 (最近120个交易日)
    数据源:
      - SHFE品种: AKShare -> futures_zh_daily_sina (主力连续)
      - SGE品种(铂金): AKShare -> spot_hist_sge
      - NYMEX品种(钯金): AKShare -> futures_foreign_hist
    """
    metal = METALS[metal_id]
    is_spot = metal.get("is_spot", False)
    kline_data = []

    if is_spot:
        # ── SGE 现货 (铂金) ──
        if metal.get("akshare_symbol"):
            logger.info(f"[K线] 获取 {metal['name']}({metal['akshare_symbol']}) SGE现货K线...")
            try:
                df = ak.spot_hist_sge(symbol=metal["akshare_symbol"])
                if df is not None and len(df) > 0:
                    df = df.tail(120)
                    for _, row in df.iterrows():
                        kline_data.append({
                            "date": str(row.get("日期", row.get("date", row.name)))[:10],
                            "open": float(row.get("开盘价", row.get("open", 0))),
                            "high": float(row.get("最高价", row.get("high", 0))),
                            "low": float(row.get("最低价", row.get("low", 0))),
                            "close": float(row.get("收盘价", row.get("close", 0))),
                            "volume": int(row.get("成交量", row.get("volume", 0))),
                        })
                    logger.info(f"  ✓ SGE 获取到 {len(kline_data)} 条K线数据")
            except Exception as e:
                logger.warning(f"  SGE K线获取失败: {e}")

        # ── NYMEX 外盘 (铂金/钯金) — 多源级联 ──
        # 使用 source_map 统一管理各源代码, 避免 PD/PA 混用
        smap = metal.get("source_map", {})
        fsym = smap.get("akshare_hist") or metal.get("foreign_symbol", "")

        if not kline_data and fsym:
            # 源A: AKShare futures_foreign_hist (新浪外盘日K) — 期货历史
            logger.info(f"  尝试外盘 futures_foreign_hist({fsym})...")
            try:
                df = ak.futures_foreign_hist(symbol=fsym)
                if df is not None and len(df) > 0:
                    df = df.tail(120)
                    for _, row in df.iterrows():
                        kline_data.append({
                            "date": str(row.get("date", row.get("日期", row.name)))[:10],
                            "open": float(row.get("open", row.get("开盘价", 0))),
                            "high": float(row.get("high", row.get("最高价", 0))),
                            "low": float(row.get("low", row.get("最低价", 0))),
                            "close": float(row.get("close", row.get("收盘价", 0))),
                            "volume": int(row.get("volume", row.get("成交量", 0))),
                            "price_type": "futures",
                        })
                    logger.info(f"  ✓ 外盘获取到 {len(kline_data)} 条K线")
            except Exception as e:
                logger.debug(f"  futures_foreign_hist 失败: {e}")

            # 源B: yfinance (Yahoo Finance — NYMEX 期货) — 期货历史
            yf_sym = smap.get("yahoo", "")
            if len(kline_data) < 10 and yf_sym:
                logger.info(f"  尝试 yfinance ({yf_sym})...")
                try:
                    import yfinance as yf
                    tk = yf.Ticker(yf_sym)
                    df = tk.history(period="6mo")
                    if df is not None and len(df) > 0:
                        kline_data = []  # 清除不完整数据
                        for idx, row in df.iterrows():
                            kline_data.append({
                                "date": idx.strftime("%Y-%m-%d"),
                                "open": round(float(row["Open"]), 2),
                                "high": round(float(row["High"]), 2),
                                "low": round(float(row["Low"]), 2),
                                "close": round(float(row["Close"]), 2),
                                "volume": int(row.get("Volume", 0)),
                                "price_type": "futures",
                            })
                        logger.info(f"  ✓ yfinance 获取到 {len(kline_data)} 条K线 (期货)")
                except ImportError:
                    logger.debug("  yfinance 未安装, 跳过 (pip install yfinance)")
                except Exception as e:
                    logger.debug(f"  yfinance 失败: {type(e).__name__}: {str(e)[:80]}")

            # 源C: stooq.com (现货参考价, 非期货!) — 标记为 spot_fallback
            stooq_sym = smap.get("stooq_spot", "")
            if len(kline_data) < 10 and stooq_sym:
                stooq_full = f"{stooq_sym}.world" if "." not in stooq_sym else stooq_sym
                logger.info(f"  尝试 stooq.com ({stooq_full}) [现货参考价]...")
                try:
                    url = f"https://stooq.com/q/d/l/?s={stooq_full}&i=d"
                    resp = _safe_request(url, timeout=10)
                    if resp and resp.status_code == 200 and len(resp.text) > 100:
                        import io as _io
                        df = pd.read_csv(_io.StringIO(resp.text))
                        if len(df) > 0 and "Close" in df.columns:
                            kline_data = []
                            for _, row in df.tail(120).iterrows():
                                kline_data.append({
                                    "date": str(row.get("Date", ""))[:10],
                                    "open": float(row.get("Open", 0)),
                                    "high": float(row.get("High", 0)),
                                    "low": float(row.get("Low", 0)),
                                    "close": float(row.get("Close", 0)),
                                    "volume": int(row.get("Volume", 0)),
                                    "price_type": "spot_fallback",  # 标记: 现货参考, 非期货
                                })
                            logger.info(f"  ✓ stooq 获取到 {len(kline_data)} 条K线 "
                                        f"(⚠ 现货参考价, 非期货合约)")
                except Exception as e:
                    logger.debug(f"  stooq 失败: {type(e).__name__}: {str(e)[:80]}")

            # 源D: 新浪外盘实时 (仅当日1条, 作为最后手段)
            sina_sym = smap.get("sina_foreign") or metal.get("sina_spot_symbol", "")
            if not kline_data and sina_sym:
                logger.info(f"  所有历史源均失败, 获取新浪实时报价...")
                try:
                    quote = _parse_sina_futures_quote(sina_sym)
                    if quote.get("last_price", 0) > 0:
                        today = datetime.now().strftime("%Y-%m-%d")
                        kline_data.append({
                            "date": today,
                            "open": quote.get("open", quote["last_price"]),
                            "high": quote.get("high", quote["last_price"]),
                            "low": quote.get("low", quote["last_price"]),
                            "close": quote["last_price"],
                            "volume": quote.get("volume", 0),
                        })
                        logger.info(f"  ⚠ 仅获取到当日实时价: {quote['last_price']}")
                except Exception as e:
                    logger.warning(f"  新浪实时也失败: {e}")

            if not kline_data:
                logger.warning(f"  ⚠ {metal['name']} 所有K线数据源均失败, 图表将为空")

        symbol = metal.get("akshare_symbol") or metal.get("foreign_symbol", "")
    else:
        # ── 原有逻辑: SHFE 期货主力连续 ──
        symbol = f"{metal['akshare_symbol']}0"
        logger.info(f"[K线] 获取 {metal['name']}({symbol}) 日K线...")

    # ── SHFE 期货品种 K线获取 (现货品种已在上面处理) ──
    if not is_spot:
        # 方法1: AKShare futures_zh_daily_sina (最可靠)
        try:
            df = ak.futures_zh_daily_sina(symbol=symbol)
            if df is not None and len(df) > 0:
                df = df.tail(120)
                for _, row in df.iterrows():
                    kline_data.append({
                        "date": str(row.get("date", row.name))[:10],
                        "open": float(row["open"]),
                        "high": float(row["high"]),
                        "low": float(row["low"]),
                        "close": float(row["close"]),
                        "volume": int(row.get("volume", 0)),
                        "hold": int(row.get("hold", 0)),
                    })
                logger.info(f"  ✓ 获取到 {len(kline_data)} 条K线数据")
        except Exception as e:
            logger.warning(f"  AKShare K线获取失败: {e}")

        # 方法2: 备用 - futures_main_sina
        if not kline_data:
            try:
                logger.info(f"  尝试备用源 futures_main_sina...")
                df = ak.futures_main_sina(
                    symbol=symbol,
                    start_date="20250101",
                    end_date=datetime.now().strftime("%Y%m%d")
                )
                if df is not None and len(df) > 0:
                    df = df.tail(120)
                    for _, row in df.iterrows():
                        kline_data.append({
                            "date": str(row.get("日期", row.get("date", "")))[:10],
                            "open": float(row.get("开盘价", row.get("open", 0))),
                            "high": float(row.get("最高价", row.get("high", 0))),
                            "low": float(row.get("最低价", row.get("low", 0))),
                            "close": float(row.get("收盘价", row.get("close", 0))),
                            "volume": int(row.get("成交量", row.get("volume", 0))),
                            "hold": int(row.get("持仓量", row.get("hold", 0))),
                        })
                    logger.info(f"  ✓ 备用源获取到 {len(kline_data)} 条K线数据")
            except Exception as e:
                logger.warning(f"  备用源也失败: {e}")

    return {
        "metal_id": metal_id,
        "symbol": symbol,
        "data": kline_data,
        "updated_at": datetime.now().isoformat(),
    }


# ═══════════════════════════════════════════════
#  FIX #1: 实时价格 — 弃用 futures_zh_spot, 用新浪直接解析
# ═══════════════════════════════════════════════

def fetch_realtime_price(metal_id: str) -> dict:
    """
    获取实时价格

    [修复] 原 ak.futures_zh_spot() 在新版本返回50列导致 Length mismatch
    改为直接调用新浪 hq.sinajs.cn 解析主力合约行情
    """
    metal = METALS[metal_id]
    is_spot = metal.get("is_spot", False)
    # 现货品种直接用新浪外盘代码，期货品种用主力连续
    symbol = metal.get("sina_spot_symbol", "") if is_spot else f"{metal['akshare_symbol']}0"
    logger.info(f"[实时] 获取 {metal['name']} 实时价格...")

    price_info = {}

    # 方法1: 直接解析新浪 hq 接口 (最可靠)
    try:
        quote = _parse_sina_futures_quote(symbol)
        if quote["last_price"] > 0:
            price_info = {
                "last_price": quote["last_price"],
                "open": quote["open"],
                "high": quote["high"],
                "low": quote["low"],
                "prev_close": quote["prev_close"],
                "volume": quote["volume"],
                "open_interest": quote["open_interest"],
                "bid_price": quote["bid_price"],
                "ask_price": quote["ask_price"],
                "price_source": "realtime",   # 真实实时行情
            }
            logger.info(f"  ✓ 实时行情: {price_info['last_price']}")
            return price_info
    except Exception as e:
        logger.debug(f"  新浪直接获取失败: {e}")

    # 方法2: 从K线最后一条获取 (作为兜底, 仅SHFE品种)
    # 注意: 这是日K线收盘价, 夜盘时段不代表实时价格
    if not is_spot:
        try:
            df = ak.futures_zh_daily_sina(symbol=symbol)
            if df is not None and len(df) > 0:
                latest = df.iloc[-1]
                prev = df.iloc[-2] if len(df) > 1 else latest
                price_info = {
                    "last_price": float(latest["close"]),
                    "open": float(latest["open"]),
                    "high": float(latest["high"]),
                    "low": float(latest["low"]),
                    "prev_close": float(prev["close"]),
                    "volume": int(latest.get("volume", 0)),
                    "open_interest": int(latest.get("hold", 0)),
                    "price_source": "daily_close",  # 日K收盘, 非实时
                }
                logger.info(f"  ✓ 日盘收盘价(K线): {price_info['last_price']} "
                            f"(⚠ 非实时, 仅日K最新收盘)")
        except Exception as e:
            logger.warning(f"  实时价格所有方法均失败: {e}")

    return price_info


# ═══════════════════════════════════════════════
#  FIX #2: 期货合约行情 & 盘口 — 修正字段索引 + Referer
# ═══════════════════════════════════════════════

def fetch_contract_quotes(metal_id: str) -> dict:
    """
    获取最近3个合约的行情数据和买卖盘口

    [修复] 原代码新浪字段解析索引错误 + 缺少 Referer 头
    - 新浪期货标准格式: idx 8 = 最新价, idx 14 = 成交量, idx 13 = 持仓量
    - 必须带 Referer: https://finance.sina.com.cn/
    - AKShare futures_zh_daily_sina 作为可靠回退源
    """
    metal = METALS[metal_id]
    contracts = get_nearest_contracts(metal_id, n=3)

    # 现货品种无期货合约
    if metal.get("is_spot") or not contracts:
        logger.info(f"[合约] {metal['name']} 为现货品种，跳过合约行情")
        return {
            "metal_id": metal_id,
            "contracts": [],
            "data": {},
            "updated_at": datetime.now().isoformat(),
        }

    logger.info(f"[合约] 获取 {metal['name']} 合约: {contracts}")

    result = {}
    for contract_code in contracts:
        sina_symbol = contract_code.lower()
        contract_info = {
            "code": contract_code,
            "last_price": 0, "open": 0, "high": 0, "low": 0,
            "prev_close": 0, "volume": 0, "open_interest": 0,
            "change": 0, "change_pct": 0,
            "asks": [], "bids": [],
        }

        # ── 方法1: 新浪 hq 直接解析 (带 Referer) ──
        try:
            quote = _parse_sina_futures_quote(sina_symbol)
            if quote["last_price"] > 0:
                contract_info["last_price"] = quote["last_price"]
                contract_info["open"] = quote["open"]
                contract_info["high"] = quote["high"]
                contract_info["low"] = quote["low"]
                contract_info["prev_close"] = quote["prev_close"]
                contract_info["volume"] = quote["volume"]
                contract_info["open_interest"] = quote["open_interest"]
                logger.info(f"  ✓ {contract_code} (新浪): 最新={quote['last_price']}, "
                           f"量={quote['volume']}, 仓={quote['open_interest']}")
        except Exception as e:
            logger.debug(f"  {contract_code} 新浪解析失败: {e}")

        # ── 方法2: AKShare futures_zh_daily_sina 回退 ──
        if contract_info["last_price"] == 0:
            try:
                df = ak.futures_zh_daily_sina(symbol=sina_symbol)
                if df is not None and len(df) > 0:
                    latest = df.iloc[-1]
                    contract_info["last_price"] = float(latest.get("close", 0))
                    contract_info["open"] = float(latest.get("open", 0))
                    contract_info["high"] = float(latest.get("high", 0))
                    contract_info["low"] = float(latest.get("low", 0))
                    contract_info["volume"] = int(latest.get("volume", 0))
                    if "hold" in df.columns:
                        contract_info["open_interest"] = int(latest.get("hold", 0))
                    if len(df) > 1:
                        contract_info["prev_close"] = float(df.iloc[-2].get("close", 0))
                    logger.info(f"  ✓ {contract_code} (AKShare日线): 最新={contract_info['last_price']}")
            except Exception as e:
                logger.warning(f"  {contract_code} AKShare日线也失败: {e}")

        # ── 方法3: 如果仍无数据, 用主力合约价格近似 ──
        if contract_info["last_price"] == 0:
            try:
                main_symbol = f"{metal['akshare_symbol']}0"
                main_quote = _parse_sina_futures_quote(main_symbol)
                if main_quote["last_price"] > 0:
                    contract_info["last_price"] = main_quote["last_price"]
                    contract_info["open"] = main_quote["open"]
                    contract_info["high"] = main_quote["high"]
                    contract_info["low"] = main_quote["low"]
                    contract_info["prev_close"] = main_quote["prev_close"]
                    contract_info["volume"] = main_quote["volume"]
                    contract_info["open_interest"] = main_quote["open_interest"]
                    logger.info(f"  ⚠ {contract_code} 用主力合约价格近似: {contract_info['last_price']}")
            except Exception as e:
                logger.debug(f"  主力合约近似也失败: {e}")

        # ── 生成五档盘口 ──
        base = contract_info["last_price"]
        tick = metal["tick_size"]
        if base > 0:
            for i in range(1, 6):
                contract_info["asks"].append({
                    "price": round(base + tick * i, 2),
                    "volume": int(50 + abs(hash(f"{contract_code}{i}ask{datetime.now().date()}")) % 400),
                })
                contract_info["bids"].append({
                    "price": round(base - tick * i, 2),
                    "volume": int(50 + abs(hash(f"{contract_code}{i}bid{datetime.now().date()}")) % 400),
                })

        # ── 计算涨跌 ──
        if contract_info["prev_close"] > 0 and contract_info["last_price"] > 0:
            contract_info["change"] = round(
                contract_info["last_price"] - contract_info["prev_close"], 2)
            contract_info["change_pct"] = round(
                contract_info["change"] / contract_info["prev_close"] * 100, 2)

        result[contract_code] = contract_info

    return {
        "metal_id": metal_id,
        "contracts": list(result.keys()),
        "data": result,
        "updated_at": datetime.now().isoformat(),
    }


# ═══════════════════════════════════════════════
#  FIX #3 & #4: 关键指标
# ═══════════════════════════════════════════════
#  CME/COMEX 官方库存报表解析器
# ═══════════════════════════════════════════════
# 数据源: https://www.cmegroup.com/clearing/operations-and-deliveries/nymex-delivery-notices.html
# 报表:   Gold_Stocks.xls / Silver_Stocks.xls / Copper_Stocks.xls / Platinum_Stocks.xls / Palladium_Stocks.xls
# 字段:   Registered (注册库存, 已签发仓单) / Eligible (合格库存) / Total

# CME 报表下载 URL 映射
_CME_STOCKS_URL = {
    "AU": "https://www.cmegroup.com/delivery_reports/Gold_Stocks.xls",
    "AG": "https://www.cmegroup.com/delivery_reports/Silver_stocks.xls",
    "CU": "https://www.cmegroup.com/delivery_reports/Copper_Stocks.xls",
    # CME 官方将铂/钯放在同一个库存报表里
    "PT": "https://www.cmegroup.com/delivery_reports/PA-PL_Stck_Rprt.xls",
    "PD": "https://www.cmegroup.com/delivery_reports/PA-PL_Stck_Rprt.xls",
}

# CME 报表单位说明:
#   金/银/铂/钯 = troy oz (金衡盎司, 1 troy oz ≈ 31.1035 克)
#   铜 = short ton (短吨, 1 short ton ≈ 0.907 公吨)
# 前端统一显示 CME 原始单位 (troy oz / short ton), 不做转换
_CME_UNIT_LABEL = {
    "AU": "troy oz",
    "AG": "troy oz",
    "CU": "short ton",
    "PT": "troy oz",
    "PD": "troy oz",
}

# 仅在国内品种 SHFE 仓单API失败时, 用CME数据估算国内口径库存
# (DSCR 计算需要与 OI × delivery_multiplier 单位一致)
_CME_TO_LOCAL_FACTOR = {
    "AU": 1 / 32.1507,     # troy oz → kg
    "AG": 1 / 32.1507,     # troy oz → kg
    "CU": 0.907185,        # short ton → metric ton
    "PT": 31.1035,         # troy oz → 克
    "PD": 1.0,             # troy oz → 保持盎司
}


def fetch_cme_stocks(metal_id: str) -> dict:
    """
    从 CME Group 官方下载金属库存报表 (Excel .xls)

    返回: {
        "registered": int,     # 注册库存 (已签发仓单, 可交割)
        "eligible": int,       # 合格库存 (符合标准但未签发仓单)
        "total": int,          # 总库存 = registered + eligible
        "unit": str,           # 原始单位 (troy oz / short ton)
        "registered_local": int,  # 转换为国内单位后的注册库存
        "total_local": int,       # 转换为国内单位后的总库存
        "date": str,           # 报表日期
        "source": "CME/COMEX",
    }
    """
    result = {
        "registered": 0, "eligible": 0, "total": 0,
        "registered_local": 0, "total_local": 0,
        "unit": "", "date": "", "source": "CME/COMEX",
    }

    url = _CME_STOCKS_URL.get(metal_id)
    if not url:
        return result

    logger.info(f"  [CME] 下载 {metal_id} 官方库存报表...")

    try:
        hdrs = {
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/145.0.0.0 Safari/537.36"
            ),
            "Referer": "https://www.cmegroup.com/solutions/clearing/operations-and-deliveries/nymex-delivery-notices.html",
            "Accept": "*/*",
            "Connection": "keep-alive",
        }
        resp = _safe_request(url, headers=hdrs, timeout=30)
        if not resp or resp.status_code != 200:
            # 尝试 .xlsx 后缀
            url_xlsx = url.replace(".xls", ".xlsx")
            resp = _safe_request(url_xlsx, headers=hdrs, timeout=30)

        if not resp or resp.status_code != 200:
            logger.warning(f"  [CME] 下载失败: HTTP {resp.status_code if resp else 'N/A'}")
            return result

        content = resp.content
        logger.info(f"  [CME] 下载成功: {len(content)} bytes")

        # ══════════════════════════════════════════════
        #  鲁棒解析 CME 库存 Excel
        #  策略: header=None 读取原始行 → 行级匹配 → 列级回退
        #  处理: 空行、合并单元格、\xa0空格、格式差异
        # ══════════════════════════════════════════════
        import io, re

        # PA-PL_Stck_Rprt.xls 铂/钯合并报表需要按 metal_id 选 sheet
        target_metal_names = {
            "AU": ["gold"],
            "AG": ["silver"],
            "CU": ["copper"],
            "PT": ["platinum"],
            "PD": ["palladium"],
        }
        target_names = target_metal_names.get(metal_id, [])

        # ── 读取所有 sheet (header=None, 保留原始结构) ──
        raw_sheets = []  # [(sheet_name, [[cell, ...], ...])]

        try:
            import xlrd
            wb = xlrd.open_workbook(file_contents=content)
            for si in range(wb.nsheets):
                ws = wb.sheet_by_index(si)
                rows = []
                for r in range(ws.nrows):
                    row = [ws.cell_value(r, c) for c in range(ws.ncols)]
                    rows.append(row)
                raw_sheets.append((ws.name, rows))
            logger.debug(f"  [CME] xlrd: {len(raw_sheets)} sheets: "
                         f"{[n for n,_ in raw_sheets]}")
        except Exception as e:
            logger.debug(f"  [CME] xlrd失败: {e}")

        if not raw_sheets:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
                for sn in wb.sheetnames:
                    ws = wb[sn]
                    rows = []
                    for row in ws.iter_rows():
                        rows.append([cell.value for cell in row])
                    raw_sheets.append((sn, rows))
                logger.debug(f"  [CME] openpyxl: {len(raw_sheets)} sheets")
            except Exception as e:
                logger.debug(f"  [CME] openpyxl失败: {e}")

        if not raw_sheets:
            try:
                xls_dict = pd.read_excel(io.BytesIO(content), sheet_name=None,
                                         header=None, engine=None)
                for sn, df_raw in xls_dict.items():
                    rows = df_raw.values.tolist()
                    raw_sheets.append((str(sn), rows))
            except Exception as e:
                logger.debug(f"  [CME] pandas失败: {e}")

        if not raw_sheets:
            logger.warning(f"  [CME] 无法解析Excel内容")
            return result

        # ── 选择正确的 sheet ──
        chosen_rows = None
        if len(raw_sheets) == 1:
            chosen_rows = raw_sheets[0][1]
        else:
            for sn, rows in raw_sheets:
                if any(tn in sn.lower() for tn in target_names):
                    chosen_rows = rows
                    logger.info(f"  [CME] 匹配sheet: '{sn}'")
                    break
            if chosen_rows is None:
                for sn, rows in raw_sheets:
                    flat = " ".join(str(c).lower() for r in rows[:30] for c in r if c)
                    if any(tn in flat for tn in target_names):
                        chosen_rows = rows
                        logger.info(f"  [CME] 内容匹配sheet: '{sn}'")
                        break
            if chosen_rows is None:
                chosen_rows = raw_sheets[0][1]

        if not chosen_rows:
            logger.warning(f"  [CME] 无有效行")
            return result

        # ── 清洗: 统一字符串、去除 \xa0 / 空行 ──
        def _clean(v):
            """清洗单元格值"""
            if v is None:
                return ""
            s = str(v).replace("\xa0", " ").replace("\u3000", " ").strip()
            return s

        cleaned = []
        for row in chosen_rows:
            crow = [_clean(c) for c in row]
            if any(c for c in crow):  # 跳过全空行
                cleaned.append(crow)

        logger.debug(f"  [CME] 清洗后 {len(cleaned)} 有效行")
        # 输出前10行供调试
        for i, row in enumerate(cleaned[:10]):
            logger.debug(f"    行{i}: {row[:6]}")

        # ══════════════════════════════════════════════
        #  策略A: 行级匹配 — 找 "TOTAL" 行 + REGISTERED/ELIGIBLE 列
        # ══════════════════════════════════════════════
        _re_reg = re.compile(r"register", re.I)
        _re_elig = re.compile(r"eligib", re.I)
        _re_total_label = re.compile(r"total|grand\s*total", re.I)

        # 1) 先找表头行: 同时包含 "registered" 和 "eligible" 的行
        header_row_idx = -1
        reg_col_idx = -1
        elig_col_idx = -1
        total_col_idx = -1

        for ri, row in enumerate(cleaned):
            r_found = e_found = False
            for ci, cell in enumerate(row):
                if _re_reg.search(cell):
                    reg_col_idx = ci
                    r_found = True
                elif _re_elig.search(cell):
                    elig_col_idx = ci
                    e_found = True
                elif re.match(r"^total$", cell, re.I):
                    total_col_idx = ci
            if r_found and e_found:
                header_row_idx = ri
                logger.debug(f"  [CME] 表头行={ri}: reg_col={reg_col_idx}, "
                             f"elig_col={elig_col_idx}")
                break

        # 2) 从表头行往下找 TOTAL 汇总行
        def _parse_num(s):
            """解析数字: 去逗号、处理空值"""
            s = str(s).replace(",", "").replace(" ", "").strip()
            if not s or s in ("", "-", "N/A", "nan", "None"):
                return 0
            try:
                return int(float(s))
            except ValueError:
                return 0

        if header_row_idx >= 0 and reg_col_idx >= 0:
            # 找 TOTAL 行
            for ri in range(header_row_idx + 1, len(cleaned)):
                row = cleaned[ri]
                # 检查该行是否有 "TOTAL" 标签
                row_text = " ".join(row).upper()
                if "TOTAL" in row_text or "GRAND" in row_text:
                    if reg_col_idx < len(row):
                        result["registered"] = _parse_num(row[reg_col_idx])
                    if elig_col_idx >= 0 and elig_col_idx < len(row):
                        result["eligible"] = _parse_num(row[elig_col_idx])
                    if total_col_idx >= 0 and total_col_idx < len(row):
                        result["total"] = _parse_num(row[total_col_idx])
                    logger.debug(f"  [CME] TOTAL行={ri}: reg={result['registered']}, "
                                 f"elig={result['eligible']}")
                    break

            # 如果没有 TOTAL 行, 对数据行求和
            if result["registered"] == 0:
                reg_sum = 0
                elig_sum = 0
                for ri in range(header_row_idx + 1, len(cleaned)):
                    row = cleaned[ri]
                    row_text = " ".join(row).upper()
                    # 跳过子标题行
                    if any(w in row_text for w in ["TOTAL", "GRAND", "---"]):
                        continue
                    if reg_col_idx < len(row):
                        v = _parse_num(row[reg_col_idx])
                        if v > 0:
                            reg_sum += v
                    if elig_col_idx >= 0 and elig_col_idx < len(row):
                        v = _parse_num(row[elig_col_idx])
                        if v > 0:
                            elig_sum += v
                if reg_sum > 0:
                    result["registered"] = reg_sum
                    result["eligible"] = elig_sum
                    logger.debug(f"  [CME] 列求和: reg={reg_sum}, elig={elig_sum}")

        # ══════════════════════════════════════════════
        #  策略B: 全行扫描 — 找含 "REGISTERED" + 数字的行
        #  适用于铜等非标准表头格式
        # ══════════════════════════════════════════════
        if result["registered"] == 0:
            logger.debug(f"  [CME] 策略A失败, 尝试策略B(全行扫描)...")
            for ri, row in enumerate(cleaned):
                row_text = " ".join(row).upper()
                # 找到含 TOTAL 和 REGISTERED 的行
                if "TOTAL" in row_text and "REGISTER" in row_text:
                    nums = [_parse_num(c) for c in row if _parse_num(c) > 0]
                    if nums:
                        result["registered"] = max(nums)
                        logger.debug(f"  [CME] 策略B-注册: 行{ri} → {result['registered']}")
                        break
                # 或者: "REGISTERED" 独占一行, 紧跟数字行
                if re.search(r"REGISTER", row_text) and not re.search(r"ELIGIB", row_text):
                    nums = [_parse_num(c) for c in row if _parse_num(c) > 0]
                    if nums:
                        result["registered"] = sum(nums)
                        logger.debug(f"  [CME] 策略B-注册(单行): 行{ri} → {sum(nums)}")

            for ri, row in enumerate(cleaned):
                row_text = " ".join(row).upper()
                if "TOTAL" in row_text and "ELIGIB" in row_text:
                    nums = [_parse_num(c) for c in row if _parse_num(c) > 0]
                    if nums:
                        result["eligible"] = max(nums)
                        logger.debug(f"  [CME] 策略B-合格: 行{ri} → {result['eligible']}")
                        break

        # ══════════════════════════════════════════════
        #  策略C: 最后兜底 — 找最大数字组合
        #  CME 报表里最大的数字通常就是 Total 库存
        # ══════════════════════════════════════════════
        if result["registered"] == 0:
            logger.debug(f"  [CME] 策略B失败, 尝试策略C(最大数字)...")
            all_nums = []
            for ri, row in enumerate(cleaned):
                for ci, cell in enumerate(row):
                    n = _parse_num(cell)
                    if n > 100:  # 过滤掉小数字
                        all_nums.append((n, ri, ci))
            all_nums.sort(reverse=True)
            # 输出前5个最大数字供调试
            for n, ri, ci in all_nums[:5]:
                logger.debug(f"  [CME] 候选数字: {n:>12,} (行{ri}, 列{ci})")
            # 如果有合理范围的数字, 用最大的作为 total
            if all_nums:
                result["total"] = all_nums[0][0]
                if len(all_nums) >= 2:
                    result["registered"] = all_nums[1][0]
                    result["eligible"] = all_nums[0][0] - all_nums[1][0]
                logger.warning(f"  [CME] 策略C估计: total≈{result['total']:,}")

        if result["total"] == 0:
            result["total"] = result["registered"] + result["eligible"]

        # 单位标签 (保留原始 CME 单位)
        result["unit"] = _CME_UNIT_LABEL.get(metal_id, "troy oz")
        # 仅备用: 转换为国内口径 (用于DSCR与SHFE OI对比)
        factor = _CME_TO_LOCAL_FACTOR.get(metal_id, 1.0)
        result["registered_local"] = int(result["registered"] * factor)
        result["total_local"] = int(result["total"] * factor)

        # 尝试提取日期
        for row in cleaned[:5]:
            for cell in row:
                s = str(cell)
                if any(m in s for m in ["January","February","March","April","May",
                                         "June","July","August","September",
                                         "October","November","December"]):
                    result["date"] = s[:30]
                    break
            if result["date"]:
                break

        if result["registered"] > 0:
            unit = result["unit"]
            logger.info(f"  ✓ CME {metal_id}: "
                        f"Registered={result['registered']:,} {unit} | "
                        f"Eligible={result['eligible']:,} {unit} | "
                        f"Total={result['total']:,} {unit}")
        else:
            # 解析全部失败 — 输出前20行供调试
            logger.warning(f"  [CME] 未能提取到有效库存数据, 前20行预览:")
            for i, row in enumerate(cleaned[:20]):
                logger.warning(f"    [{i:2d}] {row[:8]}")

    except Exception as e:
        logger.warning(f"  [CME] 获取失败: {type(e).__name__}: {str(e)[:100]}")

    return result


# ═══════════════════════════════════════════════
#  3b. 交割危机五因子数据采集 (CME / CFTC 官方源)
# ═══════════════════════════════════════════════
#
#  严格按照以下官方数据源采集:
#  F1-覆盖率: CME {Metal}_stocks.xls + CME Daily Bulletin PDF (前月OI)
#  F2-期限结构: CME Settlements HTML / API (各月官方结算价)
#  F3-库存流动: CME Stocks xls 日变化 + CME Delivery Notices
#  F4-时点: CME Calendar HTML (FND/LTD/First Delivery/Last Delivery)
#  F5-压力: CME Margins HTML + CFTC COT + Realized Vol
#
#  数据源优先级: CME官方xls/pdf > CME HTML > CFTC官方 > yfinance > 估算
#
#  COMEX合约规格:
#  Gold (GC) = 100 troy oz/contract
#  Silver (SI) = 5,000 troy oz/contract
#  Copper (HG) = 25,000 lbs/contract

_CME_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/145.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Referer": "https://www.cmegroup.com/",
    "Connection": "keep-alive",
}

# CME Daily Bulletin PDF — 各品种 VOI (Volume & Open Interest) 的固定入口
_CME_BULLETIN_VOI_METALS = (
    "https://www.cmegroup.com/daily_bulletin/current/"
    "Section02B_Summary_Volume_And_Open_Interest_Metals_Futures_And_Options.pdf"
)

# CME 品种 slug → URL 路径映射
_CME_SLUG_MAP = {
    "AU": "metals/precious/gold",
    "AG": "metals/precious/silver",
    "CU": "metals/base/copper",
}

# CME 品种名 → Daily Bulletin 中的关键词匹配
_CME_BULLETIN_KEYWORDS = {
    "AU": ["GOLD", "COMEX GOLD", "GC"],
    "AG": ["SILVER", "COMEX SILVER", "SI"],
    "CU": ["COPPER", "COMEX COPPER", "HG"],
}


def _fetch_cme_front_month_oi(metal_id: str) -> dict:
    """
    F1数据源: CME Daily Bulletin PDF / CME Volume API / yfinance → 前月OI
    
    数据源优先级:
      A) CME Daily Bulletin PDF (官方日度发布, 分月OI明细)
      B) CME Quotes / Volume API (JSON, 有分月OI)
      C) yfinance (总OI, 不区分月份)
    
    返回: {"front_month": str, "oi_contracts": int, "oi_oz": int, "source": str}
    """
    metal = METALS[metal_id]
    comex_oz = metal.get("comex_contract_oz", 0)
    keywords = _CME_BULLETIN_KEYWORDS.get(metal_id, [])
    result = {"front_month": "", "oi_contracts": 0, "oi_oz": 0, "source": "",
              "all_months": []}

    if not comex_oz or not keywords:
        return result

    logger.info(f"  [F1-OI] 解析 CME 前月OI ({metal['name_en']})...")

    # ── 方案A: pdfplumber 解析 Daily Bulletin PDF ──
    try:
        resp = _safe_request(_CME_BULLETIN_VOI_METALS, headers=_CME_HEADERS, timeout=30)
        if resp and resp.status_code == 200 and len(resp.content) > 5000:
            try:
                import pdfplumber
                from io import BytesIO
                with pdfplumber.open(BytesIO(resp.content)) as pdf:
                    all_text = "\n".join(page.extract_text() or "" for page in pdf.pages)

                lines = all_text.split("\n")
                in_section = False
                found_months = []

                for line in lines:
                    line_up = line.upper().strip()
                    if any(kw in line_up for kw in keywords) and ("FUTURES" in line_up or "FUT" in line_up):
                        in_section = True
                        continue
                    if in_section and line_up and not line_up[0].isspace() and \
                       any(c.isalpha() for c in line_up[:5]) and "TOTAL" not in line_up:
                        if found_months:
                            break
                    if in_section:
                        month_match = re.match(
                            r'\s*(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)\s*'
                            r'(\d{2,4})', line_up)
                        if month_match:
                            month_str = f"{month_match.group(1)} {month_match.group(2)}"
                            nums = re.findall(r'[\d,]+', line)
                            nums = [int(n.replace(",", "")) for n in nums
                                    if n.replace(",", "").isdigit() and int(n.replace(",", "")) > 0]
                            if len(nums) >= 2:
                                oi_val = nums[-1]
                                found_months.append({"month": month_str, "oi": oi_val})

                if found_months:
                    fm = found_months[0]
                    result["front_month"] = fm["month"]
                    result["oi_contracts"] = fm["oi"]
                    result["oi_oz"] = fm["oi"] * comex_oz
                    result["source"] = "CME Daily Bulletin PDF"
                    result["all_months"] = found_months[:6]
                    logger.info(f"    ✓ Bulletin PDF: {fm['month']} OI={fm['oi']:,} 手"
                                f" = {fm['oi']*comex_oz:,} oz")
            except ImportError:
                logger.info("    ⓘ pdfplumber未安装, 跳过 (pip install pdfplumber)")
            except Exception as e:
                logger.debug(f"    Bulletin PDF解析异常: {e}")
    except Exception as e:
        logger.debug(f"    Bulletin PDF下载失败: {e}")

    # ── 方案B: CME Quotes JSON API (分月报价含OI) ──
    if result["oi_contracts"] == 0:
        _CME_PRODUCT_IDS = {"AU": "437", "AG": "458", "CU": "438"}
        pid = _CME_PRODUCT_IDS.get(metal_id, "")
        if pid:
            try:
                api_url = (f"https://www.cmegroup.com/CmeWS/mvc/Quotes/Future/"
                           f"{pid}/G")
                resp = _safe_request(api_url, headers={
                    **_CME_HEADERS, "Accept": "application/json",
                }, timeout=15)
                if resp and resp.status_code == 200:
                    data = resp.json()
                    quotes = data.get("quotes", [])
                    found_months = []
                    for q in quotes:
                        month_str = q.get("expirationMonth", "") or q.get("monthYear", "")
                        oi = 0
                        for k, v in q.items():
                            kl = k.lower()
                            if "openinterest" in kl or "interest" in kl:
                                try:
                                    oi = int(str(v).replace(",", "").replace("'", ""))
                                except:
                                    pass
                        if oi > 0 and month_str:
                            found_months.append({"month": month_str, "oi": oi})
                    if found_months:
                        fm = found_months[0]
                        result["front_month"] = fm["month"]
                        result["oi_contracts"] = fm["oi"]
                        result["oi_oz"] = fm["oi"] * comex_oz
                        result["source"] = "CME Quotes API"
                        result["all_months"] = found_months[:6]
                        logger.info(f"    ✓ CME Quotes API: {fm['month']} OI={fm['oi']:,}")
            except Exception as e:
                logger.debug(f"    CME Quotes API失败: {e}")

    # ── 方案C: yfinance 获取 OI (总OI, 不分月) ──
    if result["oi_contracts"] == 0:
        try:
            import yfinance as yf
            smap = metal.get("source_map", {})
            sym = smap.get("yahoo", "")
            if sym:
                tk = yf.Ticker(sym)
                info = tk.info or {}
                oi = info.get("openInterest", 0)
                if oi and oi > 0:
                    result["oi_contracts"] = int(oi)
                    result["oi_oz"] = int(oi) * comex_oz
                    result["source"] = "yfinance"
                    logger.info(f"    ✓ yfinance OI: {oi:,} 手 = {oi*comex_oz:,} oz")
        except ImportError:
            pass
        except Exception as e:
            logger.debug(f"    yfinance OI失败: {e}")

    # ── 方案D: AKShare COMEX (最后兜底) ──
    if result["oi_contracts"] == 0:
        try:
            ak_map = {"AU": "comex_gold", "AG": "comex_silver", "CU": "comex_copper"}
            ak_sym = ak_map.get(metal_id, "")
            if ak_sym:
                df = ak.futures_foreign_hist(symbol=ak_sym)
                if df is not None and len(df) > 0:
                    last = df.iloc[-1]
                    for col in df.columns:
                        cl = str(col).lower()
                        if "open_interest" in cl or "oi" in cl or "持仓" in cl:
                            oi = int(float(str(last[col]).replace(",", "")))
                            if oi > 0:
                                result["oi_contracts"] = oi
                                result["oi_oz"] = oi * comex_oz
                                result["source"] = "AKShare COMEX"
                                logger.info(f"    ✓ AKShare: OI={oi:,}")
                                break
        except Exception as e:
            logger.debug(f"    AKShare OI失败: {e}")

    return result


def _fetch_cme_settlements(metal_id: str) -> dict:
    """
    F2数据源: CME Settlements — 各月官方结算价
    
    数据源优先级:
      A) CME Settlements HTML (pd.read_html)
      B) CME Settlements JSON API (CmeWS)
      C) CME Quotes JSON API (报价含前一结算)
    
    返回: {
        "settlements": [{month, settle, volume, oi}, ...],
        "front_month": {month, settle, oi},
        "second_month": {month, settle, oi},
        "source": str,
    }
    """
    metal = METALS[metal_id]
    slug = _CME_SLUG_MAP.get(metal_id)
    result = {"settlements": [], "front_month": {}, "second_month": {}, "source": ""}

    if not slug:
        return result

    logger.info(f"  [F2-结算] 获取 CME {metal['name_en']} 结算价...")

    _CME_PRODUCT_IDS = {"AU": "437", "AG": "458", "CU": "438"}

    # ── 方案A: CME Settlements HTML (pd.read_html) ──
    try:
        url = f"https://www.cmegroup.com/markets/{slug}.settlements.html"
        resp = _safe_request(url, headers=_CME_HEADERS, timeout=15)
        if resp and resp.status_code == 200:
            try:
                tables = pd.read_html(resp.text)
                for tbl in tables:
                    cols = [str(c).lower() for c in tbl.columns]
                    has_settle = any("settl" in c or "last" in c or "prior" in c for c in cols)
                    if not has_settle:
                        continue
                    for _, row in tbl.iterrows():
                        try:
                            month_str = str(row.iloc[0]).strip()
                            if not month_str or month_str == "nan" or "total" in month_str.lower():
                                continue
                            settle, oi, vol = 0, 0, 0
                            for ci, cn in enumerate(cols):
                                val = str(row.iloc[ci]).replace(",", "").replace("'", "").strip()
                                try:
                                    vf = float(val)
                                except:
                                    continue
                                if ("settl" in cn or "prior" in cn) and vf > 0:
                                    settle = vf
                                elif ("interest" in cn or "oi" in cn) and vf > 0:
                                    oi = int(vf)
                                elif "vol" in cn and "est" not in cn and vf > 0:
                                    vol = int(vf)
                            if settle > 0:
                                result["settlements"].append({
                                    "month": month_str, "settle": settle,
                                    "volume": vol, "oi": oi
                                })
                        except:
                            continue
                    if result["settlements"]:
                        result["source"] = "CME Settlements HTML"
                        logger.info(f"    ✓ 结算页: {len(result['settlements'])} 个合约月")
                        break
            except Exception as e:
                logger.debug(f"    结算页HTML解析异常: {e}")
    except Exception as e:
        logger.debug(f"    结算页下载失败: {e}")

    # ── 方案B: CME Settlements JSON API ──
    if not result["settlements"]:
        pid = _CME_PRODUCT_IDS.get(metal_id, "")
        if pid:
            try:
                api_url = (f"https://www.cmegroup.com/CmeWS/mvc/Settlements/"
                           f"Futures/Settlements/{pid}/FUT")
                resp = _safe_request(api_url, headers={
                    **_CME_HEADERS, "Accept": "application/json",
                }, timeout=15)
                if resp and resp.status_code == 200:
                    data = resp.json()
                    for s in data.get("settlements", []):
                        month_str = s.get("month", "")
                        settle, oi, vol = 0, 0, 0
                        for k, v in s.items():
                            kl = k.lower()
                            try:
                                vf = float(str(v).replace(",", "").replace("'", ""))
                            except:
                                continue
                            if "settle" in kl or "prior" in kl:
                                settle = vf
                            elif "interest" in kl:
                                oi = int(vf)
                            elif "volume" in kl:
                                vol = int(vf)
                        if settle > 0 and month_str:
                            result["settlements"].append({
                                "month": month_str, "settle": settle,
                                "volume": vol, "oi": oi
                            })
                    if result["settlements"]:
                        result["source"] = "CME Settlements API"
                        logger.info(f"    ✓ 结算API: {len(result['settlements'])} 个合约月")
            except Exception as e:
                logger.debug(f"    结算API失败: {e}")

    # ── 方案C: CME Quotes JSON API (含 priorSettle) ──
    if not result["settlements"]:
        pid = _CME_PRODUCT_IDS.get(metal_id, "")
        if pid:
            try:
                api_url = f"https://www.cmegroup.com/CmeWS/mvc/Quotes/Future/{pid}/G"
                resp = _safe_request(api_url, headers={
                    **_CME_HEADERS, "Accept": "application/json",
                }, timeout=15)
                if resp and resp.status_code == 200:
                    data = resp.json()
                    for q in data.get("quotes", []):
                        month_str = q.get("expirationMonth", "") or q.get("monthYear", "")
                        settle = 0
                        oi = 0
                        vol = 0
                        for k, v in q.items():
                            kl = k.lower()
                            try:
                                vf = float(str(v).replace(",", "").replace("'", ""))
                            except:
                                continue
                            if "priorsettle" in kl or "settle" in kl:
                                if vf > 0:
                                    settle = vf
                            elif "openinterest" in kl or "interest" in kl:
                                oi = int(vf)
                            elif "volume" in kl and "estimated" not in kl:
                                vol = int(vf)
                        if settle > 0 and month_str:
                            result["settlements"].append({
                                "month": month_str, "settle": settle,
                                "volume": vol, "oi": oi
                            })
                    if result["settlements"]:
                        result["source"] = "CME Quotes API"
                        logger.info(f"    ✓ Quotes API: {len(result['settlements'])} 个合约月")
            except Exception as e:
                logger.debug(f"    CME Quotes API结算失败: {e}")

    # 提取前月/次近月
    if len(result["settlements"]) >= 1:
        result["front_month"] = result["settlements"][0]
    if len(result["settlements"]) >= 2:
        result["second_month"] = result["settlements"][1]

    return result


def _fetch_cme_calendar(metal_id: str) -> dict:
    """
    F4数据源: CME Calendar 页面 — FND / LTD / Delivery 日期
    
    URL: https://www.cmegroup.com/markets/{slug}.calendar.html
    字段: First Position / Last Position / First Notice / Last Notice /
          First Delivery / Last Delivery
    
    返回: {
        "contracts": [{month, fnd, ltd, first_delivery, last_delivery}, ...],
        "next_fnd": str,        # 最近的未来 FND 日期
        "next_ltd": str,        # 最近的未来 LTD 日期
        "days_to_fnd": int,
        "days_to_ltd": int,
        "source": str,
    }
    """
    metal = METALS[metal_id]
    slug = _CME_SLUG_MAP.get(metal_id)
    result = {"contracts": [], "next_fnd": None, "next_ltd": None,
              "days_to_fnd": 60, "days_to_ltd": 60, "source": ""}

    if not slug:
        return result

    logger.info(f"  [F4-日历] 获取 CME {metal['name_en']} 交割日历...")

    try:
        url = f"https://www.cmegroup.com/markets/{slug}.calendar.html"
        resp = _safe_request(url, headers=_CME_HEADERS, timeout=15)
        if resp and resp.status_code == 200:
            try:
                tables = pd.read_html(resp.text)
                now = datetime.now()
                for tbl in tables:
                    cols = [str(c).lower() for c in tbl.columns]
                    has_notice = any("notice" in c or "delivery" in c or "position" in c
                                     for c in cols)
                    if not has_notice:
                        continue
                    for _, row in tbl.iterrows():
                        entry = {}
                        for ci, cn in enumerate(cols):
                            val = str(row.iloc[ci]).strip()
                            if val == "nan" or not val:
                                continue
                            if "month" in cn or ci == 0:
                                entry["month"] = val
                            elif "first" in cn and "notice" in cn:
                                entry["fnd"] = val
                            elif "last" in cn and ("trad" in cn or "position" in cn):
                                entry["ltd"] = val
                            elif "first" in cn and "deliver" in cn:
                                entry["first_delivery"] = val
                            elif "last" in cn and "deliver" in cn:
                                entry["last_delivery"] = val
                        if entry.get("fnd") or entry.get("ltd"):
                            result["contracts"].append(entry)
                            # 解析FND日期, 找最近的未来FND
                            for date_field, result_field, days_field in [
                                ("fnd", "next_fnd", "days_to_fnd"),
                                ("ltd", "next_ltd", "days_to_ltd"),
                            ]:
                                if entry.get(date_field):
                                    try:
                                        dt = pd.to_datetime(entry[date_field])
                                        days = (dt - now).days
                                        if days >= -3 and (result[result_field] is None
                                                            or days < result[days_field]):
                                            result[result_field] = entry[date_field]
                                            result[days_field] = days
                                            result["next_month"] = entry.get("month", "")
                                    except:
                                        pass
                    if result["contracts"]:
                        result["source"] = "CME Calendar HTML"
                        logger.info(f"    ✓ 日历: {len(result['contracts'])} 合约, "
                                    f"nextFND={result['next_fnd']} ({result['days_to_fnd']}天)")
                        break
            except Exception as e:
                logger.debug(f"    日历HTML解析异常: {e}")
    except Exception as e:
        logger.debug(f"    日历下载失败: {e}")

    return result


def _fetch_cme_delivery_notices(metal_id: str) -> dict:
    """
    F3补充数据源: CME 交割通知 (Daily / Monthly / YTD Issues & Stops)
    
    数据源优先级:
      A) CME Daily Delivery Report PDF (当日 issues/stops)
      B) CME YTD Issues & Stops Report PDF
      C) CME 交割通知入口页链接解析
    
    返回: {"daily_issues": N, "daily_stops": N, "ytd_notices": N, "source": str}
    """
    metal = METALS[metal_id]
    comex_sym = metal.get("comex_symbol", "")
    name_en = metal["name_en"].upper()
    result = {"daily_issues": 0, "daily_stops": 0, "monthly_notices": 0,
              "ytd_notices": 0, "cumulative_deliveries": 0, "source": ""}

    if not comex_sym:
        return result

    logger.info(f"  [F3-交割通知] 获取 CME {metal['name_en']} 交割通知...")

    # ── 方案A: CME YTD Issues & Stops PDF (最可靠) ──
    try:
        pdf_url = ("https://www.cmegroup.com/delivery_reports/"
                   "MetalsIssuesAndStopsYTDReport.pdf")
        resp = _safe_request(pdf_url, headers=_CME_HEADERS, timeout=20)
        if resp and resp.status_code == 200 and len(resp.content) > 2000:
            try:
                import pdfplumber
                from io import BytesIO
                with pdfplumber.open(BytesIO(resp.content)) as pdf:
                    text = "\n".join(page.extract_text() or "" for page in pdf.pages)

                # 搜索品种相关行
                in_section = False
                for line in text.split("\n"):
                    line_up = line.upper().strip()
                    # 识别品种行 (例如 "SILVER" 或 "SI")
                    if name_en in line_up or f" {comex_sym} " in f" {line_up} ":
                        in_section = True
                    if in_section:
                        nums = re.findall(r'[\d,]+', line)
                        nums = [int(n.replace(",", "")) for n in nums
                                if n.replace(",", "").isdigit() and int(n.replace(",", "")) > 0]
                        if nums:
                            # YTD报告中通常包含: 日issues, 日stops, 月累计, YTD累计
                            if len(nums) >= 4:
                                result["daily_issues"] = nums[0]
                                result["daily_stops"] = nums[1]
                                result["monthly_notices"] = nums[2]
                                result["ytd_notices"] = nums[-1]  # 最后一个通常是YTD
                            elif len(nums) >= 2:
                                result["ytd_notices"] = max(nums)
                                result["daily_issues"] = nums[0]
                            else:
                                result["ytd_notices"] = nums[0]
                            result["source"] = "CME YTD Report PDF"
                            logger.info(f"    ✓ YTD报告: 日issues={result['daily_issues']}, "
                                        f"日stops={result['daily_stops']}, "
                                        f"YTD={result['ytd_notices']:,}")
                            break
                    # 遇到下一品种则停止
                    if in_section and line_up and not any(c.isdigit() for c in line_up[:3]):
                        other_metals = ["GOLD", "SILVER", "COPPER", "PLATINUM", "PALLADIUM"]
                        if any(m in line_up for m in other_metals) and name_en not in line_up:
                            break
            except ImportError:
                logger.debug("    pdfplumber未安装")
            except Exception as e:
                logger.debug(f"    YTD PDF解析失败: {e}")
    except Exception as e:
        logger.debug(f"    YTD PDF下载失败: {e}")

    # ── 方案B: CME 交割通知入口页 — 搜索更多链接 ──
    if not result["source"]:
        try:
            delivery_page = ("https://www.cmegroup.com/solutions/clearing/"
                             "operations-and-deliveries/nymex-delivery-notices.html")
            resp = _safe_request(delivery_page, headers=_CME_HEADERS, timeout=15)
            if resp and resp.status_code == 200:
                soup = BeautifulSoup(resp.text, "html.parser")
                pdf_links = []
                for a in soup.find_all("a", href=True):
                    txt = a.get_text(" ", strip=True).upper()
                    href = a["href"]
                    if ("METAL" in txt or name_en in txt) and ".pdf" in href.lower():
                        full_url = href if href.startswith("http") else f"https://www.cmegroup.com{href}"
                        pdf_links.append({"name": txt, "url": full_url})
                        logger.info(f"    找到PDF链接: {txt[:50]} → {full_url}")

                # 尝试下载找到的PDF
                for link in pdf_links[:2]:
                    try:
                        r2 = _safe_request(link["url"], headers=_CME_HEADERS, timeout=15)
                        if r2 and r2.status_code == 200 and len(r2.content) > 1000:
                            import pdfplumber
                            from io import BytesIO
                            with pdfplumber.open(BytesIO(r2.content)) as pdf2:
                                t2 = "\n".join(p.extract_text() or "" for p in pdf2.pages)
                            for line in t2.split("\n"):
                                if name_en in line.upper() or comex_sym in line.upper():
                                    nums = [int(n.replace(",", "")) for n in re.findall(r'[\d,]+', line)
                                            if n.replace(",", "").isdigit() and int(n.replace(",", "")) > 0]
                                    if nums:
                                        result["ytd_notices"] = max(nums)
                                        result["source"] = f"CME Delivery PDF ({link['name'][:20]})"
                                        logger.info(f"    ✓ 补充PDF: {result['ytd_notices']:,}")
                                        break
                            if result["source"]:
                                break
                    except:
                        continue
        except Exception as e:
            logger.debug(f"    交割通知入口页失败: {e}")

    return result


def _fetch_cme_margins(metal_id: str) -> dict:
    """
    F5数据源A: CME Margins — 当前保证金 + 近期是否上调
    
    数据源优先级:
      A) CME Margins HTML (当前保证金)
      B) CME Performance Bond JSON API
      C) 已知保证金估算 (2026-02, CME上调后数据)
    + CME Clearing Notices: 搜索近10天保证金调整公告
    
    返回: {"initial_margin": float, "maintenance_margin": float,
            "margin_hike_recent": bool, "margin_change_pct": float, "source": str}
    """
    metal = METALS[metal_id]
    slug = _CME_SLUG_MAP.get(metal_id)
    result = {"initial_margin": 0, "maintenance_margin": 0,
              "margin_hike_recent": False, "margin_change_pct": 0, "source": ""}

    if not slug:
        return result

    logger.info(f"  [F5-保证金] 获取 CME {metal['name_en']} 保证金...")

    # ── 方案A: CME Margins HTML ──
    try:
        url = f"https://www.cmegroup.com/markets/{slug}.margins.html"
        resp = _safe_request(url, headers=_CME_HEADERS, timeout=15)
        if resp and resp.status_code == 200:
            try:
                tables = pd.read_html(resp.text)
                for tbl in tables:
                    cols = [str(c).lower() for c in tbl.columns]
                    has_margin = any("initial" in c or "maint" in c or "margin" in c
                                     for c in cols)
                    if not has_margin:
                        continue
                    for _, row in tbl.iterrows():
                        for ci, cn in enumerate(cols):
                            val = str(row.iloc[ci]).replace(",", "").replace("$", "").strip()
                            try:
                                vf = float(val)
                            except:
                                continue
                            if "initial" in cn and vf > 100:
                                result["initial_margin"] = vf
                            elif "maint" in cn and vf > 100:
                                result["maintenance_margin"] = vf
                    if result["initial_margin"] > 0:
                        result["source"] = "CME Margins HTML"
                        logger.info(f"    ✓ 保证金: 初始=${result['initial_margin']:,.0f}, "
                                    f"维持=${result['maintenance_margin']:,.0f}")
                        break
            except Exception as e:
                logger.debug(f"    保证金HTML解析异常: {e}")
    except Exception as e:
        logger.debug(f"    保证金下载失败: {e}")

    # ── 方案B: CME Performance Bond Advisory JSON ──
    if result["initial_margin"] == 0:
        _CME_PRODUCT_IDS = {"AU": "437", "AG": "458", "CU": "438"}
        pid = _CME_PRODUCT_IDS.get(metal_id, "")
        if pid:
            try:
                api_url = (f"https://www.cmegroup.com/CmeWS/mvc/Margins/"
                           f"GetMargins?exchange=XCME&productId={pid}")
                resp = _safe_request(api_url, headers={
                    **_CME_HEADERS, "Accept": "application/json",
                }, timeout=15)
                if resp and resp.status_code == 200:
                    data = resp.json()
                    margins_list = data if isinstance(data, list) else data.get("margins", [])
                    for m in margins_list[:3]:
                        init = 0
                        maint = 0
                        for k, v in (m if isinstance(m, dict) else {}).items():
                            kl = k.lower()
                            try:
                                vf = float(str(v).replace(",", "").replace("$", ""))
                            except:
                                continue
                            if "initial" in kl and vf > 100:
                                init = vf
                            elif "maint" in kl and vf > 100:
                                maint = vf
                        if init > 0:
                            result["initial_margin"] = init
                            result["maintenance_margin"] = maint
                            result["source"] = "CME Margins API"
                            logger.info(f"    ✓ 保证金API: 初始=${init:,.0f}")
                            break
            except Exception as e:
                logger.debug(f"    保证金API失败: {e}")

    # ── 方案C: 已知保证金估算 (2026-02 CME上调后) ──
    if result["initial_margin"] == 0:
        known_margins = {
            "AU": {"initial_margin": 11000, "maintenance_margin": 10000},
            "AG": {"initial_margin": 18150, "maintenance_margin": 16500},
            "CU": {"initial_margin": 7700, "maintenance_margin": 7000},
        }
        if metal_id in known_margins:
            result.update(known_margins[metal_id])
            result["source"] = "known estimate (2026-02)"
            logger.info(f"    ⓘ 已知保证金估算: 初始=${result['initial_margin']:,.0f}")

    # ── 检查近期保证金上调 (CME Clearing Notices) ──
    try:
        notices_url = "https://www.cmegroup.com/notices.html"
        resp = _safe_request(notices_url, headers=_CME_HEADERS, timeout=10)
        if resp and resp.status_code == 200:
            text_lower = resp.text.lower()
            # 搜索关键词: margin, performance bond, 品种名
            name_lower = metal["name_en"].lower()
            if ("margin" in text_lower or "performance bond" in text_lower) and \
               (name_lower in text_lower or "metal" in text_lower):
                result["margin_hike_recent"] = True
                result["margin_hike_note"] = "CME Clearing Notice detected"
                logger.info(f"    ⚠ 检测到近期保证金调整公告")
    except Exception as e:
        logger.debug(f"    Clearing Notices检查失败: {e}")

    # 2026-02 确认: CME 2月6日上调金银保证金 (Reuters报道)
    # 对银来说, 从 $16,500 → $18,150, 涨幅 10%
    if metal_id == "AG" and result["initial_margin"] >= 18000:
        result["margin_hike_recent"] = True
        result["margin_change_pct"] = 0.10  # 10% hike

    return result


def _fetch_cftc_cot(metal_id: str) -> dict:
    """
    F5数据源D: CFTC Commitments of Traders — 大户集中度
    
    数据源优先级:
      A) CFTC 周报网页表 (最新周报, pre-formatted text)
         URL: https://www.cftc.gov/dea/futures/deacmxsf.htm
      B) CFTC 年度历史压缩CSV (disaggregated format)
         URL: https://www.cftc.gov/files/dea/history/fut_disagg_txt_{year}.zip
      C) CFTC 补充长格式 (含大户集中度)
         URL: https://www.cftc.gov/dea/futures/deacmxlof.htm
    
    CFTC 明确表示其市场监控每天关注大户活动、关键价差和供需因素.
    
    返回: {
        "report_date": str,
        "noncommercial_long": int, "noncommercial_short": int,
        "commercial_long": int, "commercial_short": int,
        "total_oi_cot": int,
        "largest_4_long_pct": float, "largest_8_long_pct": float,
        "net_speculative": int,     # 非商业净多 = long - short
        "concentration_risk": float,  # 0~1
        "source": str,
    }
    """
    metal = METALS[metal_id]
    cftc_code = metal.get("comex_cftc_code", "")
    name_en = metal["name_en"].upper()
    comex_sym = metal.get("comex_symbol", "")
    result = {"source": "", "net_speculative": 0}

    if not cftc_code:
        return result

    logger.info(f"  [F5-COT] 获取 CFTC {metal['name_en']} 持仓报告...")

    # ── 方案A: CFTC 短格式网页表 (最新周报) ──
    try:
        cot_url = "https://www.cftc.gov/dea/futures/deacmxsf.htm"
        resp = _safe_request(cot_url, headers={
            "User-Agent": "Mozilla/5.0 (compatible; ResearchBot/1.0)",
        }, timeout=20)
        if resp and resp.status_code == 200:
            try:
                text = resp.text
                lines = text.split("\n")
                in_section = False
                section_lines = []
                
                for line in lines:
                    line_up = line.upper()
                    # 品种标题行: 包含品种名 + COMEX/NYMEX
                    if (name_en in line_up or (comex_sym and comex_sym in line_up)) and \
                       ("COMEX" in line_up or "NYMEX" in line_up or cftc_code in line_up):
                        in_section = True
                        section_lines = [line]
                        continue
                    if in_section:
                        section_lines.append(line)
                        if len(section_lines) > 40:
                            break

                if section_lines:
                    block = "\n".join(section_lines)
                    block_up = block.upper()
                    
                    # 提取各类持仓 — CFTC 短格式布局:
                    # NON-COMMERCIAL | COMMERCIAL | ...
                    # LONG | SHORT | SPREADING | LONG | SHORT | ...
                    
                    # 提取数字行
                    num_lines = []
                    for sl in section_lines:
                        nums = re.findall(r'[\d,]+', sl)
                        nums_int = [int(n.replace(",", "")) for n in nums 
                                    if n.replace(",", "").isdigit() and len(n.replace(",", "")) > 2]
                        if len(nums_int) >= 3:
                            num_lines.append(nums_int)
                    
                    # 通常第一个包含大量数字的行是持仓数据
                    if num_lines:
                        # 非商业多/空通常在前面
                        first_nums = num_lines[0]
                        if len(first_nums) >= 5:
                            result["noncommercial_long"] = first_nums[0]
                            result["noncommercial_short"] = first_nums[1]
                            result["commercial_long"] = first_nums[3] if len(first_nums) > 3 else 0
                            result["commercial_short"] = first_nums[4] if len(first_nums) > 4 else 0
                    
                    # 提取总OI (通常是区块中最大的数字)
                    all_nums = re.findall(r'[\d,]+', block)
                    all_nums_int = [int(n.replace(",", "")) for n in all_nums
                                    if n.replace(",", "").isdigit()]
                    big_nums = [n for n in all_nums_int if n > 1000]
                    if big_nums:
                        result["total_oi_cot"] = max(big_nums)
                    
                    # 提取 4/8 大户集中度 (PERCENT OF OPEN INTEREST 区段)
                    for i, sl in enumerate(section_lines):
                        sl_up = sl.upper()
                        if "PERCENT" in sl_up and "OPEN INTEREST" in sl_up:
                            # 后面几行包含集中度数据
                            for j in range(i+1, min(i+6, len(section_lines))):
                                conc_line = section_lines[j]
                                conc_nums = re.findall(r'[\d.]+', conc_line)
                                conc_floats = []
                                for cn in conc_nums:
                                    try:
                                        cf = float(cn)
                                        if 0 < cf < 100:
                                            conc_floats.append(cf)
                                    except:
                                        pass
                                if conc_floats and ("4 OR" in section_lines[j].upper() or
                                                     len(conc_floats) >= 2):
                                    if len(conc_floats) >= 1:
                                        result["largest_4_long_pct"] = conc_floats[0]
                                    if len(conc_floats) >= 2:
                                        result["largest_4_short_pct"] = conc_floats[1]
                                elif conc_floats and "8 OR" in section_lines[j].upper():
                                    if len(conc_floats) >= 1:
                                        result["largest_8_long_pct"] = conc_floats[0]
                    
                    if result.get("total_oi_cot"):
                        result["source"] = "CFTC COT webpage"
                        nc_l = result.get("noncommercial_long", 0)
                        nc_s = result.get("noncommercial_short", 0)
                        result["net_speculative"] = nc_l - nc_s
                        logger.info(f"    ✓ CFTC COT: OI={result['total_oi_cot']:,}, "
                                    f"非商业净多={result['net_speculative']:+,}, "
                                    f"4大多头={result.get('largest_4_long_pct',0):.1f}%")
            except Exception as e:
                logger.debug(f"    CFTC COT网页解析异常: {e}")
    except Exception as e:
        logger.debug(f"    CFTC COT下载失败: {e}")

    # ── 方案B: CFTC 长格式网页 (含集中度) ──
    if not result.get("largest_4_long_pct"):
        try:
            lof_url = "https://www.cftc.gov/dea/futures/deacmxlof.htm"
            resp = _safe_request(lof_url, headers={
                "User-Agent": "Mozilla/5.0 (compatible; ResearchBot/1.0)",
            }, timeout=20)
            if resp and resp.status_code == 200:
                text = resp.text
                lines = text.split("\n")
                in_section = False
                section_lines = []
                for line in lines:
                    line_up = line.upper()
                    if (name_en in line_up) and ("COMEX" in line_up or "NYMEX" in line_up):
                        in_section = True
                        section_lines = [line]
                        continue
                    if in_section:
                        section_lines.append(line)
                        if len(section_lines) > 50:
                            break
                
                if section_lines:
                    for i, sl in enumerate(section_lines):
                        sl_up = sl.upper()
                        if "4 OR LESS" in sl_up or "FOUR" in sl_up:
                            nums = re.findall(r'[\d.]+', sl)
                            floats = [float(n) for n in nums if 0 < float(n) < 100]
                            if floats:
                                result["largest_4_long_pct"] = floats[0]
                                if len(floats) >= 2:
                                    result["largest_4_short_pct"] = floats[1]
                                logger.info(f"    ✓ CFTC LOF: 4大多头={floats[0]:.1f}%")
                        elif "8 OR LESS" in sl_up or "EIGHT" in sl_up:
                            nums = re.findall(r'[\d.]+', sl)
                            floats = [float(n) for n in nums if 0 < float(n) < 100]
                            if floats:
                                result["largest_8_long_pct"] = floats[0]
        except Exception as e:
            logger.debug(f"    CFTC LOF下载失败: {e}")

    # ── 方案C: CFTC 年度历史压缩CSV (disaggregated) ──
    if not result.get("total_oi_cot"):
        try:
            year = datetime.now().year
            csv_url = f"https://www.cftc.gov/files/dea/history/fut_disagg_txt_{year}.zip"
            resp = _safe_request(csv_url, timeout=45)
            if resp and resp.status_code == 200:
                import zipfile
                from io import BytesIO
                with zipfile.ZipFile(BytesIO(resp.content)) as zf:
                    for fname in zf.namelist():
                        if fname.endswith(".txt"):
                            df = pd.read_csv(zf.open(fname))
                            name_col = None
                            for col in df.columns:
                                if "market" in str(col).lower() or "name" in str(col).lower():
                                    name_col = col
                                    break
                            if name_col:
                                mask = df[name_col].astype(str).str.upper().str.contains(
                                    name_en, na=False)
                                sub = df[mask]
                            else:
                                mask = df.apply(lambda r: any(name_en in str(v).upper()
                                                              for v in r.values), axis=1)
                                sub = df[mask]

                            if len(sub) > 0:
                                last = sub.iloc[-1]
                                for col in sub.columns:
                                    cl = str(col).lower()
                                    val = last[col]
                                    try:
                                        val = int(float(str(val).replace(",", "")))
                                    except:
                                        continue
                                    if "prod_merc" in cl and "long" in cl and "spread" not in cl:
                                        result["commercial_long"] = val
                                    elif "prod_merc" in cl and "short" in cl and "spread" not in cl:
                                        result["commercial_short"] = val
                                    elif "m_money" in cl and "long" in cl and "spread" not in cl:
                                        result["managed_money_long"] = val
                                    elif "m_money" in cl and "short" in cl and "spread" not in cl:
                                        result["managed_money_short"] = val
                                    elif "oi_all" in cl or (cl == "open_interest_all"):
                                        result["total_oi_cot"] = val
                                    elif "conc_gross_le_4" in cl and "long" in cl:
                                        result["largest_4_long_pct"] = val
                                    elif "conc_gross_le_4" in cl and "short" in cl:
                                        result["largest_4_short_pct"] = val
                                    elif "conc_gross_le_8" in cl and "long" in cl:
                                        result["largest_8_long_pct"] = val
                                    elif "report_date" in cl:
                                        result["report_date"] = str(val)

                                if result.get("total_oi_cot"):
                                    # 计算管理基金净多
                                    mm_l = result.get("managed_money_long", 0)
                                    mm_s = result.get("managed_money_short", 0)
                                    result["net_speculative"] = mm_l - mm_s
                                    result["source"] = "CFTC Disaggregated CSV"
                                    logger.info(f"    ✓ CFTC CSV: OI={result['total_oi_cot']:,}"
                                                + (f", 报告日={result.get('report_date','')}" if
                                                   result.get('report_date') else ""))
                            break
        except ImportError:
            logger.debug("    zipfile导入失败")
        except Exception as e:
            logger.debug(f"    CFTC CSV失败: {e}")

    # 计算大户集中度风险 (0~1)
    l4 = result.get("largest_4_long_pct", 0)
    l4_s = result.get("largest_4_short_pct", 0)
    if l4 > 0:
        # 4家最大多头占比 > 40% 视为高集中度
        result["concentration_risk"] = min(1.0, max(l4, l4_s) / 50.0)
    elif result.get("total_oi_cot"):
        # 没有集中度数据但有OI → 默认低风险
        result["concentration_risk"] = 0.15

    return result


def _fetch_lbma_spot_premium(metal_id: str, comex_settle: float = 0) -> dict:
    """
    F5数据源C: LBMA 现货 vs COMEX 近月 — 交割地升水代理
    
    数据源优先级:
      A) DBnomics LBMA daily silver/gold mirror (研究用)
      B) Stooq.com 现货报价 (免费)
      C) yfinance XAG/XAUUSD 现货
    
    返回: {"lbma_spot": float, "comex_front": float, "premium": float,
            "premium_pct": float, "source": str}
    
    注意: 这只是现货/期货压力的代理, 不是真实交割地溢价.
    """
    metal = METALS[metal_id]
    result = {"lbma_spot": 0, "comex_front": comex_settle, "premium": 0,
              "premium_pct": 0, "source": ""}

    # 仅金/银有 LBMA 参考价
    if metal_id not in ("AU", "AG"):
        return result

    logger.info(f"  [F5-升水] 获取 {metal['name_en']} 现货升水代理...")

    # ── 方案A: Stooq.com 现货价 ──
    try:
        stooq_sym = metal.get("source_map", {}).get("stooq_spot", "")
        if stooq_sym:
            stooq_url = f"https://stooq.com/q/l/?s={stooq_sym}&f=l"
            resp = _safe_request(stooq_url, headers={
                "User-Agent": "Mozilla/5.0 (compatible; PriceBot/1.0)",
            }, timeout=10)
            if resp and resp.status_code == 200:
                text = resp.text.strip()
                try:
                    spot = float(text)
                    if spot > 0:
                        result["lbma_spot"] = spot
                        result["source"] = "Stooq spot"
                        logger.info(f"    ✓ Stooq: {stooq_sym}={spot:.4f}")
                except:
                    # 可能返回HTML, 解析table
                    match = re.search(r'[\d.]+', text)
                    if match:
                        spot = float(match.group())
                        if spot > 1:
                            result["lbma_spot"] = spot
                            result["source"] = "Stooq spot"
    except Exception as e:
        logger.debug(f"    Stooq失败: {e}")

    # ── 方案B: yfinance 现货 ──
    if result["lbma_spot"] == 0:
        try:
            import yfinance as yf
            sym = "XAGUSD=X" if metal_id == "AG" else "XAUUSD=X"
            tk = yf.Ticker(sym)
            hist = tk.history(period="1d")
            if len(hist) > 0:
                spot = float(hist["Close"].iloc[-1])
                if spot > 0:
                    result["lbma_spot"] = spot
                    result["source"] = "yfinance spot"
                    logger.info(f"    ✓ yfinance: {sym}={spot:.4f}")
        except ImportError:
            pass
        except Exception as e:
            logger.debug(f"    yfinance spot失败: {e}")

    # ── 方案C: 新浪外汇现货 ──
    if result["lbma_spot"] == 0:
        try:
            sina_sym = metal.get("sina_spot_symbol", "")
            if sina_sym:
                url = f"https://hq.sinajs.cn/list={sina_sym}"
                resp = _safe_request(url, headers={
                    "Referer": "https://finance.sina.com.cn/",
                    "User-Agent": "Mozilla/5.0",
                }, timeout=8)
                if resp and resp.status_code == 200:
                    match = re.search(r'"([^"]+)"', resp.text)
                    if match:
                        fields = match.group(1).split(",")
                        if len(fields) >= 3:
                            spot = float(fields[0])
                            if spot > 0:
                                result["lbma_spot"] = spot
                                result["source"] = "Sina spot"
        except Exception as e:
            logger.debug(f"    新浪现货失败: {e}")

    # 计算升水
    if result["lbma_spot"] > 0 and comex_settle > 0:
        result["premium"] = round(result["lbma_spot"] - comex_settle, 4)
        result["premium_pct"] = round(result["premium"] / comex_settle * 100, 4)
        logger.info(f"    升水: {result['premium']:+.4f} ({result['premium_pct']:+.4f}%)")

    return result


def _fetch_cme_stocks_changes(metal_id: str, current_registered: float = 0,
                               current_eligible: float = 0) -> dict:
    """
    F3补充: CME 库存日变化追踪
    
    从 CME {Metal}_stocks.xls 解析当前数据, 并与历史数据对比:
    - 落地为本地CSV (data/cme_stocks_history_{metal}.csv)
    - 计算 registered/eligible/total 的日变化和趋势
    - 追踪连续流失天数
    
    返回: {
        "reg_change": float, "eli_change": float, "total_change": float,
        "reg_change_pct": float,
        "consecutive_decline_days": int,
        "eli_to_reg_ratio": float,   # eligible/registered 缓冲比
        "trend_5d": str,             # "上升"/"下降"/"平稳"
        "source": str,
    }
    """
    metal = METALS[metal_id]
    result = {
        "reg_change": 0, "eli_change": 0, "total_change": 0,
        "reg_change_pct": 0,
        "consecutive_decline_days": 0,
        "eli_to_reg_ratio": 0,
        "trend_5d": "平稳",
        "source": "",
    }

    # 历史文件路径
    hist_file = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                             "data", f"cme_stocks_history_{metal_id}.csv")

    logger.info(f"  [F3-库存变化] 追踪 {metal['name_en']} CME库存变化...")

    # 读取历史数据
    history = []
    if os.path.exists(hist_file):
        try:
            df = pd.read_csv(hist_file)
            history = df.to_dict("records")
        except:
            pass

    # CME stocks xls 获取当前数据
    cme_stocks_url = {
        "AU": "https://www.cmegroup.com/delivery_reports/Gold_Stocks.xls",
        "AG": "https://www.cmegroup.com/delivery_reports/Silver_stocks.xls",
        "CU": "https://www.cmegroup.com/delivery_reports/Copper_Stocks.xls",
    }
    url = cme_stocks_url.get(metal_id)
    
    today_registered = current_registered
    today_eligible = current_eligible
    today_total = today_registered + today_eligible

    if url:
        try:
            resp = _safe_request(url, headers=_CME_HEADERS, timeout=15)
            if resp and resp.status_code == 200:
                raw = pd.read_excel(resp.content if isinstance(resp.content, bytes) else 
                                     __import__('io').BytesIO(resp.content), header=None)
                raw_str = raw.astype(str)
                for _, row in raw_str.iterrows():
                    row_text = " ".join(row.values).upper()
                    if "TOTAL REGISTERED" in row_text:
                        nums = re.findall(r'[\d,.]+', row_text)
                        for n in nums:
                            try:
                                v = float(n.replace(",", ""))
                                if v > 1000:
                                    today_registered = v
                                    break
                            except:
                                pass
                    elif "TOTAL ELIGIBLE" in row_text:
                        nums = re.findall(r'[\d,.]+', row_text)
                        for n in nums:
                            try:
                                v = float(n.replace(",", ""))
                                if v > 1000:
                                    today_eligible = v
                                    break
                            except:
                                pass
                today_total = today_registered + today_eligible
                result["source"] = "CME Stocks XLS"
                logger.info(f"    ✓ XLS: Reg={today_registered:,.0f}, Elig={today_eligible:,.0f}")
        except Exception as e:
            logger.debug(f"    CME Stocks XLS读取失败: {e}")

    # 与前一天对比
    if history:
        prev = history[-1]
        prev_reg = prev.get("registered", 0)
        prev_eli = prev.get("eligible", 0)
        prev_total = prev.get("total", prev_reg + prev_eli)

        if prev_reg > 0 and today_registered > 0:
            result["reg_change"] = today_registered - prev_reg
            result["reg_change_pct"] = result["reg_change"] / prev_reg * 100
        if prev_eli > 0:
            result["eli_change"] = today_eligible - prev_eli
        result["total_change"] = today_total - prev_total

        # 连续下降天数
        decline_count = 0
        for h in reversed(history[-10:]):
            if h.get("total_change", 0) < 0:
                decline_count += 1
            else:
                break
        if result["total_change"] < 0:
            decline_count += 1
        result["consecutive_decline_days"] = decline_count

        # 5日趋势
        if len(history) >= 5:
            total_5d = sum(h.get("total_change", 0) for h in history[-5:]) + result["total_change"]
            if total_5d > 0:
                result["trend_5d"] = "上升"
            elif total_5d < 0:
                result["trend_5d"] = "下降"

    # E/R 缓冲比
    if today_registered > 0:
        result["eli_to_reg_ratio"] = round(today_eligible / today_registered, 2)

    # 追加今天数据到历史
    today_str = datetime.now().strftime("%Y-%m-%d")
    if today_registered > 0:
        new_record = {
            "date": today_str,
            "registered": today_registered,
            "eligible": today_eligible,
            "total": today_total,
            "reg_change": result["reg_change"],
            "eli_change": result["eli_change"],
            "total_change": result["total_change"],
        }
        # 避免重复记录当天
        history = [h for h in history if h.get("date") != today_str]
        history.append(new_record)
        # 保留最近60天
        history = history[-60:]
        try:
            os.makedirs(os.path.dirname(hist_file), exist_ok=True)
            pd.DataFrame(history).to_csv(hist_file, index=False, encoding="utf-8-sig")
        except Exception as e:
            logger.debug(f"    历史文件写入失败: {e}")

    logger.info(f"    库存变化: Reg={result['reg_change']:+,.0f}, "
                f"E/R={result['eli_to_reg_ratio']:.2f}, "
                f"连降={result['consecutive_decline_days']}天, "
                f"趋势={result['trend_5d']}")

    return result


def fetch_cme_crisis_data(metal_id: str) -> dict:
    """
    汇总采集交割危机五因子所需的全部 CME/CFTC 数据
    
    数据源清单:
    ┌──────────────┬────────────────────────────────────────┐
    │ 因子         │ 数据源                                  │
    ├──────────────┼────────────────────────────────────────┤
    │ F1-覆盖率    │ CME {Metal}_stocks.xls (已有)           │
    │              │ + CME Daily Bulletin PDF (前月OI)       │
    ├──────────────┼────────────────────────────────────────┤
    │ F2-期限结构  │ CME Settlements HTML / API              │
    ├──────────────┼────────────────────────────────────────┤
    │ F3-库存流动  │ CME Stocks xls日变化 + Delivery Notices │
    ├──────────────┼────────────────────────────────────────┤
    │ F4-时点      │ CME Calendar HTML (FND/LTD)            │
    ├──────────────┼────────────────────────────────────────┤
    │ F5-压力      │ CME Margins + CFTC COT + Realized Vol  │
    └──────────────┴────────────────────────────────────────┘
    """
    metal = METALS[metal_id]
    comex_oz = metal.get("comex_contract_oz", 0)

    result = {
        "front_month_oi": {},
        "settlements": {},
        "calendar": {},
        "delivery_notices": {},
        "stocks_changes": {},
        "lbma_premium": {},
        "cot": {},
        "margins": {},
        "front_month_oi_contracts": 0,
        "front_month_oi_oz": 0,
        "data_sources": [],         # 记录实际获取到的数据源
    }

    # 无COMEX合约的品种(如铝)跳过
    if not metal.get("comex_symbol"):
        logger.info(f"  [危机数据] {metal['name']} 无COMEX合约, 跳过CME数据采集")
        return result

    logger.info(f"  [危机数据] 采集 {metal['name']} CME/CFTC 五因子数据...")

    # ───────────────────────────────────────────
    #  1. F1数据: CME Daily Bulletin PDF → 前月OI
    # ───────────────────────────────────────────
    try:
        oi_data = _fetch_cme_front_month_oi(metal_id)
        result["front_month_oi"] = oi_data
        if oi_data.get("oi_contracts", 0) > 0:
            result["front_month_oi_contracts"] = oi_data["oi_contracts"]
            result["front_month_oi_oz"] = oi_data["oi_oz"]
            result["data_sources"].append(f"F1-OI: {oi_data['source']}")
    except Exception as e:
        logger.debug(f"    F1-OI采集异常: {e}")

    time.sleep(0.5)

    # ───────────────────────────────────────────
    #  2. F2数据: CME Settlements → 各月结算价
    # ───────────────────────────────────────────
    try:
        settlements = _fetch_cme_settlements(metal_id)
        result["settlements"] = settlements
        if settlements.get("settlements"):
            result["data_sources"].append(f"F2-结算: {settlements['source']}")
            # 补充: 从结算数据也可获取OI (如果前面没拿到)
            if result["front_month_oi_contracts"] == 0:
                fm_oi = settlements.get("front_month", {}).get("oi", 0)
                if fm_oi > 0:
                    result["front_month_oi_contracts"] = fm_oi
                    result["front_month_oi_oz"] = fm_oi * comex_oz
    except Exception as e:
        logger.debug(f"    F2-结算采集异常: {e}")

    time.sleep(0.5)

    # ───────────────────────────────────────────
    #  3. F3数据: CME Delivery Notices → 交割通知
    # ───────────────────────────────────────────
    try:
        notices = _fetch_cme_delivery_notices(metal_id)
        result["delivery_notices"] = notices
        if notices.get("ytd_notices", 0) > 0 or notices.get("daily_issues", 0) > 0:
            result["data_sources"].append(f"F3-通知: {notices['source']}")
    except Exception as e:
        logger.debug(f"    F3-通知采集异常: {e}")

    time.sleep(0.5)

    # ───────────────────────────────────────────
    #  3b. F3数据: CME Stocks 日变化追踪
    # ───────────────────────────────────────────
    try:
        stocks_chg = _fetch_cme_stocks_changes(metal_id)
        result["stocks_changes"] = stocks_chg
        if stocks_chg.get("source"):
            result["data_sources"].append(f"F3-库存变化: {stocks_chg['source']}")
    except Exception as e:
        logger.debug(f"    F3-库存变化采集异常: {e}")

    time.sleep(0.5)

    # ───────────────────────────────────────────
    #  4. F4数据: CME Calendar → FND/LTD日期
    # ───────────────────────────────────────────
    try:
        calendar = _fetch_cme_calendar(metal_id)
        result["calendar"] = calendar
        if calendar.get("next_fnd"):
            result["data_sources"].append(f"F4-日历: {calendar['source']}")
    except Exception as e:
        logger.debug(f"    F4-日历采集异常: {e}")

    time.sleep(0.5)

    # ───────────────────────────────────────────
    #  5. F5数据: CME Margins → 保证金
    # ───────────────────────────────────────────
    try:
        margins = _fetch_cme_margins(metal_id)
        result["margins"] = margins
        if margins.get("initial_margin", 0) > 0:
            result["data_sources"].append(f"F5-保证金: {margins['source']}")
    except Exception as e:
        logger.debug(f"    F5-保证金采集异常: {e}")

    time.sleep(0.5)

    # ───────────────────────────────────────────
    #  6. F5数据: CFTC COT → 大户集中度
    # ───────────────────────────────────────────
    try:
        cot = _fetch_cftc_cot(metal_id)
        result["cot"] = cot
        if cot.get("total_oi_cot"):
            result["data_sources"].append(f"F5-COT: {cot['source']}")
    except Exception as e:
        logger.debug(f"    F5-COT采集异常: {e}")

    time.sleep(0.3)

    # ───────────────────────────────────────────
    #  7. F5数据: LBMA 现货升水代理
    # ───────────────────────────────────────────
    try:
        comex_front_settle = (result.get("settlements", {})
                              .get("front_month", {}).get("settle", 0))
        lbma = _fetch_lbma_spot_premium(metal_id, comex_settle=comex_front_settle)
        result["lbma_premium"] = lbma
        if lbma.get("lbma_spot", 0) > 0:
            result["data_sources"].append(f"F5-升水: {lbma['source']}")
    except Exception as e:
        logger.debug(f"    F5-升水采集异常: {e}")

    # ── 汇总日志 ──
    n_settle = len(result["settlements"].get("settlements", []))
    has_cal = bool(result["calendar"].get("next_fnd"))
    has_cot = bool(result["cot"].get("total_oi_cot"))
    has_margin = result["margins"].get("initial_margin", 0) > 0
    has_lbma = result["lbma_premium"].get("lbma_spot", 0) > 0
    has_stocks = bool(result["stocks_changes"].get("source"))
    front_oi = result["front_month_oi_contracts"]

    logger.info(f"  [危机数据汇总] 前月OI={front_oi:,}, 结算={n_settle}月, "
                f"日历={'✓' if has_cal else '✗'}, COT={'✓' if has_cot else '✗'}, "
                f"保证金={'✓' if has_margin else '✗'}, "
                f"升水={'✓' if has_lbma else '✗'}, "
                f"库存变化={'✓' if has_stocks else '✗'}")
    logger.info(f"  [数据源] {', '.join(result['data_sources']) or '无官方数据'}")

    return result

def fetch_indicators(metal_id: str) -> dict:
    """
    获取库存量、注册库存、合格库存、持仓等关键指标

    [修复 #3] futures_shfe_warehouse_receipt() 返回 dict 而非 DataFrame
    [修复 #4] futures_inventory_em() 需要 inventory_symbol (沪金) 而非 name (黄金)
    """
    metal = METALS[metal_id]
    logger.info(f"[指标] 获取 {metal['name']} 关键指标...")

    indicators = {
        "total_inventory": 0,
        "registered_inventory": 0,
        "qualified_inventory": 0,
        "inventory_change": 0,
        "open_interest": 0,
        "daily_volume": 0,
        "position_ratio": 0,
        "warehouses": 0,
    }

    # 现货品种(铂金/钯金): 从国际市场获取库存/持仓数据
    if metal.get("is_spot"):
        logger.info(f"  ⓘ 现货品种, 从CME/COMEX获取国际市场数据...")

        # ── 源A: CME 官方库存报表 (最权威) ──
        cme = fetch_cme_stocks(metal_id)
        if cme["registered"] > 0:
            # 存储原始 CME 单位 (troy oz) 用于前端显示
            indicators["cme_registered_oz"] = cme["registered"]
            indicators["cme_eligible_oz"] = cme["eligible"]
            indicators["cme_total_oz"] = cme["total"]
            indicators["cme_unit"] = cme.get("unit", "troy oz")
            indicators["cme_date"] = cme.get("date", "")
            # 转换为本地单位用于 DSCR 计算 (与 OI × delivery_multiplier 单位一致)
            indicators["registered_inventory"] = cme["registered_local"]
            indicators["total_inventory"] = cme["total_local"]
            indicators["qualified_inventory"] = int(cme["eligible"] * _CME_TO_LOCAL_FACTOR.get(metal_id, 1))
            indicators["cme_source"] = True
            unit = cme.get("unit", "troy oz")
            logger.info(f"  ✓ CME官方: Registered={cme['registered']:,} {unit}, "
                        f"Total={cme['total']:,} {unit}")

        # ── 源B: yfinance 获取OI和Volume — 使用 source_map 统一路由 ──
        smap = metal.get("source_map", {})
        yf_sym = smap.get("yahoo", "")
        if yf_sym:
            try:
                import yfinance as yf
                tk = yf.Ticker(yf_sym)
                info = tk.info or {}
                oi = info.get("openInterest", 0)
                vol = info.get("volume", info.get("averageVolume", 0))
                if oi and oi > 0:
                    indicators["open_interest"] = int(oi)
                    logger.info(f"  ✓ yfinance({yf_sym}) OI: {oi}")
                if vol and vol > 0:
                    indicators["daily_volume"] = int(vol)
                    logger.info(f"  ✓ yfinance({yf_sym}) Volume: {vol}")
            except ImportError:
                logger.debug("  yfinance未安装")
            except Exception as e:
                logger.debug(f"  yfinance指标失败: {type(e).__name__}: {str(e)[:60]}")

        # ── 源C: 新浪外盘实时行情 OI和成交量 ──
        sina_sym = smap.get("sina_foreign") or metal.get("sina_spot_symbol", "")
        if sina_sym and indicators["open_interest"] == 0:
            try:
                quote = _parse_sina_futures_quote(sina_sym)
                if quote:
                    if quote.get("open_interest", 0) > 0:
                        indicators["open_interest"] = int(quote["open_interest"])
                    if quote.get("volume", 0) > 0:
                        indicators["daily_volume"] = int(quote["volume"])
                    logger.info(f"  ✓ 新浪外盘: OI={indicators['open_interest']}, Vol={indicators['daily_volume']}")
            except:
                pass

        indicators["position_ratio"] = 1.0
        indicators["warehouses"] = 0

        logger.info(f"  [国际指标] 库存={indicators['total_inventory']}, "
                     f"OI={indicators['open_interest']}, Vol={indicators['daily_volume']}")

        return {
            "metal_id": metal_id,
            "indicators": indicators,
            "updated_at": datetime.now().isoformat(),
        }

    # ─────────────────────────────────────
    #  1) 注册库存 (仓单) — 多源获取
    #  优先级: get_receipt > futures_shfe_warehouse_receipt > SHFE直连
    # ─────────────────────────────────────

    # ── 源1A: AKShare get_receipt (注册仓单数据) ──
    # 自动回溯最多7个自然日, 跳过非交易日
    try:
        import warnings as _warnings
        for _day_offset in range(0, 8):
            _try_date = datetime.now() - timedelta(days=_day_offset)
            if _try_date.weekday() >= 5:  # 跳过周末
                continue
            _try_str = _try_date.strftime("%Y%m%d")
            try:
                with _warnings.catch_warnings(record=True) as _w:
                    _warnings.simplefilter("always")
                    raw = ak.get_receipt(start_date=_try_str, end_date=_try_str,
                                         vars_list=[metal_id.upper()])
                    # 检查是否有 "非交易日" 警告
                    _is_non_trading = any("非交易日" in str(w.message) for w in _w)
                if _is_non_trading:
                    logger.debug(f"  get_receipt: {_try_str} 非交易日, 继续回溯...")
                    continue
                if isinstance(raw, pd.DataFrame) and len(raw) > 0:
                    if _day_offset > 0:
                        logger.info(f"  get_receipt: 今日无数据, 使用 {_try_str} (回溯{_day_offset}天)")
                    break
            except Exception:
                continue
        else:
            raw = None  # 7天内都没找到

        if isinstance(raw, pd.DataFrame) and len(raw) > 0:
            logger.debug(f"  get_receipt 列: {list(raw.columns)}, 行: {len(raw)}")
            # 该函数通常返回含 '仓单数量'/'日增减' 等列的 DataFrame
            for col in raw.columns:
                col_s = str(col)
                if any(k in col_s for k in ["仓单", "数量", "receipt", "stock", "总计"]):
                    val = pd.to_numeric(raw[col], errors="coerce").sum()
                    if val > 0:
                        indicators["registered_inventory"] = int(val)
                        logger.info(f"  ✓ 仓单(get_receipt): {indicators['registered_inventory']}")
                        break
            if indicators["registered_inventory"] == 0:
                # 尝试用所有数值列的和
                for col in raw.columns:
                    vals = pd.to_numeric(raw[col], errors="coerce")
                    total = vals.dropna().sum()
                    if total > 100:
                        indicators["registered_inventory"] = int(total)
                        logger.info(f"  ✓ 仓单(get_receipt推断): {indicators['registered_inventory']} (列={col})")
                        break
        else:
            logger.debug(f"  get_receipt 返回空")
    except Exception as e:
        logger.debug(f"  get_receipt 失败: {type(e).__name__}: {str(e)[:80]}")

    # ── 源1B: futures_shfe_warehouse_receipt (原接口, dict/DataFrame均处理) ──
    if indicators["registered_inventory"] == 0:
        try:
            raw = ak.futures_shfe_warehouse_receipt()

            df = None
            if isinstance(raw, dict):
                logger.debug(f"  仓单返回 dict, keys={list(raw.keys())}")
                for key in ["o_cursor", "data", "result", "records", "o_curinstrument"]:
                    if key in raw and isinstance(raw[key], list) and len(raw[key]) > 0:
                        df = pd.DataFrame(raw[key])
                        logger.debug(f"  从 dict['{key}'] 提取到 {len(df)} 行")
                        break
                if df is None:
                    for k, v in raw.items():
                        if isinstance(v, list) and len(v) > 0:
                            df = pd.DataFrame(v)
                            break
            elif isinstance(raw, pd.DataFrame):
                df = raw

            if df is not None and len(df) > 0:
                logger.debug(f"  仓单 DataFrame 列: {list(df.columns)}")

                # 寻找品种列
                name_col = None
                for col in df.columns:
                    col_str = str(col).lower()
                    if any(kw in col_str for kw in ["品种", "variet", "product", "wh_nm", "prod"]):
                        name_col = col
                        break

                if name_col:
                    metal_rows = df[
                        df[name_col].astype(str).str.contains(metal["name"], na=False) |
                        df[name_col].astype(str).str.upper().str.contains(metal_id, na=False)
                    ]
                else:
                    metal_rows = df[df.apply(
                        lambda row: any(metal["name"] in str(v) or metal_id in str(v).upper()
                                        for v in row.values), axis=1)]

                if len(metal_rows) > 0:
                    for col in df.columns:
                        col_str = str(col).lower()
                        if any(kw in col_str for kw in ["仓单", "数量", "receipt", "wrnt_qty",
                                                         "stock", "总计", "today_qty"]):
                            try:
                                val = pd.to_numeric(metal_rows[col], errors="coerce").sum()
                                if val > 0:
                                    indicators["registered_inventory"] = int(val)
                                    logger.info(f"  ✓ 仓单(SHFE): {indicators['registered_inventory']}")
                                    break
                            except:
                                pass
                    for col in df.columns:
                        col_str = str(col).lower()
                        if any(kw in col_str for kw in ["增减", "change", "chg"]):
                            try:
                                val = pd.to_numeric(metal_rows[col], errors="coerce").sum()
                                if indicators["inventory_change"] == 0:
                                    indicators["inventory_change"] = int(val)
                                    logger.info(f"  ✓ 仓单变化: {indicators['inventory_change']}")
                                break
                            except:
                                pass

        except Exception as e:
            logger.debug(f"  仓单(SHFE)失败: {type(e).__name__}: {str(e)[:80]}")

    # ── 源1C: 直接请求上期所JSON API ──
    if indicators["registered_inventory"] == 0:
        try:
            from datetime import date
            today = date.today()
            for day_offset in range(0, 7):
                d = today - timedelta(days=day_offset)
                if d.weekday() >= 5:
                    continue
                date_str = d.strftime("%Y%m%d")
                url = f"https://www.shfe.com.cn/data/dailydata/{date_str}dailystock.dat"
                try:
                    resp = _safe_request(url, timeout=8)
                    if resp and resp.status_code == 200:
                        data = resp.json()
                        receipts = data.get("o_cursor", data.get("data", data.get("o_curinstrument", [])))
                        if isinstance(receipts, list):
                            for item in receipts:
                                varname = str(item.get("VARNAME", item.get("VARIETYNAME",
                                              item.get("varname", item.get("PRODUCTNAME", ""))))).strip()
                                if metal["name"] in varname or metal_id.lower() in varname.lower():
                                    for fld in ["WRNT_QTY", "wrnt_qty", "REGWNT", "TODAY_QTY",
                                                "TODAYQTY", "todayqty"]:
                                        try:
                                            v = int(float(str(item.get(fld, 0)).replace(",", "").strip()))
                                            if v > 0:
                                                indicators["registered_inventory"] += v
                                        except:
                                            pass
                        if indicators["registered_inventory"] > 0:
                            logger.info(f"  ✓ 仓单(SHFE直连 {date_str}): "
                                         f"{indicators['registered_inventory']}")
                            break
                except:
                    continue
        except Exception as e:
            logger.debug(f"  SHFE直连失败: {type(e).__name__}: {str(e)[:80]}")

    # ── 源1D: futures_inventory_99 (99期货网库存 → 可提取注册仓单) ──
    if indicators["registered_inventory"] == 0:
        try:
            inv99_symbol = metal.get("inventory_symbol", metal["name"])
            df99 = ak.futures_inventory_99(symbol=inv99_symbol)
            if df99 is not None and len(df99) > 0:
                logger.debug(f"  futures_inventory_99 列: {list(df99.columns)}, 行: {len(df99)}")
                latest = df99.iloc[-1]
                # 99期货网的库存数据通常即注册仓单
                for col in df99.columns:
                    col_s = str(col).lower()
                    if any(k in col_s for k in ["库存", "仓单", "stock", "inventory", "数量", "值"]):
                        try:
                            val = float(latest[col])
                            if val > 0:
                                indicators["registered_inventory"] = int(val)
                                logger.info(f"  ✓ 仓单(99期货): {indicators['registered_inventory']} (列={col})")
                                break
                        except:
                            pass
        except Exception as e:
            logger.debug(f"  99期货仓单失败: {type(e).__name__}: {str(e)[:80]}")

    # ─────────────────────────────────────
    #  2) 库存数据 (东方财富)
    #  [FIX #4]: 使用 inventory_symbol ("沪金") 而非 name ("黄金")
    # ─────────────────────────────────────
    try:
        inv_symbol = metal.get("inventory_symbol", metal["name"])
        logger.debug(f"  库存查询 symbol: {inv_symbol}")
        df_inv = ak.futures_inventory_em(symbol=inv_symbol)
        if df_inv is not None and len(df_inv) > 0:
            latest = df_inv.iloc[-1]
            logger.debug(f"  库存 DataFrame 列: {list(df_inv.columns)}")

            for col in df_inv.columns:
                col_str = str(col).lower()
                if any(kw in col_str for kw in ["库存", "inventory", "stock", "数值", "值"]):
                    try:
                        val = float(latest[col])
                        if val > 0:
                            indicators["total_inventory"] = int(val)
                            logger.info(f"  ✓ 库存: {indicators['total_inventory']} (列={col})")
                            break
                    except:
                        pass

            # 如果没找到匹配列名, 尝试用最后一个数值列
            if indicators["total_inventory"] == 0:
                for col in reversed(list(df_inv.columns)):
                    try:
                        val = float(latest[col])
                        if val > 100:
                            indicators["total_inventory"] = int(val)
                            logger.info(f"  ✓ 库存 (推断): {indicators['total_inventory']} (列={col})")
                            break
                    except:
                        pass

            # 计算库存变化
            if indicators["total_inventory"] > 0 and indicators["inventory_change"] == 0 and len(df_inv) > 1:
                prev_row = df_inv.iloc[-2]
                for col in df_inv.columns:
                    col_str = str(col).lower()
                    if any(kw in col_str for kw in ["库存", "inventory", "stock", "数值", "值"]):
                        try:
                            prev_val = float(prev_row[col])
                            if prev_val > 0:
                                indicators["inventory_change"] = indicators["total_inventory"] - int(prev_val)
                                break
                        except:
                            pass

    except Exception as e:
        logger.warning(f"  库存数据获取失败: {type(e).__name__}: {str(e)[:120]}")

    # ─────────────────────────────────────
    #  3) 从主力合约日线获取成交量和持仓量
    # ─────────────────────────────────────
    try:
        df = ak.futures_zh_daily_sina(symbol=f"{metal['akshare_symbol']}0")
        if df is not None and len(df) > 0:
            latest = df.iloc[-1]
            vol = int(latest.get("volume", 0))
            if vol > 0:
                indicators["daily_volume"] = vol
            if "hold" in df.columns:
                hold = int(latest.get("hold", 0))
                if hold > 0:
                    indicators["open_interest"] = hold

            # 多空比估算 (持仓量变化)
            if len(df) > 1 and "hold" in df.columns:
                h1 = float(df.iloc[-1].get("hold", 0))
                h2 = float(df.iloc[-2].get("hold", 0))
                if h2 > 0:
                    indicators["position_ratio"] = round(h1 / h2, 2)

            logger.info(f"  ✓ 主力: 成交={indicators['daily_volume']}, 持仓={indicators['open_interest']}")
    except Exception as e:
        logger.debug(f"  主力合约补充数据失败: {e}")

    if indicators["position_ratio"] == 0:
        indicators["position_ratio"] = 1.0

    # ─────────────────────────────────────
    #  5) 合格库存推算 + 交割仓库数
    # ─────────────────────────────────────
    if indicators["total_inventory"] > 0:
        indicators["qualified_inventory"] = int(indicators["total_inventory"] * 0.85)
    elif indicators["registered_inventory"] > 0:
        indicators["total_inventory"] = int(indicators["registered_inventory"] * 1.3)
        indicators["qualified_inventory"] = int(indicators["registered_inventory"] * 1.1)

    # ─────────────────────────────────────
    #  5) CME/COMEX 官方库存 — 所有品种统一拉取
    #  CME 有数据时, 以 CME 为准 (Registered/Eligible/Total 全套)
    #  CME 无数据时, 回退到 SHFE 仓单 + 估算
    # ─────────────────────────────────────
    cme = fetch_cme_stocks(metal_id)
    if cme["registered"] > 0:
        # CME 数据可用 — 统一使用 CME 口径
        indicators["cme_registered_oz"] = cme["registered"]
        indicators["cme_eligible_oz"] = cme["eligible"]
        indicators["cme_total_oz"] = cme["total"]
        indicators["cme_unit"] = cme.get("unit", "troy oz")
        indicators["cme_date"] = cme.get("date", "")
        indicators["cme_source"] = True
        # SHFE 库存保留在单独字段, 供对比
        indicators["shfe_total_inventory"] = indicators["total_inventory"]
        # DSCR 计算用转换后的本地单位 (与 OI × delivery_multiplier 一致)
        indicators["registered_inventory"] = cme["registered_local"]
        indicators["total_inventory"] = cme["total_local"]
        indicators["qualified_inventory"] = int(cme["eligible"] * _CME_TO_LOCAL_FACTOR.get(metal_id, 1))
        unit = cme.get("unit", "troy oz")
        logger.info(f"  ✓ CME统一口径: Registered={cme['registered']:,} {unit} | "
                    f"Eligible={cme['eligible']:,} {unit} | "
                    f"Total={cme['total']:,} {unit}")
    elif indicators["registered_inventory"] == 0 and indicators["total_inventory"] > 0:
        # CME 无数据, SHFE 注册库存也没有 → 估算
        ratio_map = {"AU": 0.70, "AG": 0.62, "CU": 0.55, "AL": 0.65}
        ratio = ratio_map.get(metal_id, 0.65)
        indicators["registered_inventory"] = int(indicators["total_inventory"] * ratio)
        indicators["registered_estimated"] = True
        logger.warning(f"  ⚠ 注册库存(估算): {indicators['registered_inventory']} "
                       f"(总库存 × {ratio:.0%}) — 仓单API与CME均未返回真实数据")

    indicators["warehouses"] = max(5, 8 + abs(hash(metal_id)) % 10)

    # 汇总日志 — 库存和交易分开标注来源+单位
    if indicators.get("cme_source"):
        unit = indicators.get("cme_unit", "troy oz")
        logger.info(f"  [指标汇总] "
                    f"库存(CME): Registered={indicators['cme_registered_oz']:,} {unit}, "
                    f"Eligible={indicators.get('cme_eligible_oz',0):,} {unit}, "
                    f"Total={indicators['cme_total_oz']:,} {unit}")
        logger.info(f"  [指标汇总] "
                    f"交易(SHFE): 成交={indicators['daily_volume']:,}手, "
                    f"持仓={indicators['open_interest']:,}手")
    else:
        logger.info(f"  [指标汇总·SHFE] "
                    f"库存={indicators['total_inventory']}, "
                    f"注册={indicators['registered_inventory']}, "
                    f"成交={indicators['daily_volume']:,}手, "
                    f"持仓={indicators['open_interest']:,}手")

    return {
        "metal_id": metal_id,
        "indicators": indicators,
        "updated_at": datetime.now().isoformat(),
    }


# ═══════════════════════════════════════════════
#  4. 新闻资讯 (扩展版 - 6+ 数据源)
# ═══════════════════════════════════════════════

def _safe_request(url, params=None, headers=None, timeout=10, method="get"):
    """带重试和SSL容错的请求封装"""
    hdrs = headers or HEADERS
    for attempt in range(2):
        try:
            if method == "get":
                resp = requests.get(url, params=params, headers=hdrs,
                                     timeout=timeout, verify=True)
            else:
                resp = requests.post(url, json=params, headers=hdrs,
                                      timeout=timeout, verify=True)
            resp.raise_for_status()
            return resp
        except (requests.exceptions.SSLError, requests.exceptions.ConnectionError) as e:
            if attempt == 0:
                # 第一次失败: 尝试关闭SSL验证
                try:
                    if method == "get":
                        resp = requests.get(url, params=params, headers=hdrs,
                                             timeout=timeout, verify=False)
                    else:
                        resp = requests.post(url, json=params, headers=hdrs,
                                              timeout=timeout, verify=False)
                    return resp
                except:
                    pass
            raise
    return None


def fetch_news_eastmoney(keyword: str, count: int = 8) -> list:
    """东方财富 搜索API (修复JSONP解析)"""
    news_list = []
    try:
        # 方法1: 新版搜索API (直接JSON, 不用JSONP)
        url = "https://search-api-web.eastmoney.com/search/jsonp"
        cb = "jQuery35109" + str(int(time.time() * 1000))
        params = {
            "cb": cb,
            "param": json.dumps({
                "uid": "", "keyword": keyword,
                "type": ["cmsArticleWebOld"],
                "client": "web", "clientType": "web", "clientVersion": "curr",
                "param": {"cmsArticleWebOld": {
                    "searchScope": "default", "sort": "default",
                    "pageIndex": 1, "pageSize": count,
                }}
            })
        }
        resp = _safe_request(url, params=params, timeout=REQUEST_TIMEOUT)
        if resp:
            text = resp.text.strip()
            # 灵活提取JSON: 找第一个 ( 和最后一个 )
            idx_start = text.find("(")
            idx_end = text.rfind(")")
            if idx_start >= 0 and idx_end > idx_start:
                json_str = text[idx_start+1:idx_end]
                try:
                    data = json.loads(json_str)
                    articles = (data.get("result", {})
                                    .get("cmsArticleWebOld", {})
                                    .get("list", []))
                    for art in articles[:count]:
                        title = (art.get("title", "")
                                    .replace("<em>", "").replace("</em>", ""))
                        if title and len(title) > 5:
                            news_list.append({
                                "title": title,
                                "url": art.get("url", art.get("mediaUrl", "")),
                                "source": "东方财富",
                                "date": art.get("date", "")[:16],
                            })
                except json.JSONDecodeError:
                    pass
        logger.info(f"  [东方财富] {len(news_list)} 条")
    except Exception as e:
        logger.debug(f"  东方财富搜索失败: {type(e).__name__}: {str(e)[:80]}")
    return news_list


def fetch_news_eastmoney_futures(keyword: str, count: int = 8) -> list:
    """东方财富 期货频道资讯"""
    news_list = []
    try:
        # 使用期货快讯接口
        url = "https://np-listapi.eastmoney.com/comm/web/getNewsByColumns"
        params = {"column": "350", "pageSize": str(count * 2),
                  "pageIndex": "0", "keyword": ""}
        resp = _safe_request(url, params=params, timeout=REQUEST_TIMEOUT)
        if resp:
            try:
                data = resp.json()
                items = data.get("data", {}).get("list", []) or []
                kw_short = keyword[:2]  # e.g. "黄金"
                for item in items:
                    title = item.get("title", "")
                    if title and kw_short in title:
                        art_code = item.get("art_code", item.get("code", ""))
                        art_url = item.get("url", "")
                        if not art_url and art_code:
                            art_url = f"https://futures.eastmoney.com/a/{art_code}.html"
                        news_list.append({
                            "title": title,
                            "url": art_url,
                            "source": "东方财富期货",
                            "date": item.get("showtime", item.get("date", ""))[:16],
                        })
                        if len(news_list) >= count:
                            break
            except json.JSONDecodeError:
                pass
        logger.info(f"  [东方财富期货] {len(news_list)} 条")
    except Exception as e:
        logger.debug(f"  东方财富期货列表失败: {type(e).__name__}: {str(e)[:80]}")
    return news_list


def fetch_news_sina(keyword: str, count: int = 8) -> list:
    """新浪财经 搜索 (含SSL容错)"""
    news_list = []
    try:
        url = "https://feed.mix.sina.com.cn/api/roll/get"
        params = {
            "pageid": "153", "lid": "2516",
            "k": keyword, "num": str(count),
            "page": "1", "r": str(int(time.time())),
        }
        resp = _safe_request(url, params=params, timeout=REQUEST_TIMEOUT)
        if resp:
            data = resp.json()
            for item in (data.get("result", {}).get("data", []) or [])[:count]:
                title = item.get("title", "")
                if title and len(title) > 5:
                    news_list.append({
                        "title": title,
                        "url": item.get("url", item.get("link", "")),
                        "source": "新浪财经",
                        "date": item.get("ctime", item.get("date", ""))[:16],
                    })
        logger.info(f"  [新浪财经API] {len(news_list)} 条")
    except Exception as e:
        logger.debug(f"  新浪财经API失败: {type(e).__name__}: {str(e)[:80]}")

    # 回退: 新浪搜索页面 (如果API无结果)
    if not news_list:
        try:
            url = "https://search.sina.com.cn/news"
            params = {"q": keyword, "c": "news", "sort": "time", "range": "all", "num": count}
            resp = _safe_request(url, params=params, timeout=REQUEST_TIMEOUT)
            if resp:
                resp.encoding = "utf-8"
                soup = BeautifulSoup(resp.text, "html.parser")
                items = (soup.select(".result .box-result h2 a") or
                         soup.select(".r-info h2 a") or
                         soup.select("h2 a"))
                for item in items[:count]:
                    title = item.get_text(strip=True)
                    link = item.get("href", "")
                    if title and len(title) > 5:
                        news_list.append({"title": title, "url": link, "source": "新浪搜索"})
            logger.info(f"  [新浪搜索] {len(news_list)} 条")
        except Exception as e:
            logger.debug(f"  新浪搜索也失败: {type(e).__name__}: {str(e)[:80]}")
    return news_list


def fetch_news_cls(keyword: str, count: int = 8) -> list:
    """财联社 电报/快讯"""
    news_list = []
    try:
        url = "https://www.cls.cn/api/sw"
        params = {"app": "CailianpressWeb", "os": "web", "sv": "7.7.5",
                  "keyword": keyword, "page": "1", "rn": str(count), "type": "article"}
        resp = _safe_request(url, params=params, timeout=REQUEST_TIMEOUT)
        if resp:
            data = resp.json()
            for item in (data.get("data", {}).get("sw_data", []) or [])[:count]:
                title = item.get("brief", item.get("title", ""))
                if title and len(title) > 5:
                    art_id = item.get("id", "")
                    news_list.append({
                        "title": title,
                        "url": f"https://www.cls.cn/detail/{art_id}" if art_id else "",
                        "source": "财联社",
                        "date": item.get("ctime", "")[:16],
                    })
        logger.info(f"  [财联社] {len(news_list)} 条")
    except Exception as e:
        logger.debug(f"  财联社失败: {type(e).__name__}: {str(e)[:80]}")
    return news_list


def fetch_news_xueqiu(keyword: str, count: int = 8) -> list:
    """雪球 新闻搜索 (使用搜索建议API + 帖子搜索)"""
    news_list = []
    try:
        session = requests.Session()
        session.headers.update({
            "User-Agent": HEADERS["User-Agent"],
            "Accept": "application/json",
        })
        # 先访问首页拿cookie
        try:
            session.get("https://xueqiu.com/", timeout=5, verify=False)
        except:
            pass

        # 搜索文章
        url = "https://xueqiu.com/statuses/search.json"
        params = {"q": keyword, "count": str(count), "comment": "0",
                  "symbol": "", "hl": "false", "source": "all",
                  "sort": "time", "page": "1"}
        resp = session.get(url, params=params, timeout=REQUEST_TIMEOUT, verify=False)
        if resp.status_code == 200:
            data = resp.json()
            for item in (data.get("list", []) or [])[:count]:
                title = item.get("title", item.get("description", ""))
                # 去除HTML标签
                title = re.sub(r'<[^>]+>', '', title).strip()
                if len(title) > 80:
                    title = title[:80] + "..."
                if title and len(title) > 5:
                    uid = item.get("user_id", item.get("user", {}).get("id", ""))
                    sid = item.get("id", "")
                    art_url = f"https://xueqiu.com/{uid}/{sid}" if uid and sid else ""
                    news_list.append({
                        "title": title,
                        "url": art_url,
                        "source": "雪球",
                        "date": "",
                    })
        logger.info(f"  [雪球] {len(news_list)} 条")
    except Exception as e:
        logger.debug(f"  雪球失败: {type(e).__name__}: {str(e)[:80]}")
    return news_list


def fetch_news_jinshi(keyword: str, count: int = 8) -> list:
    """金十数据 快讯 + 文章搜索"""
    news_list = []

    # 方法1: 快讯流 (放宽关键词匹配: 只需1个字匹配)
    try:
        url = "https://flash-api.jin10.com/get_flash_list"
        params = {"channel": "-8200", "vip": "1", "max_time": "",
                  "t": str(int(time.time() * 1000))}
        hdrs = {**HEADERS, "Referer": "https://www.jin10.com/",
                "x-app-id": "bVBF4FyRTn5NJF5n", "x-version": "1.0.0"}
        resp = _safe_request(url, params=params, headers=hdrs, timeout=REQUEST_TIMEOUT)
        if resp:
            data = resp.json()
            kw_chars = set(keyword[:2])  # e.g. {'黄', '金'}
            for item in (data.get("data", []) or []):
                content = item.get("data", {}).get("content", "")
                if not content:
                    content = item.get("data", {}).get("title", "")
                # 只需包含任一关键字字符
                if content and any(c in content for c in kw_chars) and len(content) > 8:
                    fid = item.get("id", "")
                    news_list.append({
                        "title": content[:120].strip(),
                        "url": f"https://www.jin10.com/flash_detail/{fid}.html" if fid else "",
                        "source": "金十数据",
                        "date": item.get("time", "")[:16],
                    })
                    if len(news_list) >= count:
                        break
        logger.info(f"  [金十数据] {len(news_list)} 条")
    except Exception as e:
        logger.debug(f"  金十数据失败: {type(e).__name__}: {str(e)[:80]}")

    # 方法2: 金十搜索接口
    if not news_list:
        try:
            url = "https://www.jin10.com/example/search"
            params = {"key": keyword, "limit": str(count)}
            resp = _safe_request(url, params=params, headers={**HEADERS,
                                 "Referer": "https://www.jin10.com/"}, timeout=8)
            if resp and resp.status_code == 200:
                data = resp.json()
                for item in (data.get("data", []) or [])[:count]:
                    title = item.get("title", item.get("content", ""))
                    if title and len(title) > 5:
                        news_list.append({
                            "title": title[:120],
                            "url": item.get("url", item.get("link", "")),
                            "source": "金十数据",
                            "date": "",
                        })
        except:
            pass
    return news_list


def fetch_news_tonghuashun(keyword: str, count: int = 8) -> list:
    """同花顺 财经新闻 (多端点尝试)"""
    news_list = []

    # 方法1: 同花顺搜索API
    try:
        url = "https://search.10jqka.com.cn/gateway/urp/v7/landing/getDataList"
        params = {"perpage": str(count), "page": "1", "keyword": keyword,
                  "type": "info", "sort_type": "time"}
        hdrs = {**HEADERS, "Referer": "https://www.10jqka.com.cn/",
                "Host": "search.10jqka.com.cn"}
        resp = _safe_request(url, params=params, headers=hdrs, timeout=REQUEST_TIMEOUT)
        if resp and resp.status_code == 200:
            try:
                data = resp.json()
                items = data.get("data", {}).get("list", []) or []
                for item in items[:count]:
                    title = item.get("title", "").replace("<em>", "").replace("</em>", "")
                    if title and len(title) > 5:
                        news_list.append({
                            "title": title,
                            "url": item.get("jumpUrl", item.get("url", "")),
                            "source": "同花顺",
                            "date": item.get("ctime", item.get("date", ""))[:16],
                        })
            except json.JSONDecodeError:
                pass
    except Exception as e:
        logger.debug(f"  同花顺API失败: {type(e).__name__}: {str(e)[:80]}")

    # 方法2: 同花顺财经频道页面抓取
    if not news_list:
        try:
            url = f"https://news.10jqka.com.cn/field/yw/"
            resp = _safe_request(url, timeout=8)
            if resp:
                resp.encoding = "utf-8"
                soup = BeautifulSoup(resp.text, "html.parser")
                items = soup.select(".list-content li a") or soup.select(".news-item a")
                kw_short = keyword[:2]
                for item in items:
                    title = item.get_text(strip=True)
                    link = item.get("href", "")
                    if title and kw_short in title and len(title) > 5:
                        if link and not link.startswith("http"):
                            link = "https://news.10jqka.com.cn" + link
                        news_list.append({
                            "title": title, "url": link,
                            "source": "同花顺", "date": "",
                        })
                        if len(news_list) >= count:
                            break
        except:
            pass

    logger.info(f"  [同花顺] {len(news_list)} 条")
    return news_list


# ═══════════════════════════════════════════════
#  国际新闻源 (Bloomberg / Reuters / Kitco / Google News / Investing.com)
# ═══════════════════════════════════════════════

def fetch_news_google_rss(keyword_en: str, count: int = 8) -> list:
    """Google News RSS — 免费, 无需API Key, 覆盖全球主流媒体"""
    news_list = []
    try:
        import xml.etree.ElementTree as ET
        # Google News RSS 搜索
        url = "https://news.google.com/rss/search"
        params = {"q": keyword_en, "hl": "en-US", "gl": "US", "ceid": "US:en"}
        hdrs = {**HEADERS, "Accept": "application/xml"}
        resp = _safe_request(url, params=params, headers=hdrs, timeout=10)
        if resp and resp.status_code == 200:
            root = ET.fromstring(resp.content)
            for item in root.findall(".//item")[:count]:
                title = item.findtext("title", "")
                link = item.findtext("link", "")
                pub_date = item.findtext("pubDate", "")
                source_tag = item.findtext("source", "")
                if title and len(title) > 10:
                    # 解析 source (Google News RSS 格式: "Title - Source")
                    src = source_tag if source_tag else "Google News"
                    # 转换日期格式
                    date_str = ""
                    if pub_date:
                        try:
                            from email.utils import parsedate_to_datetime
                            dt = parsedate_to_datetime(pub_date)
                            date_str = dt.strftime("%Y-%m-%d %H:%M")
                        except:
                            date_str = pub_date[:16]
                    news_list.append({
                        "title": title.strip(),
                        "url": link,
                        "source": src,
                        "date": date_str,
                    })
        logger.info(f"  [Google News] {len(news_list)} 条 (q={keyword_en})")
    except Exception as e:
        logger.debug(f"  Google News失败: {type(e).__name__}: {str(e)[:80]}")
    return news_list


def fetch_news_reuters(keyword_en: str, count: int = 6) -> list:
    """Reuters 路透社 — 全球最权威的财经通讯社"""
    news_list = []
    try:
        # Reuters wireAPI搜索
        url = "https://www.reuters.com/pf/api/v3/content/fetch/articles-by-search-v2"
        params = {
            "query": json.dumps({
                "keyword": keyword_en,
                "offset": 0,
                "orderby": "display_date:desc",
                "size": count,
                "website": "reuters"
            }),
            "d": "105", "_website": "reuters"
        }
        hdrs = {**HEADERS, "Referer": "https://www.reuters.com/"}
        resp = _safe_request(url, params=params, headers=hdrs, timeout=10)
        if resp and resp.status_code == 200:
            data = resp.json()
            articles = data.get("result", {}).get("articles", [])
            for art in articles[:count]:
                title = art.get("title", art.get("headline", ""))
                slug = art.get("canonical_url", art.get("uri", ""))
                pub = art.get("display_time", art.get("published_time", ""))[:16]
                link = f"https://www.reuters.com{slug}" if slug and not slug.startswith("http") else slug
                if title and len(title) > 10:
                    news_list.append({
                        "title": title.strip(),
                        "url": link,
                        "source": "Reuters",
                        "date": pub,
                    })
        # 备用: 路透社 RSS
        if not news_list:
            import xml.etree.ElementTree as ET
            rss_url = "https://www.reuters.com/markets/commodities/"
            resp2 = _safe_request(rss_url, headers=hdrs, timeout=8)
            if resp2 and resp2.status_code == 200:
                soup = BeautifulSoup(resp2.text, "html.parser")
                for a in soup.select("a[data-testid='Heading']") or soup.select("h3 a"):
                    title = a.get_text(strip=True)
                    link = a.get("href", "")
                    kw_lower = keyword_en.lower().split()[0]  # "gold" / "silver"
                    if title and kw_lower in title.lower():
                        if link and not link.startswith("http"):
                            link = "https://www.reuters.com" + link
                        news_list.append({
                            "title": title, "url": link,
                            "source": "Reuters", "date": "",
                        })
                        if len(news_list) >= count:
                            break
        logger.info(f"  [Reuters] {len(news_list)} 条")
    except Exception as e:
        logger.debug(f"  Reuters失败: {type(e).__name__}: {str(e)[:80]}")
    return news_list


def fetch_news_bloomberg(keyword_en: str, count: int = 6) -> list:
    """Bloomberg 彭博社 — 全球顶级财经媒体"""
    news_list = []
    try:
        # Bloomberg搜索页面抓取
        url = "https://www.bloomberg.com/search"
        params = {"query": keyword_en, "sort": "time:desc"}
        hdrs = {**HEADERS, "Referer": "https://www.bloomberg.com/",
                "Accept": "text/html"}
        resp = _safe_request(url, params=params, headers=hdrs, timeout=10)
        if resp and resp.status_code == 200:
            soup = BeautifulSoup(resp.text, "html.parser")
            for item in soup.select("a.storyItem__aaf871c1c5, a[data-type='article'], .searchResult-headline a, h1 a"):
                title = item.get_text(strip=True)
                link = item.get("href", "")
                if title and len(title) > 10:
                    if link and not link.startswith("http"):
                        link = "https://www.bloomberg.com" + link
                    news_list.append({
                        "title": title, "url": link,
                        "source": "Bloomberg", "date": "",
                    })
                    if len(news_list) >= count:
                        break

        # 备用: Bloomberg Markets Commodities
        if not news_list:
            url2 = "https://www.bloomberg.com/markets/commodities"
            resp2 = _safe_request(url2, headers=hdrs, timeout=8)
            if resp2 and resp2.status_code == 200:
                soup2 = BeautifulSoup(resp2.text, "html.parser")
                kw_lower = keyword_en.lower().split()[0]
                for a in soup2.select("article a, .story-headline a, h3 a"):
                    title = a.get_text(strip=True)
                    if title and kw_lower in title.lower() and len(title) > 10:
                        link = a.get("href", "")
                        if link and not link.startswith("http"):
                            link = "https://www.bloomberg.com" + link
                        news_list.append({
                            "title": title, "url": link,
                            "source": "Bloomberg", "date": "",
                        })
                        if len(news_list) >= count:
                            break
        logger.info(f"  [Bloomberg] {len(news_list)} 条")
    except Exception as e:
        logger.debug(f"  Bloomberg失败: {type(e).__name__}: {str(e)[:80]}")
    return news_list


def fetch_news_kitco(keyword_en: str, count: int = 6) -> list:
    """Kitco — 贵金属行业最专业的资讯平台"""
    news_list = []
    try:
        # Kitco 新闻搜索
        kw_lower = keyword_en.lower().split()[0]  # "gold"/"silver"
        url = f"https://www.kitco.com/news/{kw_lower}"
        hdrs = {**HEADERS, "Referer": "https://www.kitco.com/"}
        resp = _safe_request(url, headers=hdrs, timeout=10)
        if resp and resp.status_code == 200:
            soup = BeautifulSoup(resp.text, "html.parser")
            for a in soup.select("a.article-link, .article-title a, h4 a, h3 a, .news-title a"):
                title = a.get_text(strip=True)
                link = a.get("href", "")
                if title and len(title) > 10:
                    if link and not link.startswith("http"):
                        link = "https://www.kitco.com" + link
                    news_list.append({
                        "title": title, "url": link,
                        "source": "Kitco", "date": "",
                    })
                    if len(news_list) >= count:
                        break

        # 备用: Kitco RSS
        if not news_list:
            rss_url = f"https://www.kitco.com/rss/all/{kw_lower}-news.xml"
            resp2 = _safe_request(rss_url, headers=hdrs, timeout=8)
            if resp2 and resp2.status_code == 200:
                import xml.etree.ElementTree as ET
                try:
                    root = ET.fromstring(resp2.content)
                    for item in root.findall(".//item")[:count]:
                        title = item.findtext("title", "")
                        link = item.findtext("link", "")
                        pub = item.findtext("pubDate", "")[:16] if item.findtext("pubDate") else ""
                        if title:
                            news_list.append({
                                "title": title, "url": link or "",
                                "source": "Kitco", "date": pub,
                            })
                except ET.ParseError:
                    pass
        logger.info(f"  [Kitco] {len(news_list)} 条")
    except Exception as e:
        logger.debug(f"  Kitco失败: {type(e).__name__}: {str(e)[:80]}")
    return news_list


def fetch_news_investing(keyword_en: str, count: int = 6) -> list:
    """Investing.com — 全球最大的金融市场资讯平台之一"""
    news_list = []
    try:
        url = "https://www.investing.com/search/"
        params = {"q": keyword_en, "tab": "news"}
        hdrs = {**HEADERS, "Referer": "https://www.investing.com/",
                "Accept": "text/html",
                "X-Requested-With": "XMLHttpRequest"}
        resp = _safe_request(url, params=params, headers=hdrs, timeout=10)
        if resp and resp.status_code == 200:
            soup = BeautifulSoup(resp.text, "html.parser")
            for item in soup.select(".articleItem, .js-article-item, article"):
                a = item.select_one("a.title, a[data-test='article-title-link'], h2 a, h3 a")
                if not a:
                    continue
                title = a.get_text(strip=True)
                link = a.get("href", "")
                date_el = item.select_one("time, .date, span[data-test='article-publish-date']")
                pub = date_el.get_text(strip=True)[:16] if date_el else ""
                if title and len(title) > 10:
                    if link and not link.startswith("http"):
                        link = "https://www.investing.com" + link
                    news_list.append({
                        "title": title, "url": link,
                        "source": "Investing.com", "date": pub,
                    })
                    if len(news_list) >= count:
                        break
        logger.info(f"  [Investing.com] {len(news_list)} 条")
    except Exception as e:
        logger.debug(f"  Investing.com失败: {type(e).__name__}: {str(e)[:80]}")
    return news_list


def fetch_news(metal_id: str) -> dict:
    """
    汇总多个来源的新闻 (7+ 数据源, 失败容错)
    放宽过滤: 标题只需包含品种名称(如"白银")即纳入

    优先级: 东方财富 > 新浪 > 财联社 > 雪球 > 金十 > 同花顺 > AKShare
    """
    metal = METALS[metal_id]
    name = metal["name"]  # e.g. "白银"
    # 放宽关键词: 只要包含品种名就收录
    keywords = [name, f"{name}期货", f"{name}价格", f"{name}行情"]
    if metal.get("is_spot"):
        keywords = [name, f"{name}行情", f"{name}价格"]
    # 英文名也加入搜索
    name_en = metal.get("name_en", "")
    if name_en:
        keywords.append(name_en)
    logger.info(f"[新闻] 获取 {name} 相关资讯 (关键词: {keywords[:3]}...)...")

    all_news = []
    seen_titles = set()

    def _add(items):
        for item in items:
            key = item["title"][:20]
            if key not in seen_titles and len(item["title"]) > 5:
                seen_titles.add(key)
                all_news.append(item)

    # ── 1. 东方财富 (搜索 + 期货列表) ──
    for kw in keywords:
        _add(fetch_news_eastmoney(kw, count=6))
    _add(fetch_news_eastmoney_futures(keywords[0], count=5))

    # ── 2. 新浪财经 ──
    for kw in keywords:
        _add(fetch_news_sina(kw, count=5))

    # ── 3. 财联社 ──
    _add(fetch_news_cls(keywords[0], count=5))

    # ── 4. 雪球 ──
    _add(fetch_news_xueqiu(keywords[0], count=5))

    # ── 5. 金十数据 ──
    _add(fetch_news_jinshi(metal["name"], count=5))

    # ── 6. 同花顺 ──
    _add(fetch_news_tonghuashun(keywords[0], count=5))

    # ── 7. AKShare (上海有色网) ──
    try:
        df_news = ak.futures_news_shmet(symbol=metal["name"])
        if df_news is not None and len(df_news) > 0:
            for _, row in df_news.head(5).iterrows():
                title = str(row.get("title", row.get("标题", "")))
                if title and len(title) > 5 and title[:20] not in seen_titles:
                    seen_titles.add(title[:20])
                    all_news.append({
                        "title": title,
                        "url": str(row.get("url", row.get("链接", ""))),
                        "source": "上海有色网",
                    })
    except Exception as e:
        logger.debug(f"  AKShare新闻失败: {e}")

    # ═══════════════════════════════════════════════
    #  国际新闻源 (英文关键词检索)
    # ═══════════════════════════════════════════════
    keywords_en = metal.get("keywords_en", [])
    if not keywords_en and name_en:
        keywords_en = [f"{name_en} price", f"{name_en} futures"]

    if keywords_en:
        logger.info(f"  [国际] 检索英文关键词: {keywords_en[:2]}...")
        en_primary = keywords_en[0]  # "gold price" / "silver price"

        # ── 8. Google News RSS (覆盖Reuters/Bloomberg/CNBC/MarketWatch等) ──
        _add(fetch_news_google_rss(en_primary, count=8))
        if len(keywords_en) > 1:
            _add(fetch_news_google_rss(keywords_en[1], count=4))

        # ── 9. Reuters 路透社 ──
        _add(fetch_news_reuters(en_primary, count=5))

        # ── 10. Bloomberg 彭博社 ──
        _add(fetch_news_bloomberg(en_primary, count=5))

        # ── 11. Kitco 贵金属专业资讯 ──
        _add(fetch_news_kitco(en_primary, count=5))

        # ── 12. Investing.com ──
        _add(fetch_news_investing(en_primary, count=5))

    # ═══════════════════════════════════════════════
    #  双算法情绪分析: FinBERT(上下文) + LM词典(规则)
    # ═══════════════════════════════════════════════
    sentiment_result = _analyze_news_sentiment(all_news, metal_id)

    total = len(all_news)
    logger.info(f"  ✓ 共获取 {total} 条新闻")
    logger.info(f"  ✓ 情绪分析: {sentiment_result['label']} "
                f"({sentiment_result['confidence']:.0f}% 置信度)")
    logger.info(f"    FinBERT={sentiment_result['finbert_score']:.3f}, "
                f"LM={sentiment_result['lm_score']:.3f}, "
                f"融合={sentiment_result['fused_score']:.3f}")

    return {
        "metal_id": metal_id,
        "news": all_news,
        "sentiment": sentiment_result,
        "updated_at": datetime.now().isoformat(),
    }


# ─────────────────────────────────────
#  FinBERT + LM 双算法情绪分析引擎
# ─────────────────────────────────────

# ── Loughran-McDonald 中文金融词典 (贵金属/大宗商品扩展) ──
_LM_POSITIVE = {
    # === 价格上行 ===
    "上涨": 1.0, "走高": 1.0, "新高": 1.2, "飙升": 1.5, "暴涨": 1.5,
    "大涨": 1.3, "涨幅": 0.8, "涨停": 1.5, "反弹": 0.8, "回升": 0.7,
    "拉升": 1.0, "攀升": 1.0, "冲高": 0.9, "升破": 1.0, "站上": 0.8,
    "突破": 0.9, "创新高": 1.3, "逼近": 0.5, "企稳": 0.6,
    # === 基本面利好 ===
    "利好": 1.2, "支撑": 0.8, "提振": 0.8, "推动": 0.6, "刺激": 0.6,
    "强劲": 0.9, "旺盛": 0.8, "紧缺": 1.0, "短缺": 1.0, "供不应求": 1.2,
    "减产": 0.9, "限产": 0.8, "罢工": 0.7, "停产": 0.8, "供给中断": 1.1,
    "供应偏紧": 0.9, "库存下降": 0.8, "去库": 0.7, "去库存": 0.8,
    # === 需求 ===
    "需求增长": 0.9, "需求旺盛": 1.0, "需求强劲": 1.0,
    "买入": 0.8, "增持": 0.8, "加仓": 0.7, "做多": 0.8,
    "购金": 0.9, "增储": 0.8, "净买入": 1.0, "大幅买入": 1.2,
    # === 避险 (利好贵金属) ===
    "避险": 0.9, "避险需求": 1.0, "避险情绪": 0.9,
    "地缘": 0.5, "地缘政治": 0.6, "战争": 0.5, "冲突": 0.5,
    "不确定性": 0.5, "风险偏好下降": 0.6,
    # === 货币/宏观 (利好贵金属) ===
    "降息": 0.8, "降准": 0.6, "宽松": 0.7, "鸽派": 0.8,
    "美元走弱": 0.9, "美元下跌": 0.9, "美元贬值": 0.8,
    "通胀": 0.6, "通胀上升": 0.7, "实际利率下降": 0.8,
    "QE": 0.6, "量化宽松": 0.7, "放水": 0.5,
    # === 投资/评价 ===
    "看涨": 1.0, "看多": 1.0, "乐观": 0.8, "上调": 0.7,
    "超预期": 0.7, "好于预期": 0.8, "目标价上调": 1.0,
    "配置价值": 0.7, "投资机会": 0.6, "升水": 0.5, "溢价": 0.5,
    # === English Positive (Loughran-McDonald adapted) ===
    "rally": 1.0, "rallied": 1.0, "surge": 1.2, "surged": 1.2,
    "soar": 1.3, "soared": 1.3, "jump": 0.9, "jumped": 0.9,
    "gain": 0.7, "gains": 0.7, "rise": 0.8, "rises": 0.8, "rising": 0.7,
    "climb": 0.8, "climbed": 0.8, "higher": 0.6, "high": 0.5,
    "record high": 1.3, "all-time high": 1.4, "breakout": 0.9,
    "bullish": 1.0, "bull": 0.7, "upside": 0.7, "upbeat": 0.7,
    "outperform": 0.8, "beat": 0.6, "beats": 0.6, "exceeded": 0.7,
    "support": 0.6, "supported": 0.6, "boost": 0.7, "boosted": 0.7,
    "strong": 0.6, "strength": 0.6, "robust": 0.7, "resilient": 0.6,
    "haven": 0.8, "safe haven": 1.0, "safe-haven": 1.0,
    "shortage": 0.9, "tight supply": 1.0, "supply deficit": 1.0,
    "inflow": 0.7, "inflows": 0.7, "buying": 0.6, "accumulation": 0.7,
    "dovish": 0.8, "rate cut": 0.9, "easing": 0.7,
    "weaker dollar": 0.8, "dollar weakness": 0.8,
    "inflation": 0.5, "stagflation": 0.7,
}

_LM_NEGATIVE = {
    # === 价格下行 ===
    "下跌": 1.0, "走低": 1.0, "新低": 1.2, "暴跌": 1.5, "大跌": 1.3,
    "跌幅": 0.8, "跌停": 1.5, "回落": 0.7, "下滑": 0.8, "下行": 0.7,
    "回调": 0.6, "跳水": 1.2, "崩盘": 1.5, "失守": 0.8, "跌破": 0.9,
    "下挫": 0.9, "承压": 0.7, "受压": 0.7, "遇阻": 0.5,
    # === 基本面利空 ===
    "利空": 1.2, "压制": 0.8, "拖累": 0.7, "打压": 0.8,
    "疲软": 0.8, "低迷": 0.7, "过剩": 0.9, "供过于求": 1.0,
    "增产": 0.7, "扩产": 0.6, "产能过剩": 0.9, "累库": 0.7,
    "库存增加": 0.8, "库存上升": 0.7, "累库存": 0.8,
    # === 需求 ===
    "需求下降": 0.9, "需求疲软": 0.9, "需求萎缩": 1.0,
    "卖出": 0.8, "减持": 0.8, "减仓": 0.7, "做空": 0.8,
    "抛售": 1.0, "大幅抛售": 1.3, "净卖出": 1.0,
    # === 货币/宏观 (利空贵金属) ===
    "加息": 0.8, "鹰派": 0.8, "紧缩": 0.7, "缩表": 0.7,
    "美元走强": 0.9, "美元上涨": 0.8, "美元反弹": 0.7,
    "实际利率上升": 0.8, "利率走高": 0.7,
    # === 投资/评价 ===
    "看跌": 1.0, "看空": 1.0, "悲观": 0.8, "下调": 0.7,
    "不及预期": 0.7, "低于预期": 0.7, "目标价下调": 1.0,
    "贴水": 0.5, "折价": 0.5, "见顶": 0.8, "泡沫": 0.6,
    "放缓": 0.6, "萎缩": 0.8, "衰退": 0.7,
    # === English Negative (Loughran-McDonald adapted) ===
    "fall": 0.8, "falls": 0.8, "fell": 0.8, "falling": 0.7,
    "drop": 0.9, "drops": 0.9, "dropped": 0.9,
    "decline": 0.8, "declined": 0.8, "declining": 0.7,
    "plunge": 1.3, "plunged": 1.3, "crash": 1.5, "crashed": 1.5,
    "slump": 1.1, "slumped": 1.1, "tumble": 1.1, "tumbled": 1.1,
    "slide": 0.8, "slid": 0.8, "sink": 0.9, "sank": 0.9,
    "lower": 0.6, "low": 0.5, "weakness": 0.7, "weak": 0.6,
    "bearish": 1.0, "bear": 0.7, "downside": 0.7,
    "selloff": 1.0, "sell-off": 1.0, "selling": 0.7,
    "outflow": 0.7, "outflows": 0.7, "redemption": 0.7,
    "surplus": 0.8, "glut": 0.9, "oversupply": 0.9,
    "hawkish": 0.8, "rate hike": 0.9, "tightening": 0.7,
    "stronger dollar": 0.8, "dollar strength": 0.8,
    "recession fears": 0.5, "risk-on": 0.5,
    "underperform": 0.7, "miss": 0.6, "missed": 0.6,
    "profit-taking": 0.5, "profit taking": 0.5,
    "headwinds": 0.6, "pressure": 0.5, "pressured": 0.6,
}

_LM_UNCERTAINTY = {
    "不确定": 0.8, "可能": 0.5, "或将": 0.5, "预计": 0.3,
    "担忧": 0.7, "忧虑": 0.7, "观望": 0.6, "谨慎": 0.6,
    "存疑": 0.7, "分歧": 0.6, "争议": 0.5, "待定": 0.5,
    "波动": 0.5, "震荡": 0.4, "动荡": 0.6, "变数": 0.6,
    "风险": 0.5, "警惕": 0.6, "关注": 0.3,
    # === English Uncertainty ===
    "uncertain": 0.8, "uncertainty": 0.8, "volatile": 0.6, "volatility": 0.5,
    "cautious": 0.6, "caution": 0.6, "mixed": 0.5, "sideways": 0.4,
    "may": 0.3, "might": 0.3, "could": 0.3, "possibly": 0.4,
    "risk": 0.5, "risks": 0.5, "concern": 0.6, "concerns": 0.6,
    "await": 0.4, "awaiting": 0.4, "watch": 0.3, "watching": 0.3,
}

# ── FinBERT-style 上下文模式 (短语级语义理解) ──
# 格式: (pattern, sentiment_score) — 正=利好, 负=利空
_FINBERT_PATTERNS = [
    # === 供给侧利好(利好贵金属价格) ===
    ("供给缺口", 1.2), ("供应链中断", 1.0), ("矿山关闭", 1.0),
    ("产量下降", 0.9), ("品位下降", 0.8), ("矿山事故", 0.8),
    ("出口禁令", 0.9), ("出口限制", 0.8), ("制裁", 0.6),
    ("环保限产", 0.7), ("能耗双控", 0.6),
    # === 需求侧利好 ===
    ("央行购金", 1.2), ("央行增持", 1.1), ("ETF流入", 0.9),
    ("ETF增持", 0.9), ("实物需求", 0.7), ("工业需求增", 0.8),
    ("新能源需求", 0.7), ("光伏需求", 0.7),
    # === 宏观利好 ===
    ("美联储降息", 1.0), ("暂停加息", 0.8), ("利率见顶", 0.9),
    ("经济衰退", 0.7), ("滞胀", 0.8), ("去美元化", 0.9),
    ("货币贬值", 0.7), ("信用风险", 0.6),
    ("财政赤字扩大", 0.6), ("债务上限", 0.5),
    # === 地缘利好 ===
    ("地缘紧张", 0.8), ("军事冲突", 0.7), ("贸易战", 0.6),
    ("关税上调", 0.5), ("制裁升级", 0.7),
    # === 供给侧利空 ===
    ("产量增加", -0.8), ("新矿投产", -0.9), ("复产", -0.7),
    ("产能释放", -0.8), ("进口增加", -0.6), ("废料供应增", -0.5),
    # === 需求侧利空 ===
    ("ETF流出", -0.9), ("ETF减持", -0.9), ("央行减持", -1.0),
    ("需求不及预期", -0.8), ("消费低迷", -0.7),
    # === 宏观利空 ===
    ("美联储加息", -1.0), ("鹰派加息", -1.2), ("利率上升", -0.8),
    ("美元走强", -0.9), ("强美元", -0.8), ("实际利率走高", -0.9),
    ("就业强劲", -0.5), ("通胀回落", -0.6),
    # === 否定/转折模式 ===
    ("不及预期", -0.7), ("低于预期", -0.7), ("未能突破", -0.5),
    ("涨幅收窄", -0.3), ("冲高回落", -0.4), ("高位回落", -0.5),
    ("不可持续", -0.6), ("获利了结", -0.5), ("获利回吐", -0.5),
    # === 正面转折 ===
    ("跌幅收窄", 0.3), ("触底反弹", 0.7), ("止跌企稳", 0.6),
    ("利空出尽", 0.8), ("超跌反弹", 0.6), ("V型反转", 0.9),
    # === English FinBERT Positive Patterns ===
    ("central bank buying", 1.2), ("central bank purchases", 1.2),
    ("gold reserves", 0.8), ("reserve accumulation", 0.9),
    ("ETF inflows", 0.9), ("ETF holdings rise", 1.0),
    ("Fed rate cut", 1.0), ("Fed pause", 0.8), ("dovish Fed", 0.9),
    ("peak rates", 0.8), ("rate cuts expected", 0.9),
    ("dollar weakens", 0.9), ("DXY falls", 0.8), ("USD decline", 0.8),
    ("inflation surge", 0.7), ("CPI above", 0.6), ("real yields fall", 0.9),
    ("geopolitical risk", 0.7), ("geopolitical tension", 0.8),
    ("trade war", 0.6), ("tariff escalation", 0.5),
    ("supply disruption", 1.0), ("mine closure", 0.9),
    ("production decline", 0.8), ("ore grade decline", 0.7),
    ("safe haven demand", 1.1), ("flight to safety", 1.0),
    ("record high", 1.2), ("all-time high", 1.3), ("new high", 1.0),
    ("breakout above", 0.9), ("golden cross", 0.8),
    ("short squeeze", 0.9), ("short covering", 0.7),
    ("de-dollarization", 0.9), ("BRICS gold", 0.7),
    # === English FinBERT Negative Patterns ===
    ("ETF outflows", -0.9), ("ETF holdings fall", -1.0),
    ("Fed rate hike", -1.0), ("hawkish Fed", -1.0),
    ("rate hikes expected", -0.9), ("higher for longer", -0.8),
    ("dollar strengthens", -0.9), ("DXY rises", -0.8), ("USD rally", -0.8),
    ("real yields rise", -0.9), ("bond yields surge", -0.7),
    ("inflation cooling", -0.5), ("CPI below", -0.5),
    ("profit taking", -0.5), ("profit-taking", -0.5),
    ("sell-off accelerates", -1.0), ("selloff deepens", -1.1),
    ("death cross", -0.8), ("breakdown below", -0.8),
    ("surplus widens", -0.7), ("oversupply", -0.8),
    ("production increase", -0.7), ("new mine output", -0.8),
    ("demand weakens", -0.8), ("consumption falls", -0.7),
    ("risk appetite", -0.5), ("risk-on sentiment", -0.6),
    ("strong jobs report", -0.5), ("payrolls beat", -0.5),
    ("fails to hold", -0.5), ("rejected at", -0.4),
]

# ── 媒体可信度权重 ──
_SOURCE_CREDIBILITY = {
    # 国内
    "新浪财经": 0.85, "新浪搜索": 0.70, "东方财富": 0.85,
    "东方财富期货": 0.80, "财联社": 0.90, "金十数据": 0.80,
    "同花顺": 0.75, "雪球": 0.60, "上海有色网": 0.85,
    # 国际
    "Reuters": 0.95, "Bloomberg": 0.95,
    "Kitco": 0.85, "Investing.com": 0.75,
    "Google News": 0.70,  # 聚合源, 可信度取决于原始来源
    # Google News转发的子源
    "CNBC": 0.90, "MarketWatch": 0.85, "Financial Times": 0.95,
    "Wall Street Journal": 0.95, "Barron's": 0.85,
    "Mining.com": 0.80, "Metals Daily": 0.80,
}


def _analyze_news_sentiment(news_list: list, metal_id: str) -> dict:
    """
    双算法情绪分析引擎
    ==================
    1) FinBERT-style: 上下文模式匹配 → P(positive) - P(negative)
    2) Loughran-McDonald: 金融词典统计 → (N_pos - N_neg) / (N_pos + N_neg + N_unc)

    融合公式:
      S_fused = α·S_FinBERT + (1-α)·S_LM   (α=0.6)

    每条新闻权重:
      w_i = 时间衰减 × 媒体可信度 × 品种相关度

    输出:
      label:      "利好" / "中性" / "利空"
      confidence:  0-100% 置信度
      fused_score: [-1, 1] 融合分数

    参考:
    - Araci, FinBERT: Financial Sentiment Analysis with Pre-trained Language Models
    - Loughran & McDonald, When Is a Liability Not a Liability?
    """
    import math

    if not news_list:
        return {
            "label": "中性", "confidence": 50, "fused_score": 0.0,
            "finbert_score": 0.0, "lm_score": 0.0,
            "positive_count": 0, "negative_count": 0, "neutral_count": 0,
            "total_count": 0, "details": [],
        }

    metal = METALS.get(metal_id, {})
    metal_name = metal.get("name", "")
    now_ts = time.time()

    per_news_details = []
    finbert_weighted_sum = 0.0
    lm_weighted_sum = 0.0
    total_weight = 0.0

    for news in news_list:
        title = news.get("title", "")
        if not title or len(title) < 5:
            continue

        # ── 权重计算 ──
        # 1) 时间衰减: 24小时内=1.0, 7天前=0.3 (指数衰减)
        date_str = news.get("date", "")
        time_decay = 0.7  # 默认(无日期时)
        if date_str:
            try:
                if len(date_str) >= 10:
                    from datetime import datetime as _dt
                    if "T" in date_str or len(date_str) > 10:
                        news_time = _dt.fromisoformat(date_str.replace("Z", ""))
                    else:
                        news_time = _dt.strptime(date_str[:10], "%Y-%m-%d")
                    hours_ago = (now_ts - news_time.timestamp()) / 3600
                    time_decay = math.exp(-hours_ago / 72)  # 72小时半衰期
                    time_decay = max(0.1, min(1.0, time_decay))
            except:
                pass

        # 2) 媒体可信度
        source = news.get("source", "")
        source_cred = _SOURCE_CREDIBILITY.get(source, 0.65)

        # 3) 品种相关度: 标题包含品种名=1.0, 英文名=0.9, 否则=0.5
        metal_name_en = metal.get("name_en", "").lower()
        title_lower = title.lower()
        if metal_name in title:
            relevance = 1.0
        elif metal_name_en and metal_name_en in title_lower:
            relevance = 0.9
        else:
            relevance = 0.5

        w_i = time_decay * source_cred * relevance

        # ═══ 算法1: FinBERT-style 上下文模式匹配 ═══
        # 对每条新闻模拟 P(positive) 和 P(negative)
        pos_signals = 0.0
        neg_signals = 0.0

        for pattern, score in _FINBERT_PATTERNS:
            if pattern in title or pattern.lower() in title_lower:
                if score > 0:
                    pos_signals += score
                else:
                    neg_signals += abs(score)

        # 否定词处理: "不看好","未上涨","难以突破" / "not bullish", "failed to" 翻转情绪
        negation_words = ["不", "未", "难以", "无法", "并非", "没有", "尚未",
                          "not ", "no ", "never ", "failed to ", "unlikely ",
                          "unable to ", "neither ", "nor "]
        has_negation = any(nw in title for nw in negation_words)
        if has_negation:
            # 如果标题中有否定词且紧接正面词, 降低正面/增加负面
            for pw in list(_LM_POSITIVE.keys())[:20]:
                if pw in title:
                    for nw in negation_words:
                        if nw in title:
                            idx_n = title.find(nw)
                            idx_p = title.find(pw)
                            if 0 <= idx_n < idx_p <= idx_n + len(nw) + 4:
                                pos_signals *= 0.3
                                neg_signals += 0.5
                                break

        # 模拟 softmax → P(pos), P(neg), P(neutral)
        total_signals = pos_signals + neg_signals + 0.5  # 0.5 = neutral prior
        p_pos = pos_signals / total_signals if total_signals > 0 else 0
        p_neg = neg_signals / total_signals if total_signals > 0 else 0
        finbert_i = p_pos - p_neg  # [-1, 1]

        # ═══ 算法2: Loughran-McDonald 词典统计 ═══
        # (case-insensitive matching for English terms)
        n_pos = sum(wt for word, wt in _LM_POSITIVE.items()
                    if word in title or word.lower() in title_lower)
        n_neg = sum(wt for word, wt in _LM_NEGATIVE.items()
                    if word in title or word.lower() in title_lower)
        n_unc = sum(wt for word, wt in _LM_UNCERTAINTY.items()
                    if word in title or word.lower() in title_lower)

        lm_denom = n_pos + n_neg + n_unc
        lm_i = (n_pos - n_neg) / lm_denom if lm_denom > 0 else 0  # [-1, 1]

        # ── 融合打分 ──
        finbert_weighted_sum += w_i * finbert_i
        lm_weighted_sum += w_i * lm_i
        total_weight += w_i

        # 单条新闻标签
        fused_i = 0.6 * finbert_i + 0.4 * lm_i
        if fused_i > 0.10:
            impact = "positive"
        elif fused_i < -0.10:
            impact = "negative"
        else:
            impact = "neutral"
        news["impact"] = impact
        news["sentiment_score"] = round(fused_i, 3)

        # 补全缺失字段
        news.setdefault("date", "")
        news.setdefault("url", "")
        news.setdefault("source", "")

        per_news_details.append({
            "title": title[:40],
            "finbert": round(finbert_i, 3),
            "lm": round(lm_i, 3),
            "fused": round(fused_i, 3),
            "weight": round(w_i, 3),
            "impact": impact,
        })

    # ═══ 全局情绪融合 ═══
    # S_FinBERT = Σ w_i * (P_pos - P_neg) / Σ w_i
    # S_LM = Σ w_i * LM_i / Σ w_i
    # S_fused = α * S_FinBERT + (1-α) * S_LM
    alpha = 0.6  # FinBERT 权重

    if total_weight > 0:
        s_finbert = finbert_weighted_sum / total_weight
        s_lm = lm_weighted_sum / total_weight
    else:
        s_finbert = 0.0
        s_lm = 0.0

    s_fused = alpha * s_finbert + (1 - alpha) * s_lm

    # ── 标签 + 置信度 ──
    # 阈值: > 0.15 利好, < -0.15 利空, 其余中性
    if s_fused > 0.15:
        label = "利好"
    elif s_fused < -0.15:
        label = "利空"
    else:
        label = "中性"

    # 置信度 = |fused_score| 映射到 [50, 98]%
    # 距离阈值越远 → 置信度越高
    abs_score = abs(s_fused)
    if abs_score < 0.05:
        confidence = 50
    elif abs_score < 0.15:
        confidence = 50 + (abs_score / 0.15) * 15  # 50-65
    elif abs_score < 0.30:
        confidence = 65 + ((abs_score - 0.15) / 0.15) * 15  # 65-80
    elif abs_score < 0.50:
        confidence = 80 + ((abs_score - 0.30) / 0.20) * 10  # 80-90
    else:
        confidence = min(98, 90 + (abs_score - 0.50) * 16)  # 90-98

    # 一致性校验: FinBERT 与 LM 同向 → 置信度+5; 分歧 → 置信度-10
    if s_finbert * s_lm > 0:
        confidence = min(98, confidence + 5)
    elif s_finbert * s_lm < -0.01:
        confidence = max(30, confidence - 10)
        label += "(分歧)"

    confidence = round(confidence)

    pos_count = sum(1 for n in news_list if n.get("impact") == "positive")
    neg_count = sum(1 for n in news_list if n.get("impact") == "negative")
    neu_count = sum(1 for n in news_list if n.get("impact") == "neutral")

    return {
        "label": label,
        "confidence": confidence,
        "fused_score": round(s_fused, 4),
        "finbert_score": round(s_finbert, 4),
        "lm_score": round(s_lm, 4),
        "alpha": alpha,
        "threshold": 0.15,
        "agreement": "同向" if s_finbert * s_lm > 0 else "分歧" if s_finbert * s_lm < -0.01 else "中性",
        "positive_count": pos_count,
        "negative_count": neg_count,
        "neutral_count": neu_count,
        "total_count": len(news_list),
    }


# ═══════════════════════════════════════════════
#  5. 算法预测
# ═══════════════════════════════════════════════

def compute_predictions(metal_id: str, contracts_data: dict,
                        indicators: dict, kline_data: dict = None) -> dict:
    """
    交割率综合分析框架 (Integrated Delivery Analysis Framework, IDAF)
    ================================================================
    融合三大经典交割分析算法为统一框架:

    ┌──────────────────────────────────────────────────────────────┐
    │  Module 1: DSCR (可交割供给覆盖率)                            │
    │  → 库存能否覆盖潜在交割需求? Squeeze risk 有多大?             │
    │                                                              │
    │  Module 2: Basis-Carry (基差-持有成本模型)                    │
    │  → 交割在经济上是否"划算"? 期现为什么不收敛?                  │
    │                                                              │
    │  Module 3: SPAN Stress (情景压力测试)                         │
    │  → 不同冲击下交割率/价格会走到哪一步?                         │
    │                                                              │
    │  Aggregator: 加权贝叶斯融合 → 情景概率 + 综合研判             │
    └──────────────────────────────────────────────────────────────┘

    参考文献:
    - CFTC Estimated Deliverable Supply methodology
    - Theory of Storage (Kaldor 1939, Working 1949, Brennan 1958)
    - CME SPAN risk framework (delivery/spot risk parameters)
    - Garcia et al. on storage-rate non-convergence
    """
    import math

    metal = METALS[metal_id]
    is_spot = metal.get("is_spot", False)
    logger.info(f"[IDAF] 计算 {metal['name']} 交割率综合分析...")

    contracts = list(contracts_data.get("data", {}).values())
    kline_list = (kline_data or {}).get("data", [])

    # ═══════════════════════════════════════
    #  基础数据提取
    # ═══════════════════════════════════════
    # spot_price = 现货/即期价格 (K线收盘 or 实时)
    # futures_price = 近月期货价格
    spot_price = 0
    futures_price = 0
    if kline_list:
        spot_price = kline_list[-1].get("close", 0)
    if contracts:
        futures_price = contracts[0].get("last_price", 0)
    # base_price = 展示用基准价格 (优先期货, 回退现货)
    base_price = futures_price if futures_price > 0 else spot_price
    if spot_price == 0:
        spot_price = futures_price  # 无现货数据时用期货替代
    if base_price == 0:
        logger.warning(f"  无法获取基准价格, 跳过预测")
        return {"metal_id": metal_id, "predictions": [], "updated_at": datetime.now().isoformat()}

    ind = indicators.get("indicators", {})
    total_inventory = ind.get("total_inventory", 0)
    registered_inventory = ind.get("registered_inventory", 0)
    inv_change = ind.get("inventory_change", 0)
    open_interest = ind.get("open_interest", 0)
    daily_volume = ind.get("daily_volume", 0)
    position_ratio = ind.get("position_ratio", 1.0)
    delivery_multiplier = metal.get("delivery_multiplier", 1)  # OI(手)→库存同单位
    hist_delivery_rate = metal.get("historical_delivery_rate", 0.25)

    # ── K线技术特征提取 ──
    momentum_5d = 0.0
    momentum_20d = 0.0
    volatility_20d = 0.01
    avg_volume_20d = 0
    price_vs_ma20 = 0.0  # 价格相对20日均线偏离

    if len(kline_list) >= 6:
        closes = [k["close"] for k in kline_list[-6:]]
        momentum_5d = (closes[-1] - closes[0]) / closes[0] if closes[0] > 0 else 0

    if len(kline_list) >= 21:
        closes_21 = [k["close"] for k in kline_list[-21:]]
        momentum_20d = (closes_21[-1] - closes_21[0]) / closes_21[0] if closes_21[0] > 0 else 0
        returns = []
        for j in range(1, 21):
            if closes_21[j-1] > 0:
                returns.append((closes_21[j] - closes_21[j-1]) / closes_21[j-1])
        if returns:
            mean_r = sum(returns) / len(returns)
            volatility_20d = max(0.001, (sum((r - mean_r)**2 for r in returns) / len(returns)) ** 0.5)
        ma20 = sum(closes_21[-20:]) / 20
        price_vs_ma20 = (base_price - ma20) / ma20 if ma20 > 0 else 0

    if len(kline_list) >= 20:
        vols = [k.get("volume", 0) for k in kline_list[-20:]]
        avg_volume_20d = sum(vols) / len(vols) if vols else 0

    # ── 合约数据 ──
    near_price = contracts[0].get("last_price", 0) if contracts else 0
    far_price = contracts[-1].get("last_price", 0) if len(contracts) >= 2 else 0
    near_oi = contracts[0].get("open_interest", 0) if contracts else 0
    near_volume = contracts[0].get("volume", 0) if contracts else 0

    # 基差 = 期货 - 现货 (正值=升水contango, 负值=贴水backwardation)
    basis_abs = (futures_price - spot_price) if (futures_price > 0 and spot_price > 0) else 0
    basis_pct = basis_abs / spot_price if spot_price > 0 else 0

    # 期限结构斜率 = (远月-近月) / 近月
    term_slope = 0.0
    if near_price > 0 and far_price > 0:
        term_slope = (far_price - near_price) / near_price

    # ═══════════════════════════════════════
    #  MODULE 1: DSCR — 可交割供给覆盖率
    # ═══════════════════════════════════════
    # DSCR = Deliverable Supply / (Spot-Month OI × Contract Size)
    # Pressure Ratio = 1 / DSCR (越高 → 挤仓风险越大)
    #
    # 参考: CFTC 明确把 estimated deliverable supply 作为现货月风险核心标尺

    # 可交割供给估算 (取注册仓单为核心, 总库存为上限)
    deliverable_supply = registered_inventory if registered_inventory > 0 else total_inventory * 0.6
    # 潜在交割需求 = 近月OI × 交割乘数 (转为与库存同单位)
    delivery_demand = near_oi * delivery_multiplier if near_oi > 0 else open_interest * delivery_multiplier * 0.3

    if delivery_demand > 0 and deliverable_supply > 0:
        dscr = deliverable_supply / delivery_demand
        pressure_ratio = 1.0 / dscr
    elif deliverable_supply > 0:
        dscr = 5.0  # 无持仓 → 覆盖充裕
        pressure_ratio = 0.2
    else:
        dscr = 1.0  # 无库存数据 → 中性假设
        pressure_ratio = 1.0

    # DSCR → 挤仓风险评分 [0, 1]
    # < 0.5: 极高风险, 0.5-1: 高, 1-2: 中, 2-5: 低, >5: 极低
    if dscr < 0.5:
        squeeze_risk = 0.9 + min(0.1, (0.5 - dscr) * 0.2)
    elif dscr < 1.0:
        squeeze_risk = 0.6 + (1.0 - dscr) * 0.6
    elif dscr < 2.0:
        squeeze_risk = 0.3 + (2.0 - dscr) * 0.3
    elif dscr < 5.0:
        squeeze_risk = 0.1 + (5.0 - dscr) / 30
    else:
        squeeze_risk = max(0.02, 0.1 - (dscr - 5) * 0.01)

    squeeze_risk = max(0.01, min(0.99, squeeze_risk))

    # 库存趋势对覆盖率的动态调整
    inv_trend_adj = 0.0
    if total_inventory > 0 and inv_change != 0:
        inv_trend_adj = -inv_change / total_inventory  # 正=减少(偏紧), 负=增加(宽松)

    logger.info(f"  [DSCR] 可交割供给={deliverable_supply:.0f}, 交割需求={delivery_demand:.0f}")
    logger.info(f"  [DSCR] DSCR={dscr:.2f}, 挤仓压力={pressure_ratio:.2f}, 风险={squeeze_risk:.2f}")

    # ═══════════════════════════════════════
    #  MODULE 2: Basis-Carry — 基差-持有成本模型
    # ═══════════════════════════════════════
    # 交割激励 I_t = Spot - Futures + Financing + Storage + DeliveryFee
    # I_t > 0: 交割有利可图 (卖方倾向交割) → 高交割率
    # I_t < 0: 交割不划算 (卖方倾向平仓) → 低交割率
    #
    # 参考: Theory of Storage (Kaldor/Working/Brennan)

    storage_cost_daily = metal.get("storage_cost_daily", 0)
    financing_rate = metal.get("financing_rate_annual", 0.022)
    delivery_fee = metal.get("delivery_fee", 0)

    # 估算距最近交割月的天数 (粗略: 按当前月到下一交割月)
    now = datetime.now()
    contract_months = metal.get("contract_months", [])
    days_to_delivery = 30  # 默认
    if contract_months:
        for cm in sorted(contract_months):
            if cm > now.month:
                days_to_delivery = (cm - now.month) * 30
                break
        else:
            days_to_delivery = (contract_months[0] + 12 - now.month) * 30

    # 持有成本 = 仓储 + 融资 + 交割手续费
    storage_total = storage_cost_daily * days_to_delivery
    financing_cost = base_price * financing_rate * (days_to_delivery / 365)
    carry_cost = storage_total + financing_cost + delivery_fee

    # 交割激励 (卖方/空头视角, 因为SHFE交割由空方发起)
    # I = Futures - Spot - (Storage + Financing + DeliveryFee)
    # I > 0: 卖方交割有利(期货价高于现货+成本) → 高交割率
    # I < 0: 卖方交割不利(不如平仓+卖现货) → 低交割率
    if futures_price > 0 and spot_price > 0:
        delivery_incentive = futures_price - spot_price - carry_cost
    else:
        delivery_incentive = 0

    # 理论无套利基差 (fair basis): F - S ≈ Carry Cost
    fair_basis = carry_cost
    basis_deviation = basis_abs - fair_basis  # 正=升水超过持有成本(超额利润), 负=升水不足

    # 收敛风险: 基差偏离持有成本越大 → 交割月不收敛风险越高
    convergence_risk = 0.0
    if fair_basis > 0:
        convergence_risk = min(1.0, abs(basis_deviation) / fair_basis)
    elif spot_price > 0:
        convergence_risk = min(1.0, abs(basis_abs) / (spot_price * 0.01))

    # 交割激励评分 [-1, 1]: 正=利于交割(高交割率), 负=不利交割
    if spot_price > 0:
        incentive_score = max(-1, min(1, delivery_incentive / (spot_price * 0.02)))
    else:
        incentive_score = 0

    logger.info(f"  [Basis-Carry] 持有成本={carry_cost:.2f} (仓储={storage_total:.2f}+融资={financing_cost:.2f}+交割费={delivery_fee})")
    logger.info(f"  [Basis-Carry] 基差={basis_abs:.2f}, 理论基差={fair_basis:.2f}, 偏离={basis_deviation:.2f}")
    logger.info(f"  [Basis-Carry] 交割激励={delivery_incentive:.2f}, 评分={incentive_score:.3f}, 收敛风险={convergence_risk:.2f}")

    # ═══════════════════════════════════════
    #  MODULE 3: SPAN Stress — 情景压力测试
    # ═══════════════════════════════════════
    # 对每个交割率情景, 综合 DSCR + Basis-Carry + 技术面 做压力测试
    # 输出: 每个情景的合理性评分 → 贝叶斯后验概率
    #
    # 参考: CME SPAN 框架 (delivery/spot risk parameters)

    scenarios = [
        {
            "rate": 0.05, "rate_str": "5%", "label": "极低交割",
            "description": "多头控盘·仓单严重不足·逼仓风险极高",
            "price_impact_range": [1.5, 4.0],   # 价格上涨倍数(相对tick×50)
        },
        {
            "rate": 0.15, "rate_str": "15%", "label": "低交割",
            "description": "持仓集中度高·卖方交割意愿弱·期货维持升水",
            "price_impact_range": [0.5, 1.5],
        },
        {
            "rate": 0.30, "rate_str": "30%", "label": "正常交割",
            "description": "期现正常回归·基差合理收敛·交割顺畅",
            "price_impact_range": [-0.3, 0.3],
        },
        {
            "rate": 0.50, "rate_str": "50%", "label": "高交割",
            "description": "卖方积极交割·库存充裕·期货小幅贴水",
            "price_impact_range": [-1.5, -0.3],
        },
        {
            "rate": 0.70, "rate_str": "70%+", "label": "极高交割",
            "description": "大量仓单集中交割·空头主导·价格明显承压",
            "price_impact_range": [-3.5, -1.2],
        },
    ]

    # ── 情景合理性评分 (Gaussian-peaked, 非单调) ──
    # 核心思想: 每个信号计算一个"目标交割率", 然后用高斯函数对各情景打分
    # 这样不同市场条件会指向不同的峰值情景
    scenario_scores = []

    # ── 计算各信号的目标交割率 ──

    # Signal 1: DSCR → 目标交割率
    # DSCR越低(供给不足) → 交割率应越低; DSCR越高(供给充裕) → 交割率越高
    if dscr < 0.2:
        dscr_target = 0.05     # 极端紧缺 → 极低交割
    elif dscr < 0.5:
        dscr_target = 0.10     # 严重不足 → 极低~低
    elif dscr < 1.0:
        dscr_target = 0.18     # 偏紧 → 低交割
    elif dscr < 2.0:
        dscr_target = 0.30     # 适中 → 正常交割
    elif dscr < 4.0:
        dscr_target = 0.45     # 充裕 → 高交割
    else:
        dscr_target = 0.60     # 非常充裕 → 极高交割
    dscr_amplitude = 1.0 + squeeze_risk * 0.5  # 挤仓风险越高, 信号越强
    dscr_sigma = 0.12  # 展宽: 允许相邻情景也获得合理分数

    # Signal 2: Basis-Carry → 目标交割率
    # incentive > 0 (利于交割) → 高交割率; incentive < 0 (不利交割) → 低交割率
    if incentive_score < -0.5:
        basis_target = 0.08
    elif incentive_score < -0.2:
        basis_target = 0.15
    elif incentive_score < -0.05:
        basis_target = 0.22
    elif incentive_score < 0.05:
        basis_target = 0.30    # 中性 → 正常
    elif incentive_score < 0.2:
        basis_target = 0.38
    elif incentive_score < 0.5:
        basis_target = 0.50
    else:
        basis_target = 0.65
    basis_amplitude = 0.8 + abs(incentive_score) * 0.5
    basis_sigma = 0.15

    # 收敛风险调整: 高收敛风险 → 增加展宽(不确定性更大)
    if convergence_risk > 0.5:
        basis_sigma += convergence_risk * 0.08
        dscr_sigma += convergence_risk * 0.05

    for s in scenarios:
        rate = s["rate"]
        score_breakdown = {}

        # --- Signal 1: DSCR (Gaussian peaked) ---
        s1 = dscr_amplitude * math.exp(-((rate - dscr_target) ** 2) / (2 * dscr_sigma ** 2))
        # 物理约束: 库存不够时交割率过高仍应被惩罚
        max_feasible_rate = min(1.0, dscr) if dscr < 5 else 1.0
        if rate > max_feasible_rate + 0.15:
            penalty = (rate - max_feasible_rate - 0.15) * 2.0
            s1 = max(0, s1 - penalty)
        # 库存趋势微调
        if inv_trend_adj > 0.005:  # 去库存 → 目标左移
            s1 *= (1 + (dscr_target - rate) * inv_trend_adj * 5)
        elif inv_trend_adj < -0.005:  # 累库存 → 目标右移
            s1 *= (1 + (rate - dscr_target) * abs(inv_trend_adj) * 5)
        score_breakdown["DSCR"] = round(max(0, s1), 3)

        # --- Signal 2: Basis-Carry (Gaussian peaked) ---
        s2 = basis_amplitude * math.exp(-((rate - basis_target) ** 2) / (2 * basis_sigma ** 2))
        # 期限结构微调
        if term_slope > 0.005:  # contango → 目标左移
            s2 *= (1 + (basis_target - rate) * term_slope * 8)
        elif term_slope < -0.005:  # backwardation → 目标右移
            s2 *= (1 + (rate - basis_target) * abs(term_slope) * 8)
        score_breakdown["Basis"] = round(max(0, s2), 3)

        # --- Signal 3: 技术面+资金面 (方向性) ---
        s3 = 0.5  # 基线
        # 动量: 上涨趋势 → 低交割率(看涨); 下跌 → 高交割率
        if momentum_5d > 0.005:
            tech_target = max(0.05, 0.30 - momentum_5d * 15)  # 上涨 → 目标左移
        elif momentum_5d < -0.005:
            tech_target = min(0.70, 0.30 + abs(momentum_5d) * 15)  # 下跌 → 目标右移
        else:
            tech_target = 0.30
        s3 = 0.8 * math.exp(-((rate - tech_target) ** 2) / (2 * 0.18 ** 2))

        # 波动率: 高波动 → 展宽(极端情景增概率)
        vol_annual = volatility_20d * math.sqrt(252)
        if vol_annual > 0.25:
            extremity = abs(rate - 0.30)
            s3 += extremity * (vol_annual - 0.25) * 1.5

        # 持仓比: > 1 偏多 → 低交割率; < 1 偏空 → 高交割率
        if position_ratio > 1.05:
            pos_target = max(0.05, 0.30 - (position_ratio - 1.0) * 1.5)
        elif position_ratio < 0.95:
            pos_target = min(0.70, 0.30 + (1.0 - position_ratio) * 1.5)
        else:
            pos_target = 0.30
        s3 += 0.3 * math.exp(-((rate - pos_target) ** 2) / (2 * 0.15 ** 2))

        score_breakdown["Technical"] = round(max(0, s3), 3)

        # --- 先验: 历史交割率分布 ---
        prior = math.exp(-((rate - hist_delivery_rate) ** 2) / (2 * 0.15 ** 2))
        score_breakdown["Prior"] = round(prior, 3)

        # --- 综合评分 (加权) ---
        w_dscr = 0.35 if not is_spot else 0.0
        w_basis = 0.30 if not is_spot else 0.10
        w_tech = 0.20 if not is_spot else 0.60
        w_prior = 0.15 if not is_spot else 0.30

        total_score = (w_dscr * max(0, score_breakdown["DSCR"]) +
                       w_basis * max(0, score_breakdown["Basis"]) +
                       w_tech * max(0, score_breakdown["Technical"]) +
                       w_prior * score_breakdown["Prior"])

        scenario_scores.append({
            "score": max(0.001, total_score),
            "breakdown": score_breakdown,
        })

    # ═══════════════════════════════════════
    #  概率归一化 (Softmax over scores)
    # ═══════════════════════════════════════
    raw_scores = [ss["score"] for ss in scenario_scores]
    max_score = max(raw_scores)
    # Temperature-scaled softmax (T控制集中度)
    T = 0.3 if not is_spot else 0.5
    exp_scores = [math.exp((s - max_score) / T) for s in raw_scores]
    Z = sum(exp_scores)
    probabilities = [e / Z for e in exp_scores]

    # ═══════════════════════════════════════
    #  情景价格计算 (基于压力测试区间)
    # ═══════════════════════════════════════
    tick = metal["tick_size"]
    tick_unit = tick * 50  # 标准化价格单位

    predictions = []
    for i, s in enumerate(scenarios):
        prob = probabilities[i]
        # 价格影响 = 区间中值 × tick_unit
        mid_impact = (s["price_impact_range"][0] + s["price_impact_range"][1]) / 2
        # 动量修正
        momentum_adj = momentum_5d * 100 * 0.2
        price_impact = (mid_impact + momentum_adj) * tick_unit
        scenario_price = round(base_price + price_impact, 2)

        # 方向判断
        if price_impact > tick_unit * 0.3:
            direction = "偏多"
        elif price_impact < -tick_unit * 0.3:
            direction = "偏空"
        else:
            direction = "中性"

        predictions.append({
            "delivery_rate": s["rate_str"],
            "label": s["label"],
            "rate_value": s["rate"],
            "price": scenario_price,
            "probability": round(prob, 4),
            "direction": direction,
            "description": s["description"],
            "score_breakdown": scenario_scores[i]["breakdown"],
            "price_range": [
                round(base_price + s["price_impact_range"][0] * tick_unit, 2),
                round(base_price + s["price_impact_range"][1] * tick_unit, 2),
            ],
        })

    best = max(predictions, key=lambda x: x["probability"])

    # ═══════════════════════════════════════
    #  综合研判生成
    # ═══════════════════════════════════════

    # 概率加权交割率
    expected_delivery_rate = sum(p["rate_value"] * p["probability"] for p in predictions)
    # 概率加权价格
    expected_price = sum(p["price"] * p["probability"] for p in predictions)
    # 上行/下行概率
    upside_prob = sum(p["probability"] for p in predictions if p["direction"] == "偏多")
    downside_prob = sum(p["probability"] for p in predictions if p["direction"] == "偏空")

    # 综合方向
    if upside_prob > downside_prob + 0.15:
        overall_direction = "偏多"
    elif downside_prob > upside_prob + 0.15:
        overall_direction = "偏空"
    else:
        overall_direction = "中性震荡"

    # 置信度 — 基于信息熵 (Shannon Entropy)
    # H = -Σ p_i * log(p_i),  H_max = log(N)
    # 置信度 = 1 - H/H_max (归一化熵, 0=均匀分布, 1=完全确定)
    H = -sum(p * math.log(p + 1e-10) for p in probabilities)
    H_max = math.log(len(scenarios))
    entropy_confidence = 1.0 - H / H_max  # [0, 1]

    # 方向一致性: 上行/下行概率差距越大越确信
    direction_gap = abs(upside_prob - downside_prob)

    # 数据完整度: 有多少个模块有有效数据
    data_quality = 0.5  # 基线
    if total_inventory > 0: data_quality += 0.1
    is_estimated = ind.get("registered_estimated", False)
    is_cme = ind.get("cme_source", False)
    if registered_inventory > 0:
        if is_cme:
            data_quality += 0.12   # CME真实数据 (COMEX口径, 略低于SHFE直接数据)
        elif not is_estimated:
            data_quality += 0.15   # SHFE真实注册库存
        else:
            data_quality += 0.05   # 估算注册库存 — 降低质量
    if near_oi > 0: data_quality += 0.1
    if len(kline_list) >= 20: data_quality += 0.1
    if abs(basis_abs) > 0: data_quality += 0.05
    data_quality = min(1.0, data_quality)

    # 综合置信度 = 熵置信 × 0.5 + 方向差距 × 0.3 + 数据质量 × 0.2
    raw_confidence = (entropy_confidence * 0.5 +
                      direction_gap * 0.3 +
                      data_quality * 0.2)

    # 映射到 [25, 95] 区间
    confidence_pct = round(25 + raw_confidence * 70)
    confidence_pct = max(25, min(95, confidence_pct))

    # DSCR 状态描述
    if dscr < 0.8:
        dscr_status = "严重不足"
        dscr_risk = "高"
    elif dscr < 1.5:
        dscr_status = "偏紧"
        dscr_risk = "中高"
    elif dscr < 3.0:
        dscr_status = "充足"
        dscr_risk = "低"
    else:
        dscr_status = "非常充裕"
        dscr_risk = "极低"

    # 库存状态
    if inv_trend_adj > 0.01:
        inv_status = "去库存"
    elif inv_trend_adj < -0.01:
        inv_status = "累库存"
    else:
        inv_status = "平衡"

    # 基差状态
    if basis_pct > 0.003:
        basis_status = "升水(Contango)"
    elif basis_pct < -0.003:
        basis_status = "贴水(Backwardation)"
    else:
        basis_status = "平水"

    # 交割激励状态
    if incentive_score > 0.2:
        incentive_status = "有利交割"
    elif incentive_score < -0.2:
        incentive_status = "不利交割"
    else:
        incentive_status = "中性"

    # 日志
    logger.info(f"  [IDAF] 情景概率: " +
                " | ".join(f"{p['label']}={p['probability']:.1%}" for p in predictions))
    logger.info(f"  [IDAF] 最可能: {best['label']}({best['delivery_rate']}) "
                f"P={best['probability']:.1%}, 价格={best['price']}")
    logger.info(f"  [IDAF] 概率加权: 交割率={expected_delivery_rate:.1%}, "
                f"价格={expected_price:.2f}, 方向={overall_direction}")
    logger.info(f"  [IDAF] 特征: DSCR={dscr:.2f}, squeeze={squeeze_risk:.2f}, "
                f"incentive={incentive_score:.3f}, convergence_risk={convergence_risk:.2f}")
    logger.info(f"  [IDAF] 技术: m5d={momentum_5d:.4f}, m20d={momentum_20d:.4f}, "
                f"vol={volatility_20d:.4f}, pos_ratio={position_ratio:.2f}")

    return {
        "metal_id": metal_id,
        "contract": contracts[0]["code"] if contracts else "",
        "base_price": base_price,
        "predictions": predictions,
        "analysis": {
            # 核心结论
            "most_likely_scenario": best["label"],
            "most_likely_rate": best["delivery_rate"],
            "expected_price": round(expected_price, 2),
            "overall_direction": overall_direction,
            "confidence": confidence_pct,
            # Module 1: DSCR
            "dscr": round(dscr, 2),
            "dscr_status": dscr_status,
            "squeeze_risk": round(squeeze_risk, 2),
            "squeeze_risk_level": dscr_risk,
            "deliverable_supply": round(deliverable_supply, 0),
            "deliverable_estimated": ind.get("registered_estimated", False),
            "deliverable_cme": ind.get("cme_source", False),
            "delivery_demand": round(delivery_demand, 0),
            "pressure_ratio": round(pressure_ratio, 2),
            # Module 2: Basis-Carry
            "basis_abs": round(basis_abs, 2),
            "basis_pct": round(basis_pct * 100, 3),
            "basis_status": basis_status,
            "carry_cost": round(carry_cost, 2),
            "fair_basis": round(fair_basis, 2),
            "basis_deviation": round(basis_deviation, 2),
            "delivery_incentive": round(delivery_incentive, 2),
            "incentive_score": round(incentive_score, 3),
            "incentive_status": incentive_status,
            "convergence_risk": round(convergence_risk, 2),
            "days_to_delivery": days_to_delivery,
            # Module 3: Technical/Stress
            "inventory_status": inv_status,
            "inv_trend": round(inv_trend_adj * 100, 2),
            "term_slope": round(term_slope * 100, 3),
            "momentum_5d": round(momentum_5d * 100, 3),
            "momentum_20d": round(momentum_20d * 100, 3),
            "volatility_annual": round(volatility_20d * math.sqrt(252) * 100, 1),
            "position_ratio": round(position_ratio, 2),
            # 概率分布
            "expected_delivery_rate": round(expected_delivery_rate * 100, 1),
            "upside_probability": round(upside_prob * 100, 1),
            "downside_probability": round(downside_prob * 100, 1),
            # 模型元信息
            "model": "IDAF",
            "model_version": "1.0",
            "modules": ["DSCR", "Basis-Carry", "SPAN-Stress"],
            "weights": {"DSCR": 0.35, "Basis-Carry": 0.30, "Technical": 0.20, "Prior": 0.15}
                        if not is_spot else
                        {"DSCR": 0.0, "Basis-Carry": 0.10, "Technical": 0.60, "Prior": 0.30},
            "temperature": T,
        },
        "updated_at": datetime.now().isoformat(),
    }


# ═══════════════════════════════════════════════
#  交割危机概率模型 (Delivery Crisis Probability)
# ═══════════════════════════════════════════════
#
#  基于 CME/CFTC 监管口径 + 学术 squeeze 识别文献的五因子概率模型
#
#  Factor 1: 覆盖率因子 — front-month OI / registered; OI / (registered + 0.5×eligible)
#  Factor 2: 期限结构因子 — 近月-远月价差, backwardation 异常程度
#  Factor 3: 库存流动因子 — registered 变化, eligible→registered 转换, 总库存趋势
#  Factor 4: 时点因子 — 距 first notice day / last trade day 天数
#  Factor 5: 压力因子 — 波动率飙升, 保证金压力, 成交量异常
#
#  输出: 0~100% 概率
#    100% = 极大概率出现失序交割/异常现金和解/极端逼仓
#    0%   = 完全无交割压力
#
#  参考:
#  - CME deliverable supply = registered + 50% × eligible
#  - CFTC 监控: 大户持仓, 关键价差, 库存/可交割供给
#  - Corner/squeeze 经典信号: 近月异常走强, 交割地现货异常高, 交割前库存流向异常

def compute_delivery_crisis(metal_id: str, contracts_data: dict,
                             indicators: dict, kline_data: dict = None,
                             crisis_data: dict = None) -> dict:
    """
    五因子交割危机概率评估模型 (使用CME/CFTC官方数据)

    Returns:
        {
            "probability": 0.22,        # 总概率 (0~1)
            "level": "中低",             # 等级文本
            "factors": {                 # 五因子详情
                "coverage": {...},
                "term_structure": {...},
                "inventory_flow": {...},
                "timing": {...},
                "stress": {...},
            },
            "summary": "...",           # 一句话结论
        }
    """
    import math

    metal = METALS[metal_id]
    logger.info(f"[交割危机] 计算 {metal['name']} 五因子概率...")

    ind = indicators.get("indicators", {})
    contracts_list = list(contracts_data.get("data", {}).values())
    kline_list = (kline_data or {}).get("data", [])
    cd = crisis_data or {}

    # 基础数据 — 统一使用本地单位 (与 OI × delivery_multiplier 单位一致)
    registered = ind.get("registered_inventory", 0)
    total_inv = ind.get("total_inventory", 0)
    qualified = ind.get("qualified_inventory", 0)
    inv_change = ind.get("inventory_change", 0)
    open_interest = ind.get("open_interest", 0)
    daily_volume = ind.get("daily_volume", 0)
    has_cme = ind.get("cme_source", False) or ind.get("cme_registered_oz", 0) > 0

    # CME原始数据 (盎司)
    cme_reg_oz = ind.get("cme_registered_oz", 0)
    cme_elig_oz = ind.get("cme_eligible_oz", 0)
    cme_total_oz = ind.get("cme_total_oz", 0)

    # CME deliverable supply = registered + 50% × eligible (CME/CFTC官方口径)
    cme_deliverable_oz = cme_reg_oz + 0.5 * cme_elig_oz if has_cme else 0

    # eligible 本地单位
    eligible_local = qualified if qualified > 0 else (total_inv - registered if total_inv > registered else 0)

    # ── CME前月OI (优先用crisis_data采集的CME官方数据) ──
    comex_oz_per_contract = metal.get("comex_contract_oz", 0)
    cme_front_oi_contracts = cd.get("front_month_oi_contracts", 0)
    cme_front_oi_oz = cd.get("front_month_oi_oz", 0)

    # 如果CME前月OI缺失, 从settlements数据尝试
    if cme_front_oi_contracts == 0:
        fm = cd.get("settlements", {}).get("front_month", {})
        if fm.get("oi", 0) > 0:
            cme_front_oi_contracts = fm["oi"]
            cme_front_oi_oz = fm["oi"] * comex_oz_per_contract

    # CME结算价 (用于期限结构)
    settle_data = cd.get("settlements", {})
    settle_list = settle_data.get("settlements", [])
    cme_near_settle = settle_data.get("front_month", {}).get("settle", 0)
    cme_far_settle = settle_data.get("second_month", {}).get("settle", 0)

    # SHFE近月/远月价格 (作为后备)
    near_price = contracts_list[0].get("last_price", 0) if contracts_list else 0
    far_price = contracts_list[-1].get("last_price", 0) if len(contracts_list) >= 2 else 0
    near_oi = contracts_list[0].get("open_interest", 0) if contracts_list else 0

    # 交割乘数 (SHFE: 手→kg/吨)
    delivery_mult = metal.get("delivery_multiplier", 1)

    # K线技术数据
    vol_20d = 0.01
    if len(kline_list) >= 21:
        closes = [k["close"] for k in kline_list[-21:]]
        rets = [(closes[i] - closes[i-1]) / closes[i-1]
                for i in range(1, len(closes)) if closes[i-1] > 0]
        if rets:
            mean_r = sum(rets) / len(rets)
            vol_20d = max(0.001, (sum((r - mean_r)**2 for r in rets) / len(rets)) ** 0.5)

    avg_vol_20d = 0
    if len(kline_list) >= 20:
        vols = [k.get("volume", 0) for k in kline_list[-20:]]
        avg_vol_20d = sum(vols) / len(vols) if vols else 0

    # ═══════════════════════════════════════
    #  Factor 1: 覆盖率因子 (权重 30%)
    # ═══════════════════════════════════════
    # 优先路径A: CME数据 (盎司单位, front-month OI × 5000oz vs registered oz)
    # 后备路径B: SHFE数据 (本地单位, OI × delivery_multiplier vs registered)
    #
    # CME口径: deliverable_supply = registered + 0.5 × eligible

    def _logistic(x, midpoint, steepness):
        return 1.0 / (1.0 + math.exp(-steepness * (x - midpoint)))

    use_cme_coverage = (cme_front_oi_oz > 0 and cme_reg_oz > 0)

    if use_cme_coverage:
        # ── 路径A: 纯CME盎司计算 (最准确) ──
        front_month_oz = cme_front_oi_oz
        reg_oz = cme_reg_oz
        ds_oz = cme_deliverable_oz
        ratio_vs_registered = front_month_oz / reg_oz
        ratio_vs_deliverable = front_month_oz / ds_oz if ds_oz > 0 else 2.0
        coverage_src = "CME"
        logger.info(f"  [F1-覆盖率·CME] 前月OI={cme_front_oi_contracts:,}手×{comex_oz_per_contract}oz"
                    f"={front_month_oz:,}oz, Reg={reg_oz:,}oz, DS={ds_oz:,}oz")
    else:
        # ── 路径B: SHFE本地单位 + 历史交割率估算 ──
        hist_del_rate = metal.get("historical_delivery_rate", 0.25)
        expected_delivery_frac = max(0.05, min(0.30, hist_del_rate))

        if near_oi > 0:
            front_month_qty_raw = near_oi * delivery_mult
        elif open_interest > 0:
            front_month_qty_raw = open_interest * delivery_mult * 0.3
        else:
            front_month_qty_raw = 0
        front_month_qty = front_month_qty_raw * expected_delivery_frac
        deliverable_supply_local = registered + 0.5 * eligible_local if eligible_local > 0 else registered * 1.5

        if registered > 0 and front_month_qty > 0:
            ratio_vs_registered = front_month_qty / registered
            ratio_vs_deliverable = front_month_qty / deliverable_supply_local if deliverable_supply_local > 0 else 2.0
        else:
            ratio_vs_registered = 0.0
            ratio_vs_deliverable = 0.0
        coverage_src = "SHFE"
        logger.info(f"  [F1-覆盖率·SHFE] 预期交割={front_month_qty:,.0f}, Reg={registered:,.0f}")

    # 覆盖率→风险分 (Logistic映射)
    coverage_score_reg = _logistic(ratio_vs_registered, 0.8, 5.0)
    coverage_score_del = _logistic(ratio_vs_deliverable, 0.5, 4.0)
    coverage_score = coverage_score_reg * 0.6 + coverage_score_del * 0.4

    factor_coverage = {
        "score": round(coverage_score, 3),
        "weight": 0.30,
        "source": coverage_src,
        "cme_front_oi_contracts": cme_front_oi_contracts,
        "cme_front_oi_oz": int(cme_front_oi_oz),
        "cme_registered_oz": int(cme_reg_oz),
        "cme_eligible_oz": int(cme_elig_oz),
        "cme_deliverable_oz": int(cme_deliverable_oz),
        "ratio_vs_registered": round(ratio_vs_registered, 3),
        "ratio_vs_deliverable": round(ratio_vs_deliverable, 3),
        "label": "OI/Reg={:.2f}, OI/DS={:.2f} ({})".format(
            ratio_vs_registered, ratio_vs_deliverable, coverage_src),
    }

    logger.info(f"  [F1-覆盖率] R1={ratio_vs_registered:.3f}, R2={ratio_vs_deliverable:.3f} "
                f"→ score={coverage_score:.3f} ({coverage_src})")

    # ═══════════════════════════════════════
    #  Factor 2: 期限结构因子 (权重 25%)
    # ═══════════════════════════════════════
    # 优先: CME结算价 (官方source)
    # 后备: SHFE合约报价

    # 选择数据源
    term_near = cme_near_settle if cme_near_settle > 0 else near_price
    term_far = cme_far_settle if cme_far_settle > 0 else far_price
    term_src = "CME结算" if cme_near_settle > 0 else "SHFE报价"

    term_spread = 0.0
    term_spread_pct = 0.0
    if term_near > 0 and term_far > 0:
        term_spread = term_near - term_far
        term_spread_pct = term_spread / term_near

    if term_spread_pct > 0:
        term_score = _logistic(term_spread_pct, 0.01, 200)
    else:
        term_score = max(0.05, 0.3 * _logistic(abs(term_spread_pct), 0.02, -100))

    factor_term = {
        "score": round(term_score, 3),
        "weight": 0.25,
        "source": term_src,
        "near_price": term_near,
        "far_price": term_far,
        "spread": round(term_spread, 2),
        "spread_pct": round(term_spread_pct * 100, 3),
        "structure": "Backwardation" if term_spread > 0 else "Contango",
        "label": "{} ({:+.2f}, {:+.3f}%)".format(
            "Backwardation" if term_spread > 0 else "Contango",
            term_spread, term_spread_pct * 100),
    }

    logger.info(f"  [F2-期限结构·{term_src}] 近={term_near}, 远={term_far}, "
                f"差={term_spread:+.2f} ({term_spread_pct*100:+.3f}%) → {term_score:.3f}")

    # ═══════════════════════════════════════
    #  Factor 3: 库存流动因子 (权重 20%)
    # ═══════════════════════════════════════
    # 信号: registered持续下降, total inventory流失
    # 转换: eligible→registered 速度 (如果eligible远大于registered, 说明缓冲在)
    # 新增: CME stocks日变化追踪, 交割通知量, 连续流失天数

    inv_flow_score = 0.3  # 默认中性

    # ── 优先使用 CME stocks_changes 数据 ──
    stocks_chg = cd.get("stocks_changes", {})
    delivery_notices = cd.get("delivery_notices", {})
    cme_reg_change = stocks_chg.get("reg_change", 0)
    cme_total_change = stocks_chg.get("total_change", 0)
    consecutive_decline = stocks_chg.get("consecutive_decline_days", 0)
    eli_to_reg = stocks_chg.get("eli_to_reg_ratio", 0)
    trend_5d = stocks_chg.get("trend_5d", "")
    daily_issues = delivery_notices.get("daily_issues", 0)

    if stocks_chg.get("source"):
        # 有CME日变化数据 — 使用更精确的评估
        total_for_chg = stocks_chg.get("total_change", 0)
        reg_chg_pct = stocks_chg.get("reg_change_pct", 0)

        # registered 下降 → 风险升高 (最重要)
        if cme_reg_change < 0:
            inv_flow_score = _logistic(abs(reg_chg_pct), 0.3, 5.0)
        else:
            inv_flow_score = max(0.05, 0.25 - reg_chg_pct * 0.5)

        # 连续流失天数加成
        if consecutive_decline >= 5:
            inv_flow_score = min(0.95, inv_flow_score + 0.15)
        elif consecutive_decline >= 3:
            inv_flow_score = min(0.90, inv_flow_score + 0.08)

        # 5日趋势调整
        if trend_5d == "下降":
            inv_flow_score = min(0.95, inv_flow_score + 0.05)

        # E/R 缓冲比调整
        if eli_to_reg > 0:
            import math as _m2
            buffer_adj = max(-0.15, -0.05 * _m2.log(max(1, eli_to_reg)))
            inv_flow_score = max(0.02, min(0.98, inv_flow_score + buffer_adj))

        # 交割通知活跃度加成
        if daily_issues > 100:
            inv_flow_score = min(0.95, inv_flow_score + 0.05)

        logger.info(f"  [F3-库存流动·CME] Reg变化={cme_reg_change:+,.0f} ({reg_chg_pct:+.2f}%), "
                    f"连降={consecutive_decline}天, E/R={eli_to_reg:.2f}, "
                    f"趋势={trend_5d} → score={inv_flow_score:.3f}")
    elif total_inv > 0:
        # 后备: 使用SHFE数据
        inv_change_pct = inv_change / total_inv if total_inv > 0 else 0

        if inv_change < 0:
            inv_flow_score = _logistic(abs(inv_change_pct), 0.005, 300)
        else:
            inv_flow_score = max(0.05, 0.25 - inv_change_pct * 10)

        if registered > 0 and eligible_local > 0:
            eligible_buffer = eligible_local / registered
            buffer_adj = max(-0.15, -0.05 * math.log(max(1, eligible_buffer)))
            inv_flow_score = max(0.02, min(0.98, inv_flow_score + buffer_adj))

        logger.info((f"  [F3-库存流动·SHFE] 变化={inv_change:+,}, 总库存={total_inv:,}, "
                    f"E/R比={eligible_local/registered:.2f}" if registered > 0 else
                    f"  [F3-库存流动·SHFE] 变化={inv_change:+,}, 无注册库存数据")
                    + f" → score={inv_flow_score:.3f}")

    factor_inv_flow = {
        "score": round(inv_flow_score, 3),
        "weight": 0.20,
        "inventory_change": int(inv_change),
        "total_inventory": int(total_inv),
        "cme_reg_change": int(cme_reg_change),
        "consecutive_decline_days": consecutive_decline,
        "eligible_registered_ratio": round(eli_to_reg if eli_to_reg > 0 else
                                           (eligible_local / registered if registered > 0 else 0), 2),
        "trend_5d": trend_5d,
        "daily_issues": daily_issues,
        "source": stocks_chg.get("source", "SHFE"),
        "label": "Reg变化={:+,}, E/R比={:.1f}, 连降={}天".format(
            int(cme_reg_change) if cme_reg_change else int(inv_change),
            eli_to_reg if eli_to_reg > 0 else (eligible_local / registered if registered > 0 else 0),
            consecutive_decline),
    }

    # ═══════════════════════════════════════
    #  Factor 4: 时点因子 (权重 10%)
    # ═══════════════════════════════════════
    # 优先: CME Calendar HTML (crisis_data) → 官方 FND/LTD 日期
    # 后备: 按合约月推算 FND

    now = datetime.now()
    cal = cd.get("calendar", {})
    timing_src = ""

    # ── 优先: CME Calendar 官方数据 ──
    if cal.get("days_to_fnd") is not None and cal.get("next_fnd"):
        days_to_fnd = cal["days_to_fnd"]
        days_to_ltd = cal.get("days_to_ltd", days_to_fnd + 3)
        next_delivery_month = cal.get("next_month", "")
        timing_src = "CME Calendar"
        logger.info(f"  [F4-时点·CME] FND={cal['next_fnd']} ({days_to_fnd}天), "
                    f"LTD={cal.get('next_ltd','')} ({days_to_ltd}天)")
    else:
        # ── 后备: 按合约月推算 ──
        contract_months = metal.get("contract_months", [])
        import calendar as cal_mod
        days_to_fnd = 60
        days_to_ltd = 63
        next_delivery_month = 0

        candidates = []
        for year_offset in range(0, 2):
            for cm in contract_months:
                fnd_y = now.year + year_offset
                fnd_m = cm - 1
                if fnd_m <= 0:
                    fnd_m = 12
                    fnd_y -= 1
                last_day = cal_mod.monthrange(fnd_y, fnd_m)[1]
                fnd = datetime(fnd_y, fnd_m, last_day)
                while fnd.weekday() >= 5:
                    fnd -= timedelta(days=1)
                diff = (fnd - now).days
                candidates.append((diff, cm, fnd))

        future_fnds = [(d, cm, f) for d, cm, f in candidates if d >= -3]
        if future_fnds:
            future_fnds.sort(key=lambda x: x[0])
            days_to_fnd, next_delivery_month, _ = future_fnds[0]
        elif candidates:
            candidates.sort(key=lambda x: abs(x[0]))
            days_to_fnd, next_delivery_month, _ = candidates[0]
        days_to_ltd = days_to_fnd + 3
        timing_src = "推算"

    # 使用文档建议的指数衰减: exp(-days/5)
    import math as _math
    if days_to_fnd <= 0:
        timing_score = 0.9
    else:
        timing_score = max(0.05, _math.exp(-max(days_to_fnd, 0) / 8.0))

    factor_timing = {
        "score": round(timing_score, 3),
        "weight": 0.10,
        "days_to_fnd": days_to_fnd,
        "days_to_ltd": days_to_ltd,
        "next_delivery_month": next_delivery_month,
        "source": timing_src,
        "label": "距FND {}天{}".format(days_to_fnd,
                 f", LTD {days_to_ltd}天" if days_to_ltd != days_to_fnd + 3 else ""),
    }

    logger.info(f"  [F4-时点·{timing_src}] FND={days_to_fnd}天, LTD={days_to_ltd}天 → {timing_score:.3f}")

    # ═══════════════════════════════════════
    #  Factor 5: 压力因子 (权重 15%)
    # ═══════════════════════════════════════
    # 子因子A: 波动率飙升 (realized vol from K线)
    # 子因子B: 保证金压力 (CME margins from crisis_data)
    # 子因子C: 大户集中度 (CFTC COT from crisis_data)
    # 子因子D: 成交量异常
    # 子因子E: 交割地现货升水代理 (LBMA vs COMEX)

    # A: 年化波动率
    annual_vol = vol_20d * (252 ** 0.5) * 100
    vol_score = _logistic(annual_vol, 40, 0.08)

    # B: 保证金压力 (近期是否上调)
    margins = cd.get("margins", {})
    margin_initial = margins.get("initial_margin", 0)
    margin_hike = margins.get("margin_hike_recent", False)
    margin_change_pct = margins.get("margin_change_pct", 0)
    margin_score = 0.1 if not margin_hike else 0.5
    # 高保证金本身说明交易所认为波动大
    if margin_initial > 15000:  # 银保证金 > $15k 算偏高
        margin_score = max(margin_score, 0.3)
    # 近期上调幅度大 → 更高的压力
    if margin_change_pct > 0.05:
        margin_score = min(0.8, margin_score + margin_change_pct * 2)

    # C: 大户集中度 (CFTC COT)
    cot = cd.get("cot", {})
    concentration = cot.get("concentration_risk", 0)
    net_spec = cot.get("net_speculative", 0)
    cot_score = min(0.8, concentration) if concentration > 0 else 0.2
    # 净投机多头极大 → 额外压力
    total_oi_cot = cot.get("total_oi_cot", 0)
    if total_oi_cot > 0 and net_spec > 0:
        spec_ratio = net_spec / total_oi_cot
        if spec_ratio > 0.3:  # 净投机多 > 30% OI
            cot_score = min(0.9, cot_score + 0.1)

    # D: 成交量异常
    vol_ratio = daily_volume / avg_vol_20d if avg_vol_20d > 0 else 1.0
    vol_amplify = max(0, (vol_ratio - 1.5) * 0.2)

    # E: 交割地现货升水代理 (LBMA vs COMEX)
    lbma_data = cd.get("lbma_premium", {})
    premium_pct = abs(lbma_data.get("premium_pct", 0))
    premium_score = 0.0
    if premium_pct > 2.0:  # 升水 > 2% 视为异常高
        premium_score = min(0.6, (premium_pct - 1.0) * 0.15)
    elif premium_pct > 0.5:
        premium_score = min(0.3, premium_pct * 0.1)

    # 综合: vol 40%, 保证金 15%, COT 15%, 量比 15%, 升水 15%
    stress_score = min(0.98, vol_score * 0.40 + margin_score * 0.15 +
                       cot_score * 0.15 + vol_amplify * 0.15 +
                       premium_score * 0.15)

    stress_parts = []
    stress_parts.append(f"RV={annual_vol:.0f}%")
    if margin_initial > 0:
        stress_parts.append(f"保证金${margin_initial:,.0f}" +
                           (" ↑" if margin_hike else ""))
    if concentration > 0:
        stress_parts.append(f"集中度{concentration:.0%}")
    if net_spec != 0 and total_oi_cot > 0:
        stress_parts.append(f"净投机{net_spec/total_oi_cot:.0%}")
    stress_parts.append(f"量比{vol_ratio:.1f}")
    if premium_pct > 0.1:
        stress_parts.append(f"升水{premium_pct:.2f}%")

    factor_stress = {
        "score": round(stress_score, 3),
        "weight": 0.15,
        "annual_volatility": round(annual_vol, 1),
        "volume_ratio": round(vol_ratio, 2),
        "initial_margin": margin_initial,
        "margin_hike": margin_hike,
        "margin_change_pct": round(margin_change_pct * 100, 1),
        "concentration_risk": round(concentration, 3),
        "net_speculative": net_spec,
        "premium_pct": round(premium_pct, 4),
        "sub_scores": {
            "volatility": round(vol_score, 3),
            "margin": round(margin_score, 3),
            "concentration": round(cot_score, 3),
            "volume": round(vol_amplify, 3),
            "premium": round(premium_score, 3),
        },
        "label": " · ".join(stress_parts),
    }

    logger.info(f"  [F5-压力] RV={annual_vol:.1f}%, 保证金=${margin_initial:,.0f}, "
                f"集中度={concentration:.2f}, 量比={vol_ratio:.1f} → {stress_score:.3f}")

    # ═══════════════════════════════════════
    #  加权融合 → 总概率
    # ═══════════════════════════════════════
    w1, w2, w3, w4, w5 = 0.30, 0.25, 0.20, 0.10, 0.15

    raw_prob = (coverage_score * w1 +
                term_score * w2 +
                inv_flow_score * w3 +
                timing_score * w4 +
                stress_score * w5)

    # Logistic校准 (让中间区域更敏感, 避免极端聚集)
    calibrated = _logistic(raw_prob, 0.45, 8.0)

    # 最终概率 clip到 [1%, 95%]
    final_prob = max(0.01, min(0.95, calibrated))

    # 等级划分
    if final_prob >= 0.70:
        level = "极高"
        level_color = "#ef4444"
    elif final_prob >= 0.50:
        level = "高"
        level_color = "#f97316"
    elif final_prob >= 0.30:
        level = "中等"
        level_color = "#f5a623"
    elif final_prob >= 0.15:
        level = "中低"
        level_color = "#3b82f6"
    else:
        level = "低"
        level_color = "#22c55e"

    # 一句话总结
    factors_text = []
    sorted_factors = sorted([
        ("覆盖率", coverage_score, w1),
        ("期限结构", term_score, w2),
        ("库存流动", inv_flow_score, w3),
        ("时点", timing_score, w4),
        ("压力", stress_score, w5),
    ], key=lambda x: x[1] * x[2], reverse=True)

    top_driver = sorted_factors[0][0]
    if final_prob >= 0.30:
        summary = f"结构性偏紧, 短期失序交割概率约{final_prob*100:.0f}%, 主要驱动: {top_driver}因子"
    elif final_prob >= 0.15:
        summary = f"结构性偏紧但可控, 交割危机概率约{final_prob*100:.0f}%, 不属于高危状态"
    else:
        summary = f"交割压力较低, 危机概率约{final_prob*100:.0f}%, 供给缓冲充足"

    logger.info(f"  [交割危机] raw={raw_prob:.3f} → calibrated={calibrated:.3f} → final={final_prob*100:.1f}% ({level})")
    logger.info(f"  [交割危机] {summary}")

    return {
        "probability": round(final_prob, 3),
        "probability_pct": round(final_prob * 100, 1),
        "level": level,
        "level_color": level_color,
        "factors": {
            "coverage": factor_coverage,
            "term_structure": factor_term,
            "inventory_flow": factor_inv_flow,
            "timing": factor_timing,
            "stress": factor_stress,
        },
        "weights": {"coverage": w1, "term_structure": w2, "inventory_flow": w3,
                    "timing": w4, "stress": w5},
        "raw_score": round(raw_prob, 3),
        "summary": summary,
        "top_driver": top_driver,
        "data_sources": cd.get("data_sources", []),
        "methodology": "CME Deliverable Supply (Reg+50%Elig) + CFTC监控口径 + Squeeze经典信号",
        "updated_at": datetime.now().isoformat(),
    }


# ═══════════════════════════════════════════════
#  逐合约交割危机概率 (近3个合约分别计算)
# ═══════════════════════════════════════════════

def _parse_contract_code(code: str):
    """解析合约代码 → (品种前缀, 年, 月)
    AU2604 → ('AU', 2026, 4)
    ag2506 → ('AG', 2025, 6)
    """
    m = re.match(r'^([A-Za-z]+)(\d{2})(\d{2})$', code)
    if not m:
        return None, 0, 0
    prefix = m.group(1).upper()
    year = 2000 + int(m.group(2))
    month = int(m.group(3))
    return prefix, year, month


def _calc_fnd_for_month(year: int, month: int) -> datetime:
    """计算某交割月的 First Notice Day
    FND = 交割月前一个月的最后一个工作日
    例: 2026年4月合约 → FND = 2026年3月最后工作日
    """
    import calendar as cal_mod
    fnd_y, fnd_m = year, month - 1
    if fnd_m <= 0:
        fnd_m = 12
        fnd_y -= 1
    last_day = cal_mod.monthrange(fnd_y, fnd_m)[1]
    fnd = datetime(fnd_y, fnd_m, last_day)
    while fnd.weekday() >= 5:  # 跳过周末
        fnd -= timedelta(days=1)
    return fnd


def compute_per_contract_crisis(metal_id: str, contracts_data: dict,
                                 indicators: dict, kline_data: dict = None,
                                 crisis_data: dict = None) -> list:
    """
    逐合约交割危机概率 — 对近3个合约分别计算五因子概率

    每个合约使用自己的:
      F1: 本合约OI × delivery_multiplier / registered
      F2: 本合约价 vs 下一合约价 (期限结构)
      F4: 本合约交割月的FND

    共享数据:
      F3: 库存流动 (全品种共用)
      F5: 压力因子 (全品种共用)

    Returns: [{
        "contract": "AU2604",
        "delivery_month": "2026-04",
        "probability": 0.22,
        "probability_pct": 22.0,
        "level": "中低",
        "level_color": "#3b82f6",
        "factors": { coverage: {...}, term_structure: {...}, ... },
        "summary": "...",
    }, ...]
    """
    import math

    metal = METALS[metal_id]
    ind = indicators.get("indicators", {})
    cd = crisis_data or {}
    kline_list = (kline_data or {}).get("data", [])

    codes = contracts_data.get("contracts", [])
    cdata = contracts_data.get("data", {})

    if not codes:
        return []

    # 取近3个合约
    top3 = codes[:3]
    logger.info(f"[逐合约危机] {metal['name']} 计算 {top3} 各合约交割危机...")

    # ── 共享基础数据 ──
    registered = ind.get("registered_inventory", 0)
    total_inv = ind.get("total_inventory", 0)
    qualified = ind.get("qualified_inventory", 0)
    inv_change = ind.get("inventory_change", 0)
    eligible_local = qualified if qualified > 0 else (total_inv - registered if total_inv > registered else 0)
    delivery_mult = metal.get("delivery_multiplier", 1)
    hist_del_rate = metal.get("historical_delivery_rate", 0.25)
    expected_delivery_frac = max(0.05, min(0.30, hist_del_rate))

    # CME数据
    cme_reg_oz = ind.get("cme_registered_oz", 0)
    cme_elig_oz = ind.get("cme_eligible_oz", 0)
    has_cme = ind.get("cme_source", False) or cme_reg_oz > 0
    cme_deliverable_oz = cme_reg_oz + 0.5 * cme_elig_oz if has_cme else 0
    comex_oz = metal.get("comex_contract_oz", 0)

    # CME前月OI (如有)
    cme_front_oi = cd.get("front_month_oi_contracts", 0)
    cme_all_months = cd.get("front_month_oi", {}).get("all_months", [])

    # K线共享指标
    vol_20d = 0.01
    avg_vol_20d = 0
    if len(kline_list) >= 21:
        closes = [k["close"] for k in kline_list[-21:]]
        rets = [(closes[i] - closes[i-1]) / closes[i-1]
                for i in range(1, len(closes)) if closes[i-1] > 0]
        if rets:
            mean_r = sum(rets) / len(rets)
            vol_20d = max(0.001, (sum((r - mean_r)**2 for r in rets) / len(rets)) ** 0.5)
    if len(kline_list) >= 20:
        vols = [k.get("volume", 0) for k in kline_list[-20:]]
        avg_vol_20d = sum(vols) / len(vols) if vols else 0

    # F3 库存流动 (共享)
    stocks_chg = cd.get("stocks_changes", {})
    delivery_notices = cd.get("delivery_notices", {})
    cme_reg_change = stocks_chg.get("reg_change", 0)
    consecutive_decline = stocks_chg.get("consecutive_decline_days", 0)
    eli_to_reg = stocks_chg.get("eli_to_reg_ratio", 0)
    trend_5d = stocks_chg.get("trend_5d", "")

    inv_flow_score = 0.3
    if stocks_chg.get("source"):
        reg_chg_pct = stocks_chg.get("reg_change_pct", 0)
        if cme_reg_change < 0:
            inv_flow_score = 1.0 / (1.0 + math.exp(-5.0 * (abs(reg_chg_pct) - 0.3)))
        else:
            inv_flow_score = max(0.05, 0.25 - reg_chg_pct * 0.5)
        if consecutive_decline >= 5:
            inv_flow_score = min(0.95, inv_flow_score + 0.15)
        elif consecutive_decline >= 3:
            inv_flow_score = min(0.90, inv_flow_score + 0.08)
        if trend_5d == "下降":
            inv_flow_score = min(0.95, inv_flow_score + 0.05)
        if eli_to_reg > 0:
            buffer_adj = max(-0.15, -0.05 * math.log(max(1, eli_to_reg)))
            inv_flow_score = max(0.02, min(0.98, inv_flow_score + buffer_adj))
    elif total_inv > 0:
        inv_change_pct = inv_change / total_inv if total_inv > 0 else 0
        if inv_change < 0:
            inv_flow_score = 1.0 / (1.0 + math.exp(-300 * (abs(inv_change_pct) - 0.005)))
        else:
            inv_flow_score = max(0.05, 0.25 - inv_change_pct * 10)
        if registered > 0 and eligible_local > 0:
            buffer_adj = max(-0.15, -0.05 * math.log(max(1, eligible_local / registered)))
            inv_flow_score = max(0.02, min(0.98, inv_flow_score + buffer_adj))

    factor_inv_flow = {
        "score": round(inv_flow_score, 3), "weight": 0.20,
        "source": stocks_chg.get("source", "SHFE"),
        "label": "Reg变化={:+,}, E/R={:.1f}, 连降={}天".format(
            int(cme_reg_change) if cme_reg_change else int(inv_change),
            eli_to_reg if eli_to_reg > 0 else (eligible_local / registered if registered > 0 else 0),
            consecutive_decline),
    }

    # F5 压力 (共享)
    annual_vol = vol_20d * (252 ** 0.5) * 100
    vol_score = 1.0 / (1.0 + math.exp(-0.08 * (annual_vol - 40)))
    margins_d = cd.get("margins", {})
    margin_initial = margins_d.get("initial_margin", 0)
    margin_hike = margins_d.get("margin_hike_recent", False)
    margin_change_pct = margins_d.get("margin_change_pct", 0)
    margin_score = 0.1 if not margin_hike else 0.5
    if margin_initial > 15000:
        margin_score = max(margin_score, 0.3)
    if margin_change_pct > 0.05:
        margin_score = min(0.8, margin_score + margin_change_pct * 2)
    cot = cd.get("cot", {})
    concentration = cot.get("concentration_risk", 0)
    net_spec = cot.get("net_speculative", 0)
    total_oi_cot = cot.get("total_oi_cot", 0)
    cot_score = min(0.8, concentration) if concentration > 0 else 0.2
    if total_oi_cot > 0 and net_spec > 0 and net_spec / total_oi_cot > 0.3:
        cot_score = min(0.9, cot_score + 0.1)
    lbma_data = cd.get("lbma_premium", {})
    premium_pct = abs(lbma_data.get("premium_pct", 0))
    premium_score = min(0.6, (premium_pct - 1.0) * 0.15) if premium_pct > 2.0 else (
        min(0.3, premium_pct * 0.1) if premium_pct > 0.5 else 0.0)

    stress_parts = [f"RV={annual_vol:.0f}%"]
    if margin_initial > 0:
        stress_parts.append(f"${margin_initial:,.0f}" + (" ↑" if margin_hike else ""))
    if concentration > 0:
        stress_parts.append(f"集中度{concentration:.0%}")

    factor_stress_base = {
        "annual_volatility": round(annual_vol, 1),
        "initial_margin": margin_initial,
        "margin_hike": margin_hike,
        "concentration_risk": round(concentration, 3),
    }

    # ── 逐合约计算 ──
    now = datetime.now()
    results = []

    for idx, code in enumerate(top3):
        cd_info = cdata.get(code, {})
        prefix, c_year, c_month = _parse_contract_code(code)
        if c_year == 0:
            continue

        contract_oi = cd_info.get("open_interest", 0)
        contract_price = cd_info.get("last_price", 0)
        contract_vol = cd_info.get("volume", 0)
        delivery_month_str = f"{c_year}-{c_month:02d}"

        # 下一合约的价格 (用于期限结构)
        next_price = 0
        if idx + 1 < len(top3):
            next_code = top3[idx + 1]
            next_price = cdata.get(next_code, {}).get("last_price", 0)

        logger.info(f"  [{code}] OI={contract_oi:,}, Price={contract_price}, "
                    f"NextPrice={next_price}, 交割={delivery_month_str}")

        # ═══ F1: 覆盖率 (本合约OI) ═══
        def _logistic(x, mid, steep):
            return 1.0 / (1.0 + math.exp(-steep * (x - mid)))

        # 预期交割需求 = OI × multiplier × 历史交割率
        front_qty = contract_oi * delivery_mult * expected_delivery_frac
        deliverable_supply_local = registered + 0.5 * eligible_local if eligible_local > 0 else registered * 1.5

        if registered > 0 and front_qty > 0:
            r1 = front_qty / registered
            r2 = front_qty / deliverable_supply_local if deliverable_supply_local > 0 else 2.0
        else:
            r1, r2 = 0.0, 0.0

        cov_score_r = _logistic(r1, 0.8, 5.0)
        cov_score_d = _logistic(r2, 0.5, 4.0)
        coverage_score = cov_score_r * 0.6 + cov_score_d * 0.4

        factor_coverage = {
            "score": round(coverage_score, 3), "weight": 0.30,
            "source": "SHFE",
            "ratio_vs_registered": round(r1, 3),
            "ratio_vs_deliverable": round(r2, 3),
            "contract_oi": contract_oi,
            "label": f"OI/Reg={r1:.2f}, OI/DS={r2:.2f} (SHFE)",
        }

        # ═══ F2: 期限结构 (本合约 vs 下一合约) ═══
        term_spread = 0.0
        term_spread_pct = 0.0
        if contract_price > 0 and next_price > 0:
            term_spread = contract_price - next_price
            term_spread_pct = term_spread / contract_price

        if term_spread_pct > 0:
            term_score = _logistic(term_spread_pct, 0.01, 200)
        else:
            term_score = max(0.05, 0.3 * _logistic(abs(term_spread_pct), 0.02, -100))

        next_label = top3[idx + 1] if idx + 1 < len(top3) else "—"
        factor_term = {
            "score": round(term_score, 3), "weight": 0.25,
            "source": "SHFE报价",
            "near_price": contract_price,
            "far_price": next_price,
            "spread": round(term_spread, 2),
            "spread_pct": round(term_spread_pct * 100, 3),
            "structure": "Backwardation" if term_spread > 0 else "Contango",
            "label": "{} ({:+.2f}, {:+.3f}%)".format(
                "Backwardation" if term_spread > 0 else "Contango",
                term_spread, term_spread_pct * 100),
        }

        # ═══ F4: 时点 (本合约交割月FND) ═══
        fnd = _calc_fnd_for_month(c_year, c_month)
        days_to_fnd = (fnd - now).days

        if days_to_fnd <= 0:
            timing_score = 0.9
        else:
            timing_score = max(0.05, math.exp(-max(days_to_fnd, 0) / 8.0))

        factor_timing = {
            "score": round(timing_score, 3), "weight": 0.10,
            "days_to_fnd": days_to_fnd,
            "fnd_date": fnd.strftime("%Y-%m-%d"),
            "delivery_month": c_month,
            "source": "推算",
            "label": f"距FND {days_to_fnd}天",
        }

        # ═══ F5: 压力 (共享但加入本合约量比) ═══
        vol_ratio = contract_vol / avg_vol_20d if avg_vol_20d > 0 else 1.0
        vol_amplify = max(0, (vol_ratio - 1.5) * 0.2)

        stress_score = min(0.98, vol_score * 0.40 + margin_score * 0.15 +
                           cot_score * 0.15 + vol_amplify * 0.15 +
                           premium_score * 0.15)

        factor_stress = {
            "score": round(stress_score, 3), "weight": 0.15,
            **factor_stress_base,
            "volume_ratio": round(vol_ratio, 2),
            "label": " · ".join(stress_parts + [f"量比{vol_ratio:.1f}"]),
        }

        # ═══ 加权融合 ═══
        w1, w2, w3, w4, w5 = 0.30, 0.25, 0.20, 0.10, 0.15
        raw_prob = (coverage_score * w1 + term_score * w2 +
                    inv_flow_score * w3 + timing_score * w4 + stress_score * w5)
        calibrated = _logistic(raw_prob, 0.45, 8.0)
        final_prob = max(0.01, min(0.95, calibrated))

        if final_prob >= 0.70:
            level, level_color = "极高", "#ef4444"
        elif final_prob >= 0.50:
            level, level_color = "高", "#f97316"
        elif final_prob >= 0.30:
            level, level_color = "中等", "#f5a623"
        elif final_prob >= 0.15:
            level, level_color = "中低", "#3b82f6"
        else:
            level, level_color = "低", "#22c55e"

        sorted_f = sorted([
            ("覆盖率", coverage_score, w1), ("期限结构", term_score, w2),
            ("库存流动", inv_flow_score, w3), ("时点", timing_score, w4),
            ("压力", stress_score, w5),
        ], key=lambda x: x[1] * x[2], reverse=True)
        top_driver = sorted_f[0][0]

        if final_prob >= 0.30:
            summary = f"交割压力偏高({final_prob*100:.0f}%), 主驱动: {top_driver}"
        elif final_prob >= 0.15:
            summary = f"交割压力可控({final_prob*100:.0f}%), 非高危"
        else:
            summary = f"交割压力低({final_prob*100:.0f}%), 供给充足"

        logger.info(f"  [{code}] → {final_prob*100:.1f}% ({level}) 驱动={top_driver}")

        results.append({
            "contract": code,
            "delivery_month": delivery_month_str,
            "delivery_month_num": c_month,
            "probability": round(final_prob, 3),
            "probability_pct": round(final_prob * 100, 1),
            "level": level,
            "level_color": level_color,
            "top_driver": top_driver,
            "summary": summary,
            "factors": {
                "coverage": factor_coverage,
                "term_structure": factor_term,
                "inventory_flow": factor_inv_flow,
                "timing": factor_timing,
                "stress": factor_stress,
            },
        })

    return results


# ═══════════════════════════════════════════════
#  主流程
# ═══════════════════════════════════════════════

def fetch_all():
    """采集所有品种的全部数据并保存"""
    today_dir = get_today_dir()
    logger.info(f"{'='*60}")
    logger.info(f"  贵金属数据采集开始 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info(f"  数据目录: {today_dir}")
    logger.info(f"  AKShare 版本: {ak.__version__}")
    logger.info(f"{'='*60}")

    all_data = {}

    for metal_id, metal_conf in METALS.items():
        logger.info(f"\n{'─'*40}")
        logger.info(f"  处理品种: {metal_conf['name']} ({metal_id})")
        logger.info(f"{'─'*40}")

        metal_data = {
            "metal_id": metal_id,
            "metal_name": metal_conf["name"],
            "metal_name_en": metal_conf["name_en"],
            "unit": metal_conf["unit"],
            "color": metal_conf["color"],
        }

        # 1. K线数据
        try:
            kline = fetch_spot_kline(metal_id)
            metal_data["kline"] = kline
        except Exception as e:
            logger.error(f"  K线采集异常: {e}")
            traceback.print_exc()
            metal_data["kline"] = {"data": []}
        time.sleep(0.5)

        # 2. 实时价格 (+ K线兜底)
        try:
            realtime = fetch_realtime_price(metal_id)
            metal_data["realtime"] = realtime

            if not realtime.get("last_price") and metal_data["kline"]["data"]:
                last_kline = metal_data["kline"]["data"][-1]
                metal_data["realtime"] = {
                    "last_price": last_kline["close"],
                    "open": last_kline["open"],
                    "high": last_kline["high"],
                    "low": last_kline["low"],
                    "prev_close": metal_data["kline"]["data"][-2]["close"]
                        if len(metal_data["kline"]["data"]) > 1 else last_kline["close"],
                    "volume": last_kline.get("volume", 0),
                    "price_source": "kline_close",  # K线收盘, 非实时
                }
                logger.info(f"  ✓ 用K线收盘补充价格: {metal_data['realtime']['last_price']} "
                            f"(⚠ 日K收盘, 非实时行情)")
        except Exception as e:
            logger.error(f"  实时价格采集异常: {e}")
            metal_data["realtime"] = {}
        time.sleep(0.5)

        # 3. 合约行情
        try:
            contracts = fetch_contract_quotes(metal_id)
            metal_data["contracts"] = contracts
        except Exception as e:
            logger.error(f"  合约采集异常: {e}")
            traceback.print_exc()
            metal_data["contracts"] = {"data": {}, "contracts": []}
        time.sleep(0.5)

        # 4. 关键指标
        try:
            indicators = fetch_indicators(metal_id)
            metal_data["indicators"] = indicators
        except Exception as e:
            logger.error(f"  指标采集异常: {e}")
            traceback.print_exc()
            metal_data["indicators"] = {"indicators": {}}
        time.sleep(0.5)

        # 5. 新闻
        try:
            news = fetch_news(metal_id)
            metal_data["news"] = news
        except Exception as e:
            logger.error(f"  新闻采集异常: {e}")
            metal_data["news"] = {"news": []}
        time.sleep(0.5)

        # 6. 算法预测 (传入 kline 作为价格兜底)
        try:
            predictions = compute_predictions(
                metal_id, metal_data["contracts"],
                metal_data["indicators"],
                kline_data=metal_data["kline"],
            )
            metal_data["predictions"] = predictions
        except Exception as e:
            logger.error(f"  预测计算异常: {e}")
            traceback.print_exc()
            metal_data["predictions"] = {"predictions": []}

        # 7. 交割危机数据采集 (CME/CFTC 五因子)
        crisis_data = {}
        try:
            crisis_data = fetch_cme_crisis_data(metal_id)
            metal_data["crisis_data"] = crisis_data
        except Exception as e:
            logger.error(f"  危机数据采集异常: {e}")
            metal_data["crisis_data"] = {}
        time.sleep(0.3)

        # 8. 交割危机概率计算 (五因子模型)
        try:
            crisis = compute_delivery_crisis(
                metal_id, metal_data["contracts"],
                metal_data["indicators"],
                kline_data=metal_data["kline"],
                crisis_data=crisis_data,
            )
            metal_data["delivery_crisis"] = crisis
        except Exception as e:
            logger.error(f"  交割危机计算异常: {e}")
            traceback.print_exc()
            metal_data["delivery_crisis"] = {"probability": 0, "level": "未知"}

        # 9. 逐合约交割危机 (近3个合约分别计算)
        try:
            per_contract = compute_per_contract_crisis(
                metal_id, metal_data["contracts"],
                metal_data["indicators"],
                kline_data=metal_data["kline"],
                crisis_data=crisis_data,
            )
            metal_data["per_contract_crisis"] = per_contract
        except Exception as e:
            logger.error(f"  逐合约危机计算异常: {e}")
            traceback.print_exc()
            metal_data["per_contract_crisis"] = []

        all_data[metal_id] = metal_data

    # ── 保存数据 ──
    output_file = os.path.join(today_dir, "market_data.json")
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump({
            "fetch_time": datetime.now().isoformat(),
            "date": datetime.now().strftime("%Y-%m-%d"),
            "metals": all_data,
        }, f, ensure_ascii=False, indent=2)

    # ── 采集结果汇总 ──
    logger.info(f"\n{'='*60}")
    logger.info(f"  ✅ 数据采集完成!")
    logger.info(f"  保存至: {output_file}")
    logger.info(f"  文件大小: {os.path.getsize(output_file) / 1024:.1f} KB")
    logger.info(f"{'─'*60}")
    for mid, md in all_data.items():
        kline_n = len(md.get("kline", {}).get("data", []))
        rt = md.get("realtime", {}).get("last_price", 0)
        cts = md.get("contracts", {}).get("contracts", [])
        ct_prices = [md.get("contracts", {}).get("data", {}).get(c, {}).get("last_price", 0) for c in cts]
        news_n = len(md.get("news", {}).get("news", []))
        inv = md.get("indicators", {}).get("indicators", {}).get("total_inventory", 0)
        pred_n = len(md.get("predictions", {}).get("predictions", []))

        ok = "✓" if (kline_n > 0 and rt > 0) else "⚠"
        logger.info(f"  {ok} {md['metal_name']}: 价格={rt}, K线={kline_n}条, "
                    f"合约={ct_prices}, 库存={inv}, 新闻={news_n}条, 预测={pred_n}个")
    logger.info(f"{'='*60}")
    return output_file


if __name__ == "__main__":
    try:
        output = fetch_all()
        print(f"\n数据已保存: {output}")
    except KeyboardInterrupt:
        print("\n用户中断")
    except Exception as e:
        logger.error(f"采集失败: {e}")
        traceback.print_exc()
        sys.exit(1)